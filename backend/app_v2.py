import sys
import os
import csv
import sqlite3
import json
import calendar
from email.utils import parsedate_to_datetime
from contextlib import contextmanager
from pathlib import Path
from flask import Flask, jsonify, request, send_file, g
from flask_cors import CORS
from datetime import datetime, timedelta, date
from io import BytesIO
from openpyxl import load_workbook

# Add project root to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from backend.config import Config, BASE_DIR, DATA_DIR, DUCKDB_PATH, RAW_DB_PATH, STORE_GOVERNANCE_DB_PATH, STORE_GOVERNANCE_SOURCE_DB_PATH
from backend.core.db_manager import AuthDBManager, RawDBManager
from backend.core.duckdb_manager import DuckDBManager
from backend.core.funnel_config_backup import FunnelConfigBackup
from backend.core.query_metadata import metadata_registry
from backend.core.query_builder import (
    build_detail_query, build_aggregate_query,
    QueryBuilderError, MAX_PAGE_SIZE, DEFAULT_PAGE_SIZE, MAX_QUERY_ROWS
)
from backend.auth.service import (
    audit_api_response,
    create_role,
    create_user,
    data_scope_types,
    delete_role,
    initialize_auth_system,
    list_audit_logs,
    list_login_logs,
    audit_log_detail,
    login_log_detail,
    list_roles,
    list_users,
    load_current_user,
    login,
    logout,
    permission_tree,
    public_user_payload,
    record_audit_log,
    require_api_access,
    require_permission,
    reset_user_password,
    role_detail,
    set_user_status,
    update_role,
    update_user,
    user_detail,
)

ENABLE_STORE_STATUS_DECORATION = os.getenv("ENABLE_STORE_STATUS_DECORATION", "true").lower() == "true"
ENABLE_STORE_STATUS_EXPORT = os.getenv("ENABLE_STORE_STATUS_EXPORT", "true").lower() == "true"
ENABLE_STORE_STATUS_FILTER = os.getenv("ENABLE_STORE_STATUS_FILTER", "true").lower() == "true"

app = Flask(__name__)
app.config.from_object(Config)
Config.init_app(app)
CORS(app, resources={r"/api/*": {"origins": Config.CORS_ORIGINS}}, supports_credentials=True)

# Initialize managers
raw_db = RawDBManager()
auth_db = AuthDBManager()
duck_db = None


def archive_unreadable_duckdb(error: Exception):
    """Archive a broken DuckDB mart so it can be rebuilt from SQLite."""
    message = str(error).lower()
    if (
        "serialization" not in message
        and "deserialize" not in message
        and "corrupted art index" not in message
        and "same row id was inserted twice" not in message
        and "art index" not in message
        and "fatal error" not in message
        and "database has been invalidated" not in message
        and "invalidated because of a previous fatal error" not in message
    ):
        return False

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    archived = False
    for path in (DUCKDB_PATH, Path(str(DUCKDB_PATH) + ".wal")):
        if path.exists():
            backup_path = path.with_name(f"{path.name}.broken-{timestamp}")
            path.rename(backup_path)
            print(f"Archived unreadable DuckDB file: {path} -> {backup_path}")
            archived = True
    return archived


@contextmanager
def duckdb_maintenance_lock(operation: str = "DuckDB maintenance"):
    import fcntl
    import time

    lock_path = DATA_DIR / '.duckdb_maintenance.lock'
    lock_file = open(lock_path, 'w')
    acquired = False
    try:
        for attempt in range(30):
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                acquired = True
                break
            except BlockingIOError:
                if attempt == 0:
                    print(f"Another worker is {operation.lower()}, waiting...")
                time.sleep(2)
        if not acquired:
            raise RuntimeError(f'Unable to acquire DuckDB maintenance lock for {operation}')
        yield
    finally:
        if acquired:
            try:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
            except Exception:
                pass
        lock_file.close()


def _reset_duck_db():
    global duck_db
    duck_db = DuckDBManager()
    return duck_db


def _raw_sqlite_has_business_tables():
    if not RAW_DB_PATH.exists():
        return False
    try:
        conn = sqlite3.connect(str(RAW_DB_PATH), timeout=30.0)
        try:
            required_tables = ("线索表", "门店表", "跟进表", "人员表")
            for table_name in required_tables:
                row = conn.execute(
                    "SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1",
                    (table_name,),
                ).fetchone()
                if not row:
                    return False
            return True
        finally:
            conn.close()
    except sqlite3.Error as exc:
        print(f"Raw SQLite business table check failed: {exc}")
        return False


def _rebuild_duckdb_from_sqlite():
    """Rebuild DuckDB from raw SQLite, preserving config tables when possible."""
    global duck_db

    config_backup = FunnelConfigBackup(duck_db)
    try:
        config_backup.backup()
    except Exception as exc:
        if archive_unreadable_duckdb(exc):
            print("DuckDB unreadable while backing up config, switching to a fresh rebuild...")
            duck_db.close()
            _reset_duck_db()
            config_backup = FunnelConfigBackup(duck_db)
        else:
            raise

    duck_db.initialize()
    if not _raw_sqlite_has_business_tables():
        print(f"Raw SQLite database has no business tables, initialized empty DuckDB mart: {RAW_DB_PATH}")
        config_backup.restore()
        return
    duck_db.load_from_sqlite()
    config_backup.restore()
    duck_db.compute_all_metrics()


@app.before_request
def authenticate_api_request():
    if request.method == 'OPTIONS':
        return None
    return require_api_access(auth_db)


@app.after_request
def audit_api_request(response):
    return audit_api_response(auth_db, response)


def init_system(force_refresh=False):
    """初始化系统"""
    global duck_db
    print("Initializing Leads Analytics System...")
    initialize_auth_system(auth_db)
    
    duck_db = DuckDBManager()

    with duckdb_maintenance_lock("initializing DuckDB"):
        data_needs_refresh = force_refresh
        metadata_error = None
        if not force_refresh:
            try:
                conn = duck_db.get_connection()
                result = conn.execute("SELECT value FROM metadata WHERE key = 'earliest_data_time' LIMIT 1").fetchone()
                if result and result[0]:
                    print("Data already initialized with full metadata!")
                else:
                    print("Data missing some metadata, refreshing...")
                    data_needs_refresh = True
            except Exception as e:
                metadata_error = e
                print("Data not found or metadata missing, initializing...")
                data_needs_refresh = True

        if data_needs_refresh:
            if metadata_error and archive_unreadable_duckdb(metadata_error):
                duck_db.close()
            _rebuild_duckdb_from_sqlite()
            print("System initialized successfully!")
        else:
            duck_db.ensure_funnel_schema()
            print("Using existing data!")
        duck_db.close()
    return True


@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    payload = request.get_json(silent=True) or {}
    user, error = login(auth_db, payload.get('username'), payload.get('password'))
    if error:
        return jsonify({'success': False, 'message': error}), 401
    record_audit_log(auth_db, '登录认证', '登录', 'user', user['id'])
    return jsonify({'success': True, 'data': public_user_payload(user)})


@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    logout(auth_db)
    return jsonify({'success': True})


@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    user = load_current_user(auth_db)
    if not user:
        return jsonify({'success': False, 'message': '未登录或登录已过期'}), 401
    return jsonify({'success': True, 'data': public_user_payload(user)})


@app.route('/api/admin/permissions', methods=['GET'])
@require_permission('admin.roles.view')
def admin_permissions():
    return jsonify({'success': True, 'data': permission_tree(auth_db)})


@app.route('/api/admin/data-scopes', methods=['GET'])
@require_permission('admin.roles.view')
def admin_data_scopes():
    return jsonify({'success': True, 'data': data_scope_types()})


@app.route('/api/admin/users', methods=['GET'])
@require_permission('admin.users.view')
def admin_users_list():
    users = list_users(
        auth_db,
        keyword=request.args.get('keyword', ''),
        role_id=request.args.get('role_id', ''),
        status=request.args.get('status', ''),
    )
    return jsonify({'success': True, 'data': users})


@app.route('/api/admin/users', methods=['POST'])
@require_permission('admin.users.create')
def admin_users_create():
    payload = request.get_json(silent=True) or {}
    try:
        user_id, actual_password = create_user(auth_db, payload, operator_id=g.current_user['id'])
        record_audit_log(auth_db, '账号管理', '新建账号', 'user', user_id, after_data={k: v for k, v in payload.items() if k != 'password'})
        return jsonify({'success': True, 'data': {'id': user_id, 'temporary_password': actual_password}})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': '登录账号已存在'}), 400
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
@require_permission('admin.users.view')
def admin_users_detail(user_id):
    user = user_detail(auth_db, user_id)
    if not user:
        return jsonify({'success': False, 'message': '账号不存在'}), 404
    return jsonify({'success': True, 'data': user})


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@require_permission('admin.users.edit')
def admin_users_update(user_id):
    payload = request.get_json(silent=True) or {}
    try:
        before = update_user(auth_db, user_id, payload)
        record_audit_log(auth_db, '账号管理', '编辑账号', 'user', user_id, before_data={'username': before['username'], 'display_name': before['display_name']}, after_data=payload)
        return jsonify({'success': True})
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/users/<int:user_id>/status', methods=['PATCH'])
@require_permission('admin.users.status')
def admin_users_status(user_id):
    payload = request.get_json(silent=True) or {}
    try:
        set_user_status(auth_db, user_id, payload.get('status'), g.current_user['id'])
        record_audit_log(auth_db, '账号管理', '启停账号', 'user', user_id, after_data={'status': payload.get('status')})
        return jsonify({'success': True})
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/users/<int:user_id>/reset-password', methods=['POST'])
@require_permission('admin.users.reset_password')
def admin_users_reset_password(user_id):
    payload = request.get_json(silent=True) or {}
    try:
        password = reset_user_password(auth_db, user_id, payload.get('password'))
        record_audit_log(auth_db, '账号管理', '重置密码', 'user', user_id)
        return jsonify({'success': True, 'data': {'temporary_password': password}})
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/roles', methods=['GET'])
@require_permission('admin.roles.view')
def admin_roles_list():
    return jsonify({'success': True, 'data': list_roles(auth_db)})


@app.route('/api/admin/roles', methods=['POST'])
@require_permission('admin.roles.create')
def admin_roles_create():
    payload = request.get_json(silent=True) or {}
    try:
        role_id = create_role(auth_db, payload)
        record_audit_log(auth_db, '角色管理', '新建角色', 'role', role_id, after_data=payload)
        return jsonify({'success': True, 'data': {'id': role_id}})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': '角色编码已存在'}), 400
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/roles/<int:role_id>', methods=['GET'])
@require_permission('admin.roles.view')
def admin_roles_detail(role_id):
    detail = role_detail(auth_db, role_id)
    if not detail:
        return jsonify({'success': False, 'message': '角色不存在'}), 404
    return jsonify({'success': True, 'data': detail})


@app.route('/api/admin/roles/<int:role_id>', methods=['PUT'])
@require_permission('admin.roles.edit')
def admin_roles_update(role_id):
    payload = request.get_json(silent=True) or {}
    try:
        update_role(auth_db, role_id, payload)
        record_audit_log(auth_db, '角色管理', '编辑角色', 'role', role_id, after_data=payload)
        return jsonify({'success': True})
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/roles/<int:role_id>/permissions', methods=['PUT'])
@require_permission('admin.roles.permissions.edit')
def admin_roles_permissions_update(role_id):
    payload = request.get_json(silent=True) or {}
    try:
        update_role(auth_db, role_id, {'permission_codes': payload.get('permission_codes') or []})
        record_audit_log(auth_db, '角色管理', '修改权限', 'role', role_id, after_data=payload)
        return jsonify({'success': True})
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/roles/<int:role_id>', methods=['DELETE'])
@require_permission('admin.roles.delete')
def admin_roles_delete(role_id):
    try:
        delete_role(auth_db, role_id)
        record_audit_log(auth_db, '角色管理', '删除角色', 'role', role_id)
        return jsonify({'success': True})
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/admin/audit-logs', methods=['GET'])
@require_permission('admin.audit_logs.view')
def admin_audit_logs():
    limit = min(int(request.args.get('limit', 100)), 500)
    filters = {
        'operator': request.args.get('operator', ''),
        'module': request.args.get('module', ''),
        'action': request.args.get('action', ''),
        'result': request.args.get('result', ''),
        'start_time': request.args.get('start_time', ''),
        'end_time': request.args.get('end_time', ''),
    }
    return jsonify({'success': True, 'data': list_audit_logs(auth_db, limit=limit, filters=filters)})


@app.route('/api/admin/audit-logs/<int:log_id>', methods=['GET'])
@require_permission('admin.audit_logs.view')
def admin_audit_log_detail(log_id):
    detail = audit_log_detail(auth_db, log_id)
    if not detail:
        return jsonify({'success': False, 'message': '日志不存在'}), 404
    return jsonify({'success': True, 'data': detail})


@app.route('/api/admin/login-logs', methods=['GET'])
@require_permission('admin.login_logs.view')
def admin_login_logs():
    limit = min(int(request.args.get('limit', 100)), 500)
    filters = {
        'username': request.args.get('username', ''),
        'result': request.args.get('result', ''),
        'start_time': request.args.get('start_time', ''),
        'end_time': request.args.get('end_time', ''),
    }
    return jsonify({'success': True, 'data': list_login_logs(auth_db, limit=limit, filters=filters)})


@app.route('/api/admin/login-logs/<int:log_id>', methods=['GET'])
@require_permission('admin.login_logs.view')
def admin_login_log_detail(log_id):
    detail = login_log_detail(auth_db, log_id)
    if not detail:
        return jsonify({'success': False, 'message': '日志不存在'}), 404
    return jsonify({'success': True, 'data': detail})


# ==================== 门店档案与门店治理管理 ====================

DEFAULT_GOVERNANCE_STATUSES = [
    ("正常", "#16a34a", 1, "治理状态"),
    ("观察中", "#2563eb", 2, "治理状态"),
    ("持续异常", "#dc2626", 3, "治理状态"),
    ("已恢复", "#059669", 4, "治理状态"),
    ("需上报", "#d97706", 5, "治理状态"),
]

DEFAULT_STORE_STATUSES = [
    ("正常经营", "#16a34a", 1, "门店状态"),
    ("异常/躺平", "#dc2626", 2, "门店状态"),
    ("退网", "#d97706", 3, "门店状态"),
    ("正常经营(新店）", "#2563eb", 4, "门店状态"),
    ("二网店", "#d97706", 5, "门店状态"),
]

DEFAULT_STORE_RATINGS = [
    ("A级", "#16a34a", 1, "评级"),
    ("B级", "#0891b2", 2, "评级"),
    ("C级", "#ca8a04", 3, "评级"),
    ("D级", "#ea580c", 4, "评级"),
    ("E级", "#dc2626", 5, "评级"),
]


def _dict_rows(cursor):
    return [dict(row) for row in cursor.fetchall()]


def _governance_conn():
    STORE_GOVERNANCE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(STORE_GOVERNANCE_DB_PATH), timeout=30.0)
    conn.row_factory = sqlite3.Row
    _ensure_governance_schema(conn)
    _bootstrap_governance_data(conn)
    _backfill_store_status_from_source(conn)
    return conn


def _ensure_governance_schema(conn):
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 日报快照 (
            id INTEGER PRIMARY KEY AUTOINCREMENT
        )
    """)
    cursor.execute("PRAGMA table_info(日报快照)")
    daily_columns = [row[1] for row in cursor.fetchall()]
    for column_name in (
        "日报数据日期",
        "大区",
        "大区督导",
        "大区经理",
        "战区",
        "战区经理",
        "巡回员",
        "店编号",
        "店简称",
        "30分钟跟进任务数",
        "30分钟及时跟进数",
        "三天三次跟进任务数",
        "三天三次跟进数",
        "线索量-本地",
        "到店数",
        "到店数来源",
        "数值性质",
    ):
        if column_name not in daily_columns:
            cursor.execute(f'ALTER TABLE "日报快照" ADD COLUMN "{column_name}" TEXT')
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 跟进任务 (
            任务ID INTEGER PRIMARY KEY AUTOINCREMENT,
            任务名称 TEXT,
            周开始日期 TEXT,
            门店列表 TEXT,
            状态 TEXT DEFAULT '进行中',
            创建时间 TEXT DEFAULT (datetime('now')),
            完成时间 TEXT,
            完成人 TEXT,
            归档时间 TEXT,
            归档人 TEXT,
            基准日期 TEXT,
            基准数据 TEXT
        )
    """)
    cursor.execute("PRAGMA table_info(跟进任务)")
    task_columns = [row[1] for row in cursor.fetchall()]
    for column_name, column_type in (
        ("任务维度", "TEXT"),
        ("筛选日期", "TEXT"),
        ("筛选条件", "TEXT"),
        ("门店快照", "TEXT"),
        ("总结状态", "TEXT"),
        ("连续上榜快照", "TEXT"),
        ("删除时间", "TEXT"),
        ("删除人", "TEXT"),
        ("删除标记", "INTEGER DEFAULT 0"),
    ):
        if column_name not in task_columns:
            cursor.execute(f'ALTER TABLE 跟进任务 ADD COLUMN "{column_name}" {column_type}')

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 跟进记录 (
            记录ID INTEGER PRIMARY KEY AUTOINCREMENT,
            任务ID INTEGER,
            日报数据日期 TEXT,
            店编号 TEXT,
            店简称 TEXT,
            "线索量-本地" TEXT,
            到店数 TEXT,
            跟进原因 TEXT,
            备注 TEXT,
            操作人 TEXT,
            创建时间 TEXT DEFAULT (datetime('now')),
            跟进时间 TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 跟进原因配置 (
            配置ID INTEGER PRIMARY KEY AUTOINCREMENT,
            原因选项 TEXT,
            排序 INTEGER,
            状态 TEXT,
            父级ID INTEGER,
            创建时间 TEXT DEFAULT (datetime('now'))
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 任务历史记录表 (
            历史ID INTEGER PRIMARY KEY AUTOINCREMENT,
            任务ID INTEGER,
            操作类型 TEXT,
            操作前状态 TEXT,
            操作后状态 TEXT,
            操作人 TEXT,
            操作时间 TEXT DEFAULT (datetime('now')),
            操作备注 TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 任务总结表 (
            总结ID INTEGER PRIMARY KEY AUTOINCREMENT,
            任务ID INTEGER NOT NULL,
            总结类型 TEXT,
            总结来源 TEXT,
            总结内容 TEXT,
            草稿内容 TEXT,
            确认状态 TEXT DEFAULT '待确认',
            确认人 TEXT,
            确认时间 TEXT,
            创建时间 TEXT DEFAULT (datetime('now')),
            更新时间 TEXT DEFAULT (datetime('now'))
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 门店管理配置表 (
            管理ID INTEGER PRIMARY KEY AUTOINCREMENT,
            店编号 TEXT UNIQUE NOT NULL,
            门店状态 TEXT DEFAULT '',
            状态备注 TEXT,
            管理员备注 TEXT,
            创建时间 TEXT DEFAULT (datetime('now')),
            更新时间 TEXT DEFAULT (datetime('now'))
        )
    """)
    cursor.execute("PRAGMA table_info(门店管理配置表)")
    config_columns = [row[1] for row in cursor.fetchall()]
    if "门店评级" not in config_columns:
        cursor.execute("ALTER TABLE 门店管理配置表 ADD COLUMN 门店评级 TEXT")
    if "治理状态" not in config_columns:
        cursor.execute("ALTER TABLE 门店管理配置表 ADD COLUMN 治理状态 TEXT")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS 门店状态配置表 (
            状态ID INTEGER PRIMARY KEY AUTOINCREMENT,
            状态名称 TEXT UNIQUE NOT NULL,
            状态颜色 TEXT,
            排序 INTEGER DEFAULT 0,
            创建时间 TEXT DEFAULT (datetime('now'))
        )
    """)
    cursor.execute("PRAGMA table_info(门店状态配置表)")
    status_columns = [row[1] for row in cursor.fetchall()]
    if "配置类型" not in status_columns:
        cursor.execute("ALTER TABLE 门店状态配置表 ADD COLUMN 配置类型 TEXT DEFAULT '状态'")
    if "启用状态" not in status_columns:
        cursor.execute("ALTER TABLE 门店状态配置表 ADD COLUMN 启用状态 INTEGER DEFAULT 1")
    if "说明" not in status_columns:
        cursor.execute("ALTER TABLE 门店状态配置表 ADD COLUMN 说明 TEXT")

    for row in DEFAULT_STORE_STATUSES + DEFAULT_GOVERNANCE_STATUSES + DEFAULT_STORE_RATINGS:
        cursor.execute("""
            INSERT OR IGNORE INTO 门店状态配置表 (状态名称, 状态颜色, 排序, 配置类型)
            VALUES (?, ?, ?, ?)
        """, row)

    default_reason_groups = [
        ("线索量问题", 1),
        ("门店配合度", 2),
        ("到店率问题", 3),
        ("其他原因", 4),
    ]
    default_reasons = [
        ("线索量正常，无需干预", "线索量问题", 1),
        ("线索量下降，已联系督导", "线索量问题", 2),
        ("线索量下降，已联系店长", "线索量问题", 3),
        ("线索量下降，等待自然恢复", "线索量问题", 4),
        ("门店积极配合，等待数据回升", "门店配合度", 5),
        ("门店不配合，已上报", "门店配合度", 6),
        ("到店率正常", "到店率问题", 7),
        ("到店率偏低，已提醒门店", "到店率问题", 8),
        ("到店率偏低，门店已采取措施", "到店率问题", 9),
        ("其他原因", "其他原因", 10),
    ]
    cursor.execute("DROP INDEX IF EXISTS idx_follow_reason_unique")
    cursor.execute("""
        DELETE FROM 跟进原因配置
        WHERE COALESCE(父级ID, -1) != 0
          AND 配置ID NOT IN (
              SELECT MIN(配置ID)
              FROM 跟进原因配置
              WHERE 原因选项 IS NOT NULL
                AND 原因选项 != ''
                AND COALESCE(父级ID, -1) != 0
              GROUP BY 原因选项
          )
    """)
    cursor.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_follow_reason_unique_active
        ON 跟进原因配置 (原因选项)
        WHERE COALESCE(父级ID, -1) != 0
    """)
    group_ids = {}
    for group_name, sort_order in default_reason_groups:
        existing_group = cursor.execute("""
            SELECT 配置ID
            FROM 跟进原因配置
            WHERE 原因选项 = ? AND COALESCE(父级ID, -1) = 0
            LIMIT 1
        """, (group_name,)).fetchone()
        if existing_group:
            group_ids[group_name] = existing_group[0]
        else:
            cursor.execute("""
                INSERT INTO 跟进原因配置 (原因选项, 排序, 状态, 父级ID)
                VALUES (?, ?, '启用', 0)
            """, (group_name, sort_order))
            group_ids[group_name] = cursor.lastrowid
    for reason, group_name, sort_order in default_reasons:
        existing = cursor.execute("""
            SELECT 配置ID
            FROM 跟进原因配置
            WHERE 原因选项 = ? AND COALESCE(父级ID, -1) != 0
            LIMIT 1
        """, (reason,)).fetchone()
        if not existing:
            cursor.execute("""
                INSERT INTO 跟进原因配置 (原因选项, 排序, 状态, 父级ID)
                VALUES (?, ?, '启用', ?)
            """, (reason, sort_order, group_ids.get(group_name)))
        elif group_ids.get(group_name):
            cursor.execute("""
                UPDATE 跟进原因配置
                SET 父级ID = ?
                WHERE 配置ID = ? AND 父级ID IS NULL
            """, (group_ids[group_name], existing[0]))
    conn.commit()


def _table_columns(conn, table_name):
    return [row[1] for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()]


def _copy_table_common_columns(source_conn, target_conn, table_name):
    if not _table_exists(source_conn, table_name):
        return 0

    source_columns = _table_columns(source_conn, table_name)
    target_columns = _table_columns(target_conn, table_name)
    common_columns = [col for col in target_columns if col in source_columns]
    if not common_columns:
        return 0

    quoted = ", ".join([f'"{col}"' for col in common_columns])
    rows = source_conn.execute(f'SELECT {quoted} FROM "{table_name}"').fetchall()
    if not rows:
        return 0

    placeholders = ", ".join(["?"] * len(common_columns))
    target_conn.executemany(
        f'INSERT OR IGNORE INTO "{table_name}" ({quoted}) VALUES ({placeholders})',
        [tuple(row[col] for col in common_columns) for row in rows],
    )
    return len(rows)


def _bootstrap_governance_data(conn):
    """首次初始化时，从异常店治理工具同步既有治理数据到当前系统治理库。"""
    if not STORE_GOVERNANCE_SOURCE_DB_PATH.exists():
        return
    if STORE_GOVERNANCE_SOURCE_DB_PATH.resolve() == STORE_GOVERNANCE_DB_PATH.resolve():
        return

    current_count = conn.execute("SELECT COUNT(*) FROM 门店管理配置表").fetchone()[0]
    daily_count = 0
    if _table_exists(conn, "日报快照"):
        daily_count = conn.execute("SELECT COUNT(*) FROM 日报快照").fetchone()[0]
    follow_count = 0
    if _table_exists(conn, "跟进记录"):
        follow_count = conn.execute("SELECT COUNT(*) FROM 跟进记录").fetchone()[0]
    if current_count > 0 or daily_count > 0 or follow_count > 0:
        return

    source_conn = sqlite3.connect(str(STORE_GOVERNANCE_SOURCE_DB_PATH), timeout=30.0)
    source_conn.row_factory = sqlite3.Row
    try:
        if not _table_exists(source_conn, "门店管理配置表") and not _table_exists(source_conn, "日报快照"):
            return

        for table_sql in (
            """
            CREATE TABLE IF NOT EXISTS 日报快照 (
                id INTEGER PRIMARY KEY AUTOINCREMENT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS 跟进任务 (
                任务ID INTEGER PRIMARY KEY AUTOINCREMENT,
                任务名称 TEXT,
                周开始日期 TEXT,
                门店列表 TEXT,
                状态 TEXT,
                创建时间 TEXT,
                完成时间 TEXT,
                完成人 TEXT,
                归档时间 TEXT,
                归档人 TEXT,
                基准日期 TEXT,
                基准数据 TEXT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS 跟进记录 (
                记录ID INTEGER PRIMARY KEY AUTOINCREMENT,
                任务ID INTEGER,
                日报数据日期 TEXT,
                店编号 TEXT,
                店简称 TEXT,
                "线索量-本地" TEXT,
                到店数 TEXT,
                跟进原因 TEXT,
                备注 TEXT,
                操作人 TEXT,
                创建时间 TEXT,
                跟进时间 TEXT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS 跟进原因配置 (
                配置ID INTEGER PRIMARY KEY AUTOINCREMENT,
                原因选项 TEXT,
                排序 INTEGER,
                状态 TEXT,
                父级ID INTEGER,
                创建时间 TEXT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS 任务历史记录表 (
                历史ID INTEGER PRIMARY KEY AUTOINCREMENT,
                任务ID INTEGER,
                操作类型 TEXT,
                操作前状态 TEXT,
                操作后状态 TEXT,
                操作人 TEXT,
                操作时间 TEXT,
                操作备注 TEXT
            )
            """,
        ):
            conn.execute(table_sql)

        for table_name in ("日报快照", "跟进任务", "跟进记录", "跟进原因配置", "任务历史记录表", "门店管理配置表", "门店状态配置表"):
            if not _table_exists(source_conn, table_name):
                continue
            for col in _table_columns(source_conn, table_name):
                if col not in _table_columns(conn, table_name):
                    try:
                        conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{col}" TEXT')
                    except sqlite3.OperationalError:
                        pass
            _copy_table_common_columns(source_conn, conn, table_name)

        conn.commit()
        print(f"Store governance data initialized from {STORE_GOVERNANCE_SOURCE_DB_PATH}")
    finally:
        source_conn.close()


def _backfill_store_status_from_source(conn):
    """Initialize system-managed store status for stores not yet maintained."""
    stores = _raw_store_rows()
    if not stores:
        return
    now = _now_text()
    existing_statuses = {
        str(row["店编号"]): (row["门店状态"] or "").strip()
        for row in conn.execute("SELECT 店编号, 门店状态 FROM 门店管理配置表").fetchall()
    }
    changed = False
    for store in stores:
        dealer_id = str(store.get("店编号") or "").strip()
        source_status = (store.get("门店状态") or "").strip()
        if not dealer_id or not source_status:
            continue
        if dealer_id not in existing_statuses:
            conn.execute("""
                INSERT INTO 门店管理配置表 (店编号, 门店状态, 更新时间)
                VALUES (?, ?, ?)
            """, (dealer_id, source_status, now))
            changed = True
        elif not existing_statuses[dealer_id] or existing_statuses[dealer_id] == "正常":
            conn.execute("""
                UPDATE 门店管理配置表
                SET 门店状态 = ?, 更新时间 = ?
                WHERE 店编号 = ?
            """, (source_status, now, dealer_id))
            changed = True
    if changed:
        conn.commit()


def _table_exists(conn, table_name):
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        (table_name,),
    ).fetchone()
    return bool(row)


def _normalize_date(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    value = str(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        return parsedate_to_datetime(value).strftime("%Y-%m-%d")
    except (TypeError, ValueError, IndexError):
        pass
    parts = value.split("/")
    if len(parts) == 3:
        try:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
        except ValueError:
            return value
    return value


def _raw_store_rows():
    if not RAW_DB_PATH.exists():
        return []
    conn = sqlite3.connect(str(RAW_DB_PATH), timeout=30.0)
    conn.row_factory = sqlite3.Row
    try:
        if not _table_exists(conn, "门店表"):
            return []
        rows = conn.execute("""
            SELECT 大区, 大区督导, 大区经理, 大区副经理,
                   战区, 战区经理, 巡回员, 店编号, 店简称,
                   商贸重点店, 非商贸重点店, 门店状态,
                   商贸店, 线索运营区域负责人, "线索运营-区域支持"
            FROM 门店表
            ORDER BY 大区, 战区, 店编号
        """).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def _store_base_map():
    return {str(row.get("店编号") or ""): row for row in _raw_store_rows() if row.get("店编号")}


def _governance_config_map(conn):
    rows = conn.execute("""
        SELECT 店编号, 门店状态, 治理状态, 门店评级, 状态备注, 管理员备注, 创建时间, 更新时间
        FROM 门店管理配置表
    """).fetchall()
    return {str(row["店编号"]): dict(row) for row in rows}


def _safe_float(value):
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _current_operator_name():
    user = getattr(g, "current_user", None) or {}
    return user.get("display_name") or user.get("username") or "系统"


def _now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _json_dumps(value):
    return json.dumps(value, ensure_ascii=False, default=str)


def _parse_json_object(value, fallback=None):
    fallback = fallback if fallback is not None else {}
    if not value:
        return fallback
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else fallback
    except Exception:
        return fallback


def _parse_json_list(value, fallback=None):
    fallback = fallback if fallback is not None else []
    if not value:
        return fallback
    if isinstance(value, list):
        return value
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else fallback
    except Exception:
        return fallback


def _add_task_history(conn, task_id, action, before_status="", after_status="", remark=""):
    conn.execute("""
        INSERT INTO 任务历史记录表 (任务ID, 操作类型, 操作前状态, 操作后状态, 操作人, 操作时间, 操作备注)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (task_id, action, before_status or "", after_status or "", _current_operator_name(), _now_text(), remark or ""))


def _shift_date_str(value, days=0):
    if not value:
        return ""
    try:
        dt = datetime.strptime(_normalize_date(value), "%Y-%m-%d")
        return (dt + timedelta(days=days)).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _same_day_last_month(value):
    if not value:
        return ""
    try:
        dt = datetime.strptime(_normalize_date(value), "%Y-%m-%d")
        year = dt.year
        month = dt.month - 1
        if month <= 0:
            month += 12
            year -= 1
        last_day = calendar.monthrange(year, month)[1]
        day = min(dt.day, last_day)
        return dt.replace(year=year, month=month, day=day).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _week_start_str(value):
    if not value:
        return ""
    try:
        dt = datetime.strptime(_normalize_date(value), "%Y-%m-%d")
        return (dt - timedelta(days=dt.weekday())).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _dealer_daily_connection():
    global duck_db
    if duck_db is None:
        duck_db = DuckDBManager()
    return duck_db.get_connection()


def _dealer_monthly_report_rows(region="", zone="", dealer_id="", dealer_name="", page=1, page_size=2000, report_date=""):
    daily_conn = _dealer_daily_connection()
    end_date = _normalize_date(report_date) or _latest_report_date(daily_conn)
    if not end_date:
        return []
    month_start = f"{end_date[:7]}-01"
    response = _get_dealer_report_custom_range(
        daily_conn,
        month_start,
        end_date,
        region,
        zone,
        dealer_id,
        dealer_name,
        "lead_count",
        "desc",
        page,
        page_size,
    )
    payload = response.get_json(silent=True) or {}
    return payload.get("data") or []


def _dealer_monthly_report_map(report_date=""):
    target_date = _normalize_date(report_date) or _latest_report_date(_dealer_daily_connection())
    cache_key = f"_dealer_monthly_report_map_{target_date or 'latest'}"
    cached = getattr(g, cache_key, None)
    if cached is not None:
        return cached
    rows = _dealer_monthly_report_rows(page=1, page_size=5000, report_date=target_date)
    result = {str(row.get("dealer_id") or ""): row for row in rows if row.get("dealer_id")}
    setattr(g, cache_key, result)
    return result


def _dealer_range_report_rows(start_date, end_date, page=1, page_size=5000):
    start_date = _normalize_date(start_date)
    end_date = _normalize_date(end_date)
    if not start_date or not end_date:
        return []
    daily_conn = _dealer_daily_connection()
    response = _get_dealer_report_custom_range(
        daily_conn,
        start_date,
        end_date,
        "",
        "",
        "",
        "",
        "lead_count",
        "desc",
        page,
        page_size,
    )
    payload = response.get_json(silent=True) or {}
    return payload.get("data") or []


def _dealer_range_report_map(start_date, end_date):
    start_date = _normalize_date(start_date)
    end_date = _normalize_date(end_date)
    cache_key = f"_dealer_range_report_map_{start_date}_{end_date}"
    cached = getattr(g, cache_key, None)
    if cached is not None:
        return cached
    rows = _dealer_range_report_rows(start_date, end_date)
    result = {str(row.get("dealer_id") or ""): row for row in rows if row.get("dealer_id")}
    setattr(g, cache_key, result)
    return result


def _range_visit_metric_for_store(store_code, start_date, end_date):
    row = _dealer_range_report_map(start_date, end_date).get(str(store_code or ""))
    if not row:
        return {"visit_count": 0, "local_lead_count": 0, "visit_rate": 0}
    local_lead_count = _safe_float(row.get("local_lead_count"))
    visit_count = _safe_float(row.get("to_shop_count"))
    return {
        "visit_count": visit_count,
        "local_lead_count": local_lead_count,
        "visit_rate": round(visit_count * 100 / local_lead_count, 2) if local_lead_count > 0 else 0,
    }


def _daily_metric_for_store(conn, store_code, report_date):
    if not report_date:
        return {"report_date": "", "local_lead_count": 0, "visit_count": 0, "visit_rate": 0}
    target_date = _normalize_date(report_date)
    try:
        row = _dealer_monthly_report_map(target_date).get(store_code)
    except Exception:
        row = None
    if not row:
        return {"report_date": target_date, "local_lead_count": 0, "visit_count": 0, "visit_rate": 0}
    local_lead_count = _safe_float(row.get("local_lead_count"))
    visit_count = _safe_float(row.get("to_shop_count"))
    return {
        "report_date": target_date,
        "local_lead_count": local_lead_count,
        "visit_count": visit_count,
        "visit_rate": round(visit_count * 100 / local_lead_count, 2) if local_lead_count > 0 else 0,
    }


def _dealer_daily_report_row(store_code, report_date):
    target_date = _normalize_date(report_date)
    if not store_code or not target_date:
        return None
    daily_conn = _dealer_daily_connection()
    try:
        row = daily_conn.execute("""
            SELECT
                report_date,
                dealer_id,
                COALESCE(d_local_lead_count, 0) AS local_lead_count,
                COALESCE(d_to_shop_count, 0) AS to_shop_count,
                COALESCE(d_local_lead_to_shop_rate, 0) AS local_lead_to_shop_rate
            FROM report_dealer_daily
            WHERE period_type = 'daily'
              AND dealer_id = ?
              AND report_date = CAST(? AS DATE)
            LIMIT 1
        """, (store_code, target_date)).fetchone()
    except Exception:
        return None
    if not row:
        return None
    return {
        "report_date": target_date,
        "local_lead_count": _safe_float(row[2]),
        "visit_count": _safe_float(row[3]),
        "visit_rate": _safe_float(row[4]),
    }


def _daily_visit_metric_for_store(store_code, report_date):
    row = _dealer_daily_report_row(store_code, report_date)
    if row:
        return row
    target_date = _normalize_date(report_date)
    return {"report_date": target_date, "local_lead_count": 0, "visit_count": 0, "visit_rate": 0}


def _available_report_dates(conn):
    daily_conn = _dealer_daily_connection()
    try:
        rows = daily_conn.execute("""
            SELECT DISTINCT report_date
            FROM report_dealer_daily
            WHERE period_type = 'daily'
            ORDER BY report_date
        """).fetchall()
    except Exception:
        return []
    dates = [_normalize_date(row[0]) for row in rows if _normalize_date(row[0])]
    return dates


def _latest_report_date(conn):
    dates = _available_report_dates(conn)
    if dates:
        return dates[-1]
    return ""


def _monthly_summary_for_store(store_code, report_date=""):
    target_date = _normalize_date(report_date) or _latest_report_date(_dealer_daily_connection())
    if not store_code or not target_date:
        return {
            "report_date": "",
            "month_start": "",
            "lead_count": 0,
            "visit_count": 0,
            "visit_rate": 0,
        }
    month_start = f"{target_date[:7]}-01"
    row = _dealer_range_report_map(month_start, target_date).get(str(store_code or ""))
    if not row:
        return {
            "report_date": target_date,
            "month_start": month_start,
            "lead_count": 0,
            "visit_count": 0,
            "visit_rate": 0,
        }
    lead_count = _safe_float(row.get("lead_count"))
    visit_count = _safe_float(row.get("to_shop_count"))
    return {
        "report_date": target_date,
        "month_start": month_start,
        "lead_count": lead_count,
        "visit_count": visit_count,
        "visit_rate": round(visit_count * 100 / lead_count, 2) if lead_count > 0 else 0,
    }


def _daily_stats_for_store(conn, store_code, start_date="", end_date=""):
    start_date = _normalize_date(start_date)
    end_date = _normalize_date(end_date)
    if not end_date:
        end_date = _latest_report_date(_dealer_daily_connection())
    if not start_date and end_date:
        start_date = (datetime.strptime(end_date, "%Y-%m-%d") - timedelta(days=29)).strftime("%Y-%m-%d")
    if not store_code or not start_date or not end_date:
        return []
    daily_conn = _dealer_daily_connection()
    try:
        rows = daily_conn.execute("""
            WITH lead_daily AS (
                SELECT
                    assign_date AS report_date,
                    COUNT(*) AS lead_count,
                    SUM(CASE WHEN lead_status != '异地' THEN 1 ELSE 0 END) AS local_lead_count
                FROM mart_leads
                WHERE channel_1 = '线上'
                  AND dealer_id = ?
                  AND assign_date >= CAST(? AS DATE)
                  AND assign_date <= CAST(? AS DATE)
                GROUP BY assign_date
            ),
            visit_daily AS (
                SELECT
                    visit_date AS report_date,
                    SUM(unique_lead_count) AS visit_count
                FROM fact_daily_visit
                WHERE period_type = 'daily'
                  AND channel_1 = '线上'
                  AND dealer_id = ?
                  AND visit_date >= CAST(? AS DATE)
                  AND visit_date <= CAST(? AS DATE)
                GROUP BY visit_date
            ),
            dates AS (
                SELECT report_date FROM lead_daily
                UNION
                SELECT report_date FROM visit_daily
            )
            SELECT
                d.report_date,
                COALESCE(l.lead_count, 0) AS lead_count,
                COALESCE(l.local_lead_count, 0) AS local_lead_count,
                COALESCE(v.visit_count, 0) AS visit_count
            FROM dates d
            LEFT JOIN lead_daily l ON d.report_date = l.report_date
            LEFT JOIN visit_daily v ON d.report_date = v.report_date
            ORDER BY d.report_date
        """, (store_code, start_date, end_date, store_code, start_date, end_date)).fetchall()
    except Exception:
        return []
    data = []
    for row in rows:
        day = _normalize_date(row[0])
        lead_count = _safe_float(row[1])
        local_lead_count = _safe_float(row[2])
        visit_count = _safe_float(row[3])
        data.append({
            "report_date": day,
            "raw_report_date": day,
            "lead_count": lead_count,
            "local_lead_count": local_lead_count,
            "visit_count": visit_count,
            "visit_rate": round(visit_count * 100 / lead_count, 2) if lead_count > 0 else 0,
        })
    data.sort(key=lambda item: item["report_date"])
    return data


def _follow_stats_map(conn):
    if not _table_exists(conn, "跟进记录"):
        return {}
    rows = conn.execute("""
        SELECT 店编号,
               COUNT(*) AS follow_count,
               MIN(COALESCE(跟进时间, 创建时间)) AS first_follow_time,
               MAX(COALESCE(跟进时间, 创建时间)) AS latest_follow_time,
               GROUP_CONCAT(DISTINCT 跟进原因) AS reason_summary
        FROM 跟进记录
        GROUP BY 店编号
    """).fetchall()
    return {str(row["店编号"]): dict(row) for row in rows}


def _store_profile_payload(store, config=None, follow=None, trend=None):
    config = config or {}
    follow = follow or {}
    trend = trend or {}
    source_store_status = store.get("门店状态") or ""
    managed_store_status = config.get("门店状态") or ""
    return {
        "store_code": store.get("店编号") or "",
        "store_name": store.get("店简称") or "",
        "region": store.get("大区") or "",
        "zone": store.get("战区") or "",
        "region_supervisor": store.get("大区督导") or "",
        "region_manager": store.get("大区经理") or "",
        "region_deputy_manager": store.get("大区副经理") or "",
        "zone_manager": store.get("战区经理") or "",
        "inspector": store.get("巡回员") or "",
        "source_store_status": source_store_status,
        "store_status": managed_store_status,
        "governance_status": config.get("治理状态") or "",
        "store_rating": config.get("门店评级") or "",
        "status_note": config.get("状态备注") or "",
        "admin_note": config.get("管理员备注") or "",
        "updated_at": config.get("更新时间") or "",
        "follow_count": int(follow.get("follow_count") or 0),
        "first_follow_time": follow.get("first_follow_time") or "",
        "latest_follow_time": follow.get("latest_follow_time") or "",
        "reason_summary": follow.get("reason_summary") or "",
        "avg_local_lead_count": trend.get("avg_local_lead_count", 0),
        "avg_visit_count": trend.get("avg_visit_count", 0),
        "avg_visit_rate": trend.get("avg_visit_rate", 0),
    }


def _split_follow_reasons(value):
    text = str(value or "").strip()
    if not text:
        return []
    normalized = text
    for sep in ("，", ",", ";", "；", "\n", "\r", "|", "/"):
        normalized = normalized.replace(sep, "、")
    return [item.strip() for item in normalized.split("、") if item.strip()]


def _reason_category_map(conn):
    if not _table_exists(conn, "跟进原因配置"):
        return {}
    rows = _dict_rows(conn.execute("""
        SELECT r.原因选项 AS reason, p.原因选项 AS category
        FROM 跟进原因配置 r
        LEFT JOIN 跟进原因配置 p ON r.父级ID = p.配置ID
        WHERE COALESCE(r.父级ID, -1) != 0
          AND r.原因选项 IS NOT NULL
          AND r.原因选项 != ''
    """))
    return {str(row.get("reason") or "").strip(): (row.get("category") or "未分类") for row in rows if row.get("reason")}


def _week_key_for_follow_record(row):
    raw_time = row.get("follow_time") or row.get("created_at") or row.get("report_date") or ""
    normalized = _normalize_date(raw_time)
    try:
        dt = datetime.strptime(normalized, "%Y-%m-%d")
    except Exception:
        return "", "", ""
    week_start = dt - timedelta(days=dt.weekday())
    week_end = week_start + timedelta(days=6)
    return week_start.strftime("%Y-%m-%d"), week_end.strftime("%Y-%m-%d"), f"{week_start.strftime('%Y-%m-%d')} 至 {week_end.strftime('%Y-%m-%d')}"


def _status_options(conn, config_type):
    rows = conn.execute("""
        SELECT 状态ID AS id, 状态名称 AS name, 状态颜色 AS color, 排序 AS sort_order,
               配置类型 AS config_type, COALESCE(启用状态, 1) AS enabled, COALESCE(说明, '') AS description
        FROM 门店状态配置表
        WHERE 配置类型 = ?
        ORDER BY 排序, 状态ID
    """, (config_type,)).fetchall()
    return [dict(row) for row in rows]


def _store_status_map(dealer_ids=None, enabled=None):
    if enabled is None:
        enabled = ENABLE_STORE_STATUS_DECORATION
    if not enabled:
        return {}
    target_ids = {str(item) for item in dealer_ids or [] if item}
    status_map = {}
    with _governance_conn() as conn:
        configs = _governance_config_map(conn)
    for dealer_id, config in configs.items():
        if dealer_id and (not target_ids or dealer_id in target_ids):
            status_map[dealer_id] = config.get("门店状态") or ""
    return status_map


def _decorate_rows_with_store_status(rows, dealer_key="dealer_id"):
    if not ENABLE_STORE_STATUS_DECORATION or not rows:
        return rows
    dealer_ids = [row.get(dealer_key) for row in rows if isinstance(row, dict)]
    status_map = _store_status_map(dealer_ids)
    for row in rows:
        if isinstance(row, dict):
            row["store_status"] = status_map.get(str(row.get(dealer_key) or ""), "")
    return rows


def _xlsx_rows_with_store_status(results, dealer_index, insert_index):
    if not ENABLE_STORE_STATUS_EXPORT:
        return [list(row) for row in results]
    dealer_ids = [row[dealer_index] for row in results if len(row) > dealer_index]
    status_map = _store_status_map(dealer_ids, enabled=True)
    output = []
    for row in results:
        row_data = list(row)
        dealer_id = str(row_data[dealer_index] or "") if len(row_data) > dealer_index else ""
        row_data.insert(insert_index, status_map.get(dealer_id, ""))
        output.append(row_data)
    return output


def _xlsx_headers_with_store_status(headers, insert_index):
    output = list(headers)
    if ENABLE_STORE_STATUS_EXPORT:
        output.insert(insert_index, "门店状态")
    return output


def _store_status_filter_options():
    if not ENABLE_STORE_STATUS_FILTER:
        return []
    try:
        with _governance_conn() as conn:
            return [row.get("name") for row in _status_options(conn, "门店状态") if row.get("enabled") != 0 and row.get("name")]
    except Exception:
        return []


def _filter_rows_by_store_status(rows, store_status, dealer_key="dealer_id"):
    if not ENABLE_STORE_STATUS_FILTER or not store_status:
        return rows
    return [row for row in rows if str(row.get("store_status") or "") == store_status]


def _store_status_filter_ids(store_status):
    if not ENABLE_STORE_STATUS_FILTER or not store_status:
        return None
    status_map = _store_status_map(enabled=True)
    return {dealer_id for dealer_id, value in status_map.items() if value == store_status}


def _store_status_filtered_tuple_rows(results, store_status, dealer_index):
    dealer_ids = _store_status_filter_ids(store_status)
    if dealer_ids is None:
        return results
    return [row for row in results if len(row) > dealer_index and str(row[dealer_index] or "") in dealer_ids]


def _dealer_id_filter_sql(alias, dealer_ids):
    if dealer_ids is None:
        return "", []
    if not dealer_ids:
        return f" AND {alias}.dealer_id IS NULL", []
    placeholders = ", ".join(["?"] * len(dealer_ids))
    return f" AND CAST({alias}.dealer_id AS VARCHAR) IN ({placeholders})", sorted(dealer_ids)


def _parse_task_store_list(value):
    if not value:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value).strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
    except Exception:
        pass
    return [item.strip() for item in text.replace("，", ",").split(",") if item.strip()]


def _unique_store_codes(stores):
    result = []
    seen = set()
    for item in stores or []:
        if isinstance(item, dict):
            code = str(item.get("store_code") or item.get("店编号") or "").strip()
        else:
            code = str(item or "").strip()
        if not code or code in seen:
            continue
        seen.add(code)
        result.append(code)
    return result


def _parse_multi_value_terms(value):
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    else:
        text = str(value or "")
        for sep in ("，", ",", ";", "；", "、", "|", "/", "\n", "\r", "\t"):
            text = text.replace(sep, " ")
        raw_items = text.split(" ")
    seen = set()
    result = []
    for item in raw_items:
        text = str(item or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _task_dimension(task_row):
    explicit = task_row.get("任务维度") if isinstance(task_row, dict) else None
    if explicit:
        return explicit
    name = str(task_row.get("任务名称") or "")
    if "日" in name and "周" not in name:
        return "日任务"
    if "周" in name:
        return "周任务"
    if task_row.get("基准日期") or task_row.get("周开始日期"):
        return "周任务"
    return "日任务"


def _task_store_streak_map(conn, store_code, current_task_row=None, include_current_task=True):
    if current_task_row:
        snapshot = _parse_json_object(current_task_row.get("连续上榜快照") or "")
        if isinstance(snapshot, dict) and store_code in snapshot:
            item = snapshot.get(store_code) or {}
            return int(item.get("continuous_days") or 0), item.get("continuous_tasks") or []
    rows = _dict_rows(conn.execute("""
        SELECT 任务ID, 任务名称, 周开始日期, 基准日期, 创建时间, 状态, 门店列表
        FROM 跟进任务
        WHERE COALESCE(删除标记, 0) = 0
        ORDER BY 任务ID DESC
    """))
    current_task_id = int(current_task_row.get("任务ID") or 0) if current_task_row else None
    streak = 0
    streak_tasks = []
    for row in rows:
        if current_task_id is not None:
            row_task_id = int(row.get("任务ID") or 0)
            if row_task_id > current_task_id:
                continue
            if row_task_id == current_task_id and not include_current_task:
                continue
        store_list = set(_parse_task_store_list(row.get("门店列表")))
        if store_code in store_list:
            streak += 1
            streak_tasks.append({
                "task_id": row.get("任务ID"),
                "task_name": row.get("任务名称") or "",
                "status": row.get("状态") or "",
                "week_start_date": row.get("周开始日期") or "",
                "baseline_date": row.get("基准日期") or "",
                "created_at": row.get("创建时间") or "",
            })
        else:
            break
    return streak, streak_tasks


def _build_continuous_snapshot(conn, store_codes, current_task_id=None):
    snapshot = {}
    task_context = {"任务ID": current_task_id} if current_task_id else None
    for code in store_codes:
        streak, streak_tasks = _task_store_streak_map(conn, code, task_context, include_current_task=False)
        snapshot[code] = {
            "continuous_days": int(streak or 0) + 1,
            "continuous_tasks": streak_tasks[:10],
        }
    return snapshot


def _task_store_streak_from_filter(conn, store_code, report_date, filters):
    """按当前人工筛选条件回看连续日报，计算连续上榜天数。"""
    if not report_date:
        return 0
    try:
        current = datetime.strptime(_normalize_date(report_date), "%Y-%m-%d")
    except Exception:
        return 0
    streak = 0
    for offset in range(0, 60):
        day = (current - timedelta(days=offset)).strftime("%Y-%m-%d")
        metric = _build_store_metric_snapshot(conn, store_code, day)
        if not _store_matches_filter(metric, filters):
            break
        streak += 1
    return streak


def _decline_rate(current_value, previous_value):
    if previous_value <= 0:
        return None
    return round((previous_value - current_value) * 100 / previous_value, 2)


def _build_store_metric_snapshot(conn, store_code, report_date):
    base_store = _store_base_map().get(store_code, {})
    current = _daily_metric_for_store(conn, store_code, report_date)
    daily_current = _daily_visit_metric_for_store(store_code, report_date)
    week_start = _week_start_str(report_date)
    prev_week_start = _shift_date_str(week_start, -7)
    prev_week_end = _shift_date_str(report_date, -7)
    month_start = f"{_normalize_date(report_date)[:7]}-01" if _normalize_date(report_date) else ""
    prev_month_start = _same_day_last_month(month_start)
    prev_month_end = _same_day_last_month(report_date)
    current_week = _range_visit_metric_for_store(store_code, week_start, report_date)
    prev_week = _range_visit_metric_for_store(store_code, prev_week_start, prev_week_end)
    current_month = _range_visit_metric_for_store(store_code, month_start, report_date)
    prev_month = _range_visit_metric_for_store(store_code, prev_month_start, prev_month_end)
    wow_diff = current_week["visit_count"] - prev_week["visit_count"]
    mom_diff = current_month["visit_count"] - prev_month["visit_count"]
    wow_rate = round(wow_diff * 100 / prev_week["visit_count"], 2) if prev_week["visit_count"] > 0 else None
    mom_rate = round(mom_diff * 100 / prev_month["visit_count"], 2) if prev_month["visit_count"] > 0 else None
    return {
        "store_code": store_code,
        "store_name": base_store.get("店简称") or "",
        "region": base_store.get("大区") or "",
        "zone": base_store.get("战区") or "",
        "source_store_status": base_store.get("门店状态") or "",
        "report_date": current.get("report_date") or _normalize_date(report_date),
        "local_lead_count": current.get("local_lead_count") or 0,
        "visit_count": current.get("visit_count") or 0,
        "visit_rate": daily_current.get("visit_rate") or 0,
        "prev_week_visit_count": prev_week.get("visit_count") or 0,
        "prev_month_visit_count": prev_month.get("visit_count") or 0,
        "wow_visit_diff": wow_diff,
        "wow_visit_rate": wow_rate,
        "mom_visit_diff": mom_diff,
        "mom_visit_rate": mom_rate,
        "wow_decline_rate": _decline_rate(current_week.get("visit_count") or 0, prev_week.get("visit_count") or 0),
        "mom_decline_rate": _decline_rate(current_month.get("visit_count") or 0, prev_month.get("visit_count") or 0),
    }


def _store_matches_filter(metric, filters):
    min_leads = filters.get("min_local_lead_count")
    min_visits = filters.get("min_visit_count")
    max_visit_rate = filters.get("max_visit_rate")
    min_wow_decline = filters.get("min_wow_decline_rate")
    min_mom_decline = filters.get("min_mom_decline_rate")

    if min_leads not in (None, "") and metric["local_lead_count"] < _safe_float(min_leads):
        return False
    if min_visits not in (None, "") and metric["visit_count"] >= _safe_float(min_visits):
        return False
    if max_visit_rate not in (None, "") and metric["visit_rate"] > _safe_float(max_visit_rate):
        return False
    if min_wow_decline not in (None, ""):
        decline = metric.get("wow_decline_rate")
        if decline is None or decline < _safe_float(min_wow_decline):
            return False
    if min_mom_decline not in (None, ""):
        decline = metric.get("mom_decline_rate")
        if decline is None or decline < _safe_float(min_mom_decline):
            return False
    return True


def _preview_governance_stores(conn, payload):
    report_date = _normalize_date(payload.get("report_date") or payload.get("baseline_date") or _latest_report_date(conn))
    if not report_date:
        return [], "", "未找到可用日报数据"

    filters = payload.get("filters") or {}
    selected_codes = {str(code).strip().lower() for code in _parse_multi_value_terms(payload.get("store_codes"))}
    selected_names = [str(name).strip().lower() for name in _parse_multi_value_terms(payload.get("store_names"))]
    keyword = (payload.get("keyword") or "").strip().lower()
    region = payload.get("region") or ""
    zone = payload.get("zone") or ""
    limit = min(max(int(payload.get("limit") or 200), 1), 1000)
    metric_map = _dealer_monthly_report_map(report_date)
    rows = []
    for store in _raw_store_rows():
        code = str(store.get("店编号") or "")
        name = str(store.get("店简称") or "")
        code_lower = code.lower()
        name_lower = name.lower()
        if selected_codes and code_lower not in selected_codes:
            continue
        if selected_names and not any(term in name_lower for term in selected_names):
            continue
        if region and store.get("大区") != region:
            continue
        if zone and store.get("战区") != zone:
            continue
        if keyword and keyword not in code_lower and keyword not in name_lower:
            continue
        if code not in metric_map:
            continue
        metric = _build_store_metric_snapshot(conn, code, report_date)
        if not _store_matches_filter(metric, filters):
            continue
        metric.pop("continuous_days", None)
        rows.append(metric)

    sort_by = payload.get("sort_by") or "visit_count"
    sort_order = payload.get("sort_order") or "desc"
    rows.sort(key=lambda item: item.get(sort_by) if item.get(sort_by) is not None else -999999, reverse=sort_order != "asc")
    return rows[:limit], report_date, ""


def _stores_with_continuous_days(conn, stores):
    result = []
    for store in stores:
        item = dict(store)
        store_code = str(item.get("store_code") or "").strip()
        if store_code:
            streak, _ = _task_store_streak_map(conn, store_code)
            item["continuous_days"] = streak + 1
        else:
            item["continuous_days"] = 1
        result.append(item)
    return result


def _task_history_map(conn):
    if not _table_exists(conn, "任务历史记录表"):
        return {}
    rows = _dict_rows(conn.execute("""
        SELECT 任务ID, 操作类型, 操作前状态, 操作后状态, 操作人, 操作时间, 操作备注
        FROM 任务历史记录表
        ORDER BY datetime(COALESCE(操作时间, '1970-01-01')) DESC, 历史ID DESC
    """))
    history = {}
    for row in rows:
        task_id = str(row.get("任务ID") or "")
        history.setdefault(task_id, []).append(row)
    return history


def _task_store_payload(conn, store_code, task_row):
    base_store = _store_base_map().get(store_code, {})
    config = _governance_config_map(conn).get(store_code, {})
    follow = _follow_stats_map(conn).get(store_code, {})
    baseline_date = _normalize_date(task_row.get("基准日期") or task_row.get("周开始日期") or task_row.get("创建时间") or "")
    latest_daily = _build_store_metric_snapshot(conn, store_code, baseline_date)
    streak, streak_tasks = _task_store_streak_map(conn, store_code, task_row)
    return {
        "store_code": store_code,
        "store_name": latest_daily.get("store_name") or base_store.get("店简称") or "",
        "region": latest_daily.get("region") or base_store.get("大区") or "",
        "zone": latest_daily.get("zone") or base_store.get("战区") or "",
        "source_store_status": latest_daily.get("source_store_status") or base_store.get("门店状态") or "",
        "governance_status": config.get("治理状态") or "",
        "store_rating": config.get("门店评级") or "",
        "status_note": config.get("状态备注") or "",
        "admin_note": config.get("管理员备注") or "",
        "follow_count": int(follow.get("follow_count") or 0),
        "latest_follow_time": follow.get("latest_follow_time") or "",
        "reason_summary": follow.get("reason_summary") or "",
        "report_date": latest_daily.get("report_date") or "",
        "local_lead_count": latest_daily.get("local_lead_count") or 0,
        "visit_count": latest_daily.get("visit_count") or 0,
        "visit_rate": latest_daily.get("visit_rate") or 0,
        "prev_week_visit_count": latest_daily.get("prev_week_visit_count") or 0,
        "prev_month_visit_count": latest_daily.get("prev_month_visit_count") or 0,
        "wow_visit_diff": latest_daily.get("wow_visit_diff") or 0,
        "wow_visit_rate": latest_daily.get("wow_visit_rate"),
        "mom_visit_diff": latest_daily.get("mom_visit_diff") or 0,
        "mom_visit_rate": latest_daily.get("mom_visit_rate"),
        "continuous_days": streak,
        "continuous_tasks": streak_tasks[:5],
    }


def _task_row_payload(conn, task_row):
    task_id = task_row.get("任务ID")
    store_list = _parse_task_store_list(task_row.get("门店列表"))
    follow_count = conn.execute("""
        SELECT COUNT(DISTINCT NULLIF(TRIM(店编号), ''))
        FROM 跟进记录
        WHERE 任务ID = ?
    """, (task_id,)).fetchone()[0] if _table_exists(conn, "跟进记录") else 0
    histories = _task_history_map(conn).get(str(task_id), [])
    return {
        "task_id": task_id,
        "task_name": task_row.get("任务名称") or f"{task_row.get('周开始日期') or task_id} 跟进任务",
        "dimension": _task_dimension(task_row),
        "status": task_row.get("状态") or "",
        "week_start_date": task_row.get("周开始日期") or "",
        "baseline_date": task_row.get("基准日期") or "",
        "filter_date": task_row.get("筛选日期") or "",
        "filter_config": _parse_json_object(task_row.get("筛选条件") or ""),
        "summary_status": task_row.get("总结状态") or "",
        "created_at": task_row.get("创建时间") or "",
        "completed_at": task_row.get("完成时间") or "",
        "completed_by": task_row.get("完成人") or "",
        "archived_at": task_row.get("归档时间") or "",
        "archived_by": task_row.get("归档人") or "",
        "store_count": len(store_list),
        "follow_count": int(follow_count or 0),
        "store_preview": store_list[:8],
        "latest_history": histories[0] if histories else None,
    }


def _task_summary_payload(conn, task_id):
    if not _table_exists(conn, "任务总结表"):
        return None
    row = conn.execute("""
        SELECT 总结ID AS summary_id, 任务ID AS task_id, 总结类型 AS summary_type, 总结来源 AS source,
               总结内容 AS content, 草稿内容 AS draft_content, 确认状态 AS confirm_status,
               确认人 AS confirmed_by, 确认时间 AS confirmed_at, 创建时间 AS created_at, 更新时间 AS updated_at
        FROM 任务总结表
        WHERE 任务ID = ?
        ORDER BY datetime(COALESCE(更新时间, 创建时间, '1970-01-01')) DESC, 总结ID DESC
        LIMIT 1
    """, (task_id,)).fetchone()
    return dict(row) if row else None


def _generate_summary_text(task_payload):
    stores = task_payload.get("stores") or []
    follow_records = task_payload.get("follow_records") or []
    followed_codes = {row.get("store_code") for row in follow_records if row.get("store_code")}
    continuous = sorted(stores, key=lambda item: item.get("continuous_days") or 0, reverse=True)[:5]
    reasons = {}
    for row in follow_records:
        reason = row.get("reason") or "未填写原因"
        reasons[reason] = reasons.get(reason, 0) + 1
    reason_text = "、".join([f"{key} {value} 次" for key, value in sorted(reasons.items(), key=lambda item: item[1], reverse=True)[:5]]) or "暂无跟进原因记录"
    continuous_text = "、".join([f"{item.get('store_name') or item.get('store_code')}（连续{item.get('continuous_days') or 0}天）" for item in continuous if (item.get("continuous_days") or 0) > 0]) or "暂无连续上榜门店"
    summary_type = "周总结" if task_payload.get("dimension") == "周任务" else "日总结"
    return "\n".join([
        f"{summary_type}：{task_payload.get('task_name') or ''}",
        f"本周期共纳入 {len(stores)} 家门店，已跟进 {len(followed_codes)} 家，未跟进 {max(len(stores) - len(followed_codes), 0)} 家。",
        f"主要跟进原因：{reason_text}。",
        f"连续上榜重点门店：{continuous_text}。",
        "建议：优先跟进连续上榜且到店数环比下降明显的门店，并在下一周期复核改善情况。",
    ])


def _load_task_detail_payload(conn, task_id):
    task = conn.execute("""
        SELECT 任务ID, 任务名称, 周开始日期, 门店列表, 状态, 创建时间, 完成时间, 完成人, 归档时间, 归档人,
               基准日期, 基准数据, 任务维度, 筛选日期, 筛选条件, 门店快照, 总结状态, 连续上榜快照
        FROM 跟进任务
        WHERE 任务ID = ? AND COALESCE(删除标记, 0) = 0
    """, (task_id,)).fetchone()
    if not task:
        return None
    task_row = dict(task)
    store_codes = _parse_task_store_list(task_row.get("门店列表"))
    stores = [_task_store_payload(conn, store_code, task_row) for store_code in store_codes]
    histories = _task_history_map(conn).get(str(task_id), [])
    follow_records = []
    if _table_exists(conn, "跟进记录"):
        follow_records = _dict_rows(conn.execute("""
            SELECT 记录ID AS record_id, 任务ID AS task_id, 日报数据日期 AS report_date, 店编号 AS store_code,
                   店简称 AS store_name, "线索量-本地" AS local_lead_count, 到店数 AS visit_count,
                   跟进原因 AS reason, 备注 AS remark, 操作人 AS operator, 创建时间 AS created_at, 跟进时间 AS follow_time
            FROM 跟进记录
            WHERE 任务ID = ?
            ORDER BY datetime(COALESCE(跟进时间, 创建时间, '1970-01-01')) DESC, 记录ID DESC
        """, (task_id,)))
    payload = _task_row_payload(conn, task_row)
    payload["baseline_data"] = task_row.get("基准数据") or ""
    payload["store_snapshot"] = []
    payload["stores"] = stores
    payload["history"] = histories
    payload["follow_records"] = follow_records
    payload["summary"] = _task_summary_payload(conn, task_id)
    return payload


def _query_governance_tasks():
    status = (request.args.get("status") or "").strip()
    dimension = (request.args.get("dimension") or "").strip()
    keyword = (request.args.get("keyword") or "").strip().lower()
    limit = min(max(int(request.args.get("limit", 50)), 1), 200)
    with _governance_conn() as conn:
        rows = _dict_rows(conn.execute("""
            SELECT 任务ID, 任务名称, 周开始日期, 门店列表, 状态, 创建时间, 完成时间, 完成人, 归档时间, 归档人,
                   基准日期, 任务维度, 筛选日期, 筛选条件, 门店快照, 总结状态, 连续上榜快照
            FROM 跟进任务
            WHERE COALESCE(删除标记, 0) = 0
            ORDER BY datetime(COALESCE(归档时间, 完成时间, 创建时间, '1970-01-01')) DESC, 任务ID DESC
        """))
        items = []
        for row in rows:
            payload = _task_row_payload(conn, row)
            if status and payload["status"] != status:
                continue
            if dimension and payload["dimension"] != dimension:
                continue
            if keyword and keyword not in str(payload["task_name"]).lower() and keyword not in str(payload["task_id"]).lower():
                continue
            items.append(payload)
        return items[:limit]


@app.route('/api/governance/overview', methods=['GET'])
@require_permission('follow.view')
def governance_overview():
    with _governance_conn() as conn:
        tasks = _query_governance_tasks()
        summary = {
            "task_count": len(tasks),
            "in_progress_count": len([item for item in tasks if item["status"] == "进行中"]),
            "completed_count": len([item for item in tasks if item["status"] == "已完成"]),
            "archived_count": len([item for item in tasks if item["status"] == "已归档"]),
            "store_count": sum(item["store_count"] for item in tasks),
            "latest_task_time": tasks[0]["created_at"] if tasks else "",
        }
    return jsonify({"success": True, "data": {"summary": summary, "tasks": tasks}})


@app.route('/api/governance/tasks', methods=['GET'])
@require_permission('follow.view')
def governance_tasks():
    tasks = _query_governance_tasks()
    return jsonify({"success": True, "data": tasks})


@app.route('/api/governance/filter-preview', methods=['POST'])
@require_permission('follow.view')
def governance_filter_preview():
    payload = request.get_json(silent=True) or {}
    with _governance_conn() as conn:
        rows, report_date, error = _preview_governance_stores(conn, payload)
    if error:
        return jsonify({"success": False, "message": error}), 400
    return jsonify({"success": True, "data": {"report_date": report_date, "stores": rows, "total": len(rows)}})


@app.route('/api/governance/tasks', methods=['POST'])
@require_permission('follow.task.create')
def governance_create_task():
    payload = request.get_json(silent=True) or {}
    dimension = payload.get("dimension") or "周任务"
    if dimension not in ("日任务", "周任务"):
        return jsonify({"success": False, "message": "任务维度必须为日任务或周任务"}), 400

    with _governance_conn() as conn:
        report_date = _normalize_date(payload.get("report_date") or payload.get("filter_date") or payload.get("baseline_date") or _latest_report_date(conn))
        if not report_date:
            return jsonify({"success": False, "message": "未找到可用日报数据，请先通过现有日报导入能力补充数据"}), 400

        stores = payload.get("stores")
        filters = payload.get("filters") or {}
        if not stores:
            stores, report_date, error = _preview_governance_stores(conn, {
                **payload,
                "report_date": report_date,
                "filters": filters,
            })
            if error:
                return jsonify({"success": False, "message": error}), 400

        store_codes = _unique_store_codes(stores)
        if not store_codes:
            return jsonify({"success": False, "message": "当前筛选条件下没有可保存的门店"}), 400

        task_name = (payload.get("task_name") or "").strip()
        if not task_name:
            task_name = f"{report_date} {'异常门店日任务' if dimension == '日任务' else '异常门店周任务'}"
        week_start_date = _normalize_date(payload.get("week_start_date") or "")
        continuous_snapshot = _build_continuous_snapshot(conn, store_codes)
        cursor = conn.execute("""
            INSERT INTO 跟进任务 (
                任务名称, 周开始日期, 门店列表, 状态, 创建时间, 基准日期, 基准数据,
                任务维度, 筛选日期, 筛选条件, 门店快照, 总结状态, 连续上榜快照
            )
            VALUES (?, ?, ?, '进行中', ?, ?, ?, ?, ?, ?, ?, '未生成', ?)
        """, (
            task_name,
            week_start_date,
            _json_dumps(store_codes),
            _now_text(),
            report_date,
            "[]",
            dimension,
            report_date,
            _json_dumps(filters),
            "[]",
            _json_dumps(continuous_snapshot),
        ))
        task_id = cursor.lastrowid
        _add_task_history(conn, task_id, "创建任务", "", "进行中", f"保存 {len(store_codes)} 家门店")
        conn.commit()

    record_audit_log(auth_db, "异常门店治理", "创建任务", "governance_task", task_id, after_data=payload)
    return jsonify({"success": True, "data": {"task_id": task_id}})


@app.route('/api/governance/tasks/<int:task_id>', methods=['GET'])
@require_permission('follow.view')
def governance_task_detail(task_id):
    with _governance_conn() as conn:
        payload = _load_task_detail_payload(conn, task_id)
        if not payload:
            return jsonify({"success": False, "message": "任务不存在"}), 404
    return jsonify({"success": True, "data": payload})


@app.route('/api/governance/tasks/<int:task_id>/stores', methods=['PUT'])
@require_permission('follow.task.edit')
def governance_update_task_stores(task_id):
    payload = request.get_json(silent=True) or {}
    with _governance_conn() as conn:
        task = conn.execute("SELECT * FROM 跟进任务 WHERE 任务ID = ? AND COALESCE(删除标记, 0) = 0", (task_id,)).fetchone()
        if not task:
            return jsonify({"success": False, "message": "任务不存在"}), 404
        stores = payload.get("stores")
        report_date = _normalize_date(payload.get("report_date") or dict(task).get("筛选日期") or dict(task).get("基准日期") or _latest_report_date(conn))
        filters = payload.get("filters") or _parse_json_object(dict(task).get("筛选条件") or "")
        if not stores:
            stores, report_date, error = _preview_governance_stores(conn, {
                **payload,
                "report_date": report_date,
                "filters": filters,
            })
            if error:
                return jsonify({"success": False, "message": error}), 400
        existing_codes = _parse_task_store_list(dict(task).get("门店列表"))
        incoming_codes = _unique_store_codes(stores)
        store_codes = _unique_store_codes(existing_codes + incoming_codes)
        if not store_codes:
            return jsonify({"success": False, "message": "门店列表不能为空"}), 400
        before_data = {"store_count": len(existing_codes)}
        previous_snapshot = _parse_json_object(dict(task).get("连续上榜快照") or "")
        continuous_snapshot = dict(previous_snapshot) if isinstance(previous_snapshot, dict) else {}
        new_codes = [code for code in incoming_codes if code not in continuous_snapshot]
        continuous_snapshot.update(_build_continuous_snapshot(conn, new_codes, task_id))
        conn.execute("""
            UPDATE 跟进任务
            SET 门店列表 = ?, 基准日期 = ?, 基准数据 = ?, 筛选日期 = ?, 筛选条件 = ?, 门店快照 = ?, 连续上榜快照 = ?
            WHERE 任务ID = ?
        """, (_json_dumps(store_codes), report_date, "[]", report_date, _json_dumps(filters), "[]", _json_dumps(continuous_snapshot), task_id))
        _add_task_history(conn, task_id, "更新门店列表", dict(task).get("状态") or "", dict(task).get("状态") or "", f"新增 {len(new_codes)} 家门店，当前共 {len(store_codes)} 家")
        conn.commit()
    record_audit_log(auth_db, "异常门店治理", "更新门店列表", "governance_task", task_id, before_data=before_data, after_data=payload)
    return jsonify({"success": True, "data": {"task_id": task_id, "store_count": len(store_codes)}})


@app.route('/api/governance/tasks/<int:task_id>', methods=['DELETE'])
@require_permission('follow.task.status')
def governance_delete_task(task_id):
    with _governance_conn() as conn:
        task = conn.execute("SELECT * FROM 跟进任务 WHERE 任务ID = ? AND COALESCE(删除标记, 0) = 0", (task_id,)).fetchone()
        if not task:
            return jsonify({"success": False, "message": "任务不存在"}), 404
        before_data = dict(task)
        operator = _current_operator_name()
        conn.execute("""
            UPDATE 跟进任务
            SET 删除标记 = 1, 删除时间 = ?, 删除人 = ?, 状态 = '已删除'
            WHERE 任务ID = ?
        """, (_now_text(), operator, task_id))
        _add_task_history(conn, task_id, "删除任务", task["状态"] or "", "已删除", "软删除任务，历史记录保留")
        conn.commit()
    record_audit_log(auth_db, "异常门店治理", "删除任务", "governance_task", task_id, before_data=before_data)
    return jsonify({"success": True, "data": {"task_id": task_id}})


@app.route('/api/governance/tasks/<int:task_id>/status', methods=['PATCH'])
@require_permission('follow.task.status')
def governance_update_task_status(task_id):
    payload = request.get_json(silent=True) or {}
    action = payload.get("action") or payload.get("status") or ""
    status_map = {
        "complete": "已完成",
        "archive": "已归档",
        "restore": "进行中",
        "已完成": "已完成",
        "已归档": "已归档",
        "进行中": "进行中",
    }
    target_status = status_map.get(action)
    if not target_status:
        return jsonify({"success": False, "message": "不支持的任务状态操作"}), 400
    with _governance_conn() as conn:
        task = conn.execute("SELECT * FROM 跟进任务 WHERE 任务ID = ? AND COALESCE(删除标记, 0) = 0", (task_id,)).fetchone()
        if not task:
            return jsonify({"success": False, "message": "任务不存在"}), 404
        before_status = task["状态"] or ""
        operator = _current_operator_name()
        now = _now_text()
        if target_status == "已完成":
            conn.execute("UPDATE 跟进任务 SET 状态 = ?, 完成时间 = ?, 完成人 = ? WHERE 任务ID = ?", (target_status, now, operator, task_id))
        elif target_status == "已归档":
            conn.execute("UPDATE 跟进任务 SET 状态 = ?, 归档时间 = ?, 归档人 = ? WHERE 任务ID = ?", (target_status, now, operator, task_id))
        else:
            conn.execute("UPDATE 跟进任务 SET 状态 = ?, 归档时间 = '', 归档人 = '' WHERE 任务ID = ?", (target_status, task_id))
        _add_task_history(conn, task_id, "状态流转", before_status, target_status, payload.get("remark") or "")
        conn.commit()
    record_audit_log(auth_db, "异常门店治理", "状态流转", "governance_task", task_id, before_data={"status": before_status}, after_data={"status": target_status})
    return jsonify({"success": True, "data": {"task_id": task_id, "status": target_status}})


@app.route('/api/governance/tasks/<int:task_id>/follow-records', methods=['POST'])
@require_permission('follow.record.edit')
def governance_save_follow_record(task_id):
    payload = request.get_json(silent=True) or {}
    store_code = str(payload.get("store_code") or "").strip()
    if not store_code:
        return jsonify({"success": False, "message": "请选择门店"}), 400
    reason = (payload.get("reason") or "").strip()
    remark = (payload.get("remark") or "").strip()
    if not reason and not remark:
        return jsonify({"success": False, "message": "请填写跟进原因或备注"}), 400
    with _governance_conn() as conn:
        task = conn.execute("SELECT * FROM 跟进任务 WHERE 任务ID = ? AND COALESCE(删除标记, 0) = 0", (task_id,)).fetchone()
        if not task:
            return jsonify({"success": False, "message": "任务不存在"}), 404
        task_row = dict(task)
        store = _task_store_payload(conn, store_code, task_row)
        if not store.get("store_code"):
            return jsonify({"success": False, "message": "门店不存在"}), 404
        now = _now_text()
        cursor = conn.execute("""
            INSERT INTO 跟进记录 (
                任务ID, 日报数据日期, 店编号, 店简称, "线索量-本地", 到店数,
                跟进原因, 备注, 操作人, 创建时间, 跟进时间
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            task_id,
            payload.get("report_date") or store.get("report_date") or task_row.get("筛选日期") or task_row.get("基准日期") or "",
            store_code,
            store.get("store_name") or "",
            str(store.get("local_lead_count") or 0),
            str(store.get("visit_count") or 0),
            reason,
            remark,
            _current_operator_name(),
            now,
            payload.get("follow_time") or now,
        ))
        record_id = cursor.lastrowid
        _add_task_history(conn, task_id, "保存跟进记录", task_row.get("状态") or "", task_row.get("状态") or "", f"{store_code} {reason}")
        conn.commit()
    record_audit_log(auth_db, "异常门店治理", "保存跟进记录", "governance_follow_record", record_id, after_data=payload)
    return jsonify({"success": True, "data": {"record_id": record_id}})


@app.route('/api/governance/follow-reasons', methods=['GET'])
@require_permission('follow.view')
def governance_follow_reasons():
    with _governance_conn() as conn:
        rows = _dict_rows(conn.execute("""
            SELECT
                r.配置ID AS id,
                r.原因选项 AS name,
                r.排序 AS sort_order,
                COALESCE(NULLIF(r.状态, ''), '启用') AS status,
                r.父级ID AS parent_id,
                p.原因选项 AS parent_name,
                r.创建时间 AS created_at
            FROM 跟进原因配置
            r
            LEFT JOIN 跟进原因配置 p ON r.父级ID = p.配置ID
            WHERE r.原因选项 IS NOT NULL
              AND r.原因选项 != ''
            ORDER BY COALESCE(p.排序, r.排序, 999999), COALESCE(r.排序, 999999), r.配置ID
        """))
    return jsonify({"success": True, "data": rows})


@app.route('/api/governance/follow-reasons', methods=['POST'])
@require_permission('follow.reason.config.manage')
def governance_create_follow_reason():
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    status = payload.get("status") or "启用"
    if not name:
        return jsonify({"success": False, "message": "跟进原因不能为空"}), 400
    with _governance_conn() as conn:
        is_group = bool(payload.get("is_group")) or payload.get("parent_id") == 0
        parent_id = 0 if is_group else payload.get("parent_id")
        existing = conn.execute("""
            SELECT 配置ID
            FROM 跟进原因配置
            WHERE 原因选项 = ? AND COALESCE(父级ID, -1) = ?
        """, (name, 0 if is_group else (parent_id or -1))).fetchone()
        if existing:
            return jsonify({"success": False, "message": "该分类或跟进原因已存在"}), 400
        max_sort = conn.execute("SELECT MAX(排序) FROM 跟进原因配置").fetchone()[0] or 0
        sort_order = payload.get("sort_order") if payload.get("sort_order") is not None else max_sort + 1
        cursor = conn.execute("""
            INSERT INTO 跟进原因配置 (原因选项, 排序, 状态, 父级ID, 创建时间)
            VALUES (?, ?, ?, ?, ?)
        """, (name, sort_order, status, parent_id, _now_text()))
        conn.commit()
        reason_id = cursor.lastrowid
    record_audit_log(auth_db, "异常门店治理", "新增跟进原因", "governance_follow_reason", reason_id, after_data=payload)
    return jsonify({"success": True, "data": {"id": reason_id}})


@app.route('/api/governance/follow-reasons/<int:reason_id>', methods=['PATCH'])
@require_permission('follow.reason.config.manage')
def governance_update_follow_reason(reason_id):
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    status = payload.get("status") or "启用"
    if not name:
        return jsonify({"success": False, "message": "跟进原因不能为空"}), 400
    with _governance_conn() as conn:
        before = conn.execute("SELECT * FROM 跟进原因配置 WHERE 配置ID = ?", (reason_id,)).fetchone()
        if not before:
            return jsonify({"success": False, "message": "跟进原因不存在"}), 404
        is_group = bool(payload.get("is_group")) or payload.get("parent_id") == 0
        parent_id = 0 if is_group else payload.get("parent_id")
        duplicate = conn.execute("""
            SELECT 配置ID
            FROM 跟进原因配置
            WHERE 原因选项 = ? AND COALESCE(父级ID, -1) = ? AND 配置ID != ?
        """, (name, 0 if is_group else (parent_id or -1), reason_id)).fetchone()
        if duplicate:
            return jsonify({"success": False, "message": "该分类或跟进原因已存在"}), 400
        conn.execute("""
            UPDATE 跟进原因配置
            SET 原因选项 = ?, 排序 = COALESCE(?, 排序), 状态 = ?, 父级ID = ?
            WHERE 配置ID = ?
        """, (name, payload.get("sort_order"), status, parent_id, reason_id))
        conn.commit()
    record_audit_log(auth_db, "异常门店治理", "编辑跟进原因", "governance_follow_reason", reason_id, before_data=dict(before), after_data=payload)
    return jsonify({"success": True})


@app.route('/api/governance/follow-reasons/<int:reason_id>', methods=['DELETE'])
@require_permission('follow.reason.config.manage')
def governance_delete_follow_reason(reason_id):
    with _governance_conn() as conn:
        before = conn.execute("SELECT * FROM 跟进原因配置 WHERE 配置ID = ?", (reason_id,)).fetchone()
        if not before:
            return jsonify({"success": False, "message": "跟进原因不存在"}), 404
        child_count = conn.execute("SELECT COUNT(*) FROM 跟进原因配置 WHERE 父级ID = ?", (reason_id,)).fetchone()[0]
        if child_count:
            return jsonify({"success": False, "message": "该一级原因下仍有二级原因，请先停用或迁移二级原因"}), 400
        conn.execute("DELETE FROM 跟进原因配置 WHERE 配置ID = ?", (reason_id,))
        conn.commit()
    record_audit_log(auth_db, "异常门店治理", "删除跟进原因", "governance_follow_reason", reason_id, before_data=dict(before))
    return jsonify({"success": True})


@app.route('/api/governance/tasks/<int:task_id>/summary', methods=['GET'])
@require_permission('follow.view')
def governance_get_summary(task_id):
    with _governance_conn() as conn:
        summary = _task_summary_payload(conn, task_id)
    return jsonify({"success": True, "data": summary})


@app.route('/api/governance/tasks/<int:task_id>/summary/generate', methods=['POST'])
@require_permission('follow.summary.generate')
def governance_generate_summary(task_id):
    with _governance_conn() as conn:
        task_payload = _load_task_detail_payload(conn, task_id)
        if not task_payload:
            return jsonify({"success": False, "message": "任务不存在"}), 404
        draft = _generate_summary_text(task_payload)
        summary_type = "周总结" if task_payload.get("dimension") == "周任务" else "日总结"
        existing = _task_summary_payload(conn, task_id)
        if existing:
            conn.execute("""
                UPDATE 任务总结表
                SET 总结类型 = ?, 总结来源 = 'AI生成', 草稿内容 = ?, 确认状态 = '待确认', 更新时间 = ?
                WHERE 总结ID = ?
            """, (summary_type, draft, _now_text(), existing["summary_id"]))
        else:
            conn.execute("""
                INSERT INTO 任务总结表 (任务ID, 总结类型, 总结来源, 草稿内容, 确认状态, 创建时间, 更新时间)
                VALUES (?, ?, 'AI生成', ?, '待确认', ?, ?)
            """, (task_id, summary_type, draft, _now_text(), _now_text()))
        conn.execute("UPDATE 跟进任务 SET 总结状态 = '待确认' WHERE 任务ID = ?", (task_id,))
        _add_task_history(conn, task_id, "生成AI总结", task_payload.get("status") or "", task_payload.get("status") or "", "已生成总结草稿，待人工确认")
        conn.commit()
        summary = _task_summary_payload(conn, task_id)
    record_audit_log(auth_db, "异常门店治理", "生成AI总结", "governance_summary", task_id)
    return jsonify({"success": True, "data": summary})


@app.route('/api/governance/tasks/<int:task_id>/summary', methods=['POST'])
@require_permission('follow.summary.confirm')
def governance_confirm_summary(task_id):
    payload = request.get_json(silent=True) or {}
    content = (payload.get("content") or "").strip()
    if not content:
        return jsonify({"success": False, "message": "总结内容不能为空"}), 400
    with _governance_conn() as conn:
        task = conn.execute("SELECT * FROM 跟进任务 WHERE 任务ID = ? AND COALESCE(删除标记, 0) = 0", (task_id,)).fetchone()
        if not task:
            return jsonify({"success": False, "message": "任务不存在"}), 404
        summary_type = "周总结" if _task_dimension(dict(task)) == "周任务" else "日总结"
        existing = _task_summary_payload(conn, task_id)
        if existing:
            conn.execute("""
                UPDATE 任务总结表
                SET 总结类型 = ?, 总结来源 = ?, 总结内容 = ?, 确认状态 = '已确认',
                    确认人 = ?, 确认时间 = ?, 更新时间 = ?
                WHERE 总结ID = ?
            """, (summary_type, payload.get("source") or existing.get("source") or "人工编辑", content, _current_operator_name(), _now_text(), _now_text(), existing["summary_id"]))
        else:
            conn.execute("""
                INSERT INTO 任务总结表 (任务ID, 总结类型, 总结来源, 总结内容, 确认状态, 确认人, 确认时间, 创建时间, 更新时间)
                VALUES (?, ?, ?, ?, '已确认', ?, ?, ?, ?)
            """, (task_id, summary_type, payload.get("source") or "人工编辑", content, _current_operator_name(), _now_text(), _now_text(), _now_text()))
        conn.execute("UPDATE 跟进任务 SET 总结状态 = '已确认' WHERE 任务ID = ?", (task_id,))
        _add_task_history(conn, task_id, "确认总结", task["状态"] or "", task["状态"] or "", "总结已人工确认入库")
        conn.commit()
        summary = _task_summary_payload(conn, task_id)
    record_audit_log(auth_db, "异常门店治理", "确认总结", "governance_summary", task_id, after_data=payload)
    return jsonify({"success": True, "data": summary})


def _store_filter_payload():
    stores = _raw_store_rows()
    region_zones = {}
    for row in stores:
        region = row.get("大区")
        zone = row.get("战区")
        if region and zone:
            region_zones.setdefault(region, set()).add(zone)
    with _governance_conn() as conn:
        store_statuses = _status_options(conn, "门店状态")
        statuses = _status_options(conn, "治理状态")
        ratings = _status_options(conn, "评级")
    return {
        "regions": sorted({row.get("大区") for row in stores if row.get("大区")}),
        "zones": sorted({row.get("战区") for row in stores if row.get("战区")}),
        "region_zones": {region: sorted(zones) for region, zones in region_zones.items()},
        "source_store_statuses": sorted({row.get("门店状态") for row in stores if row.get("门店状态")}),
        "store_statuses": store_statuses,
        "governance_statuses": statuses,
        "ratings": ratings,
    }


def _mart_dealer_region_zone_filters(conn):
    rows = conn.execute("""
        SELECT DISTINCT region, zone
        FROM mart_dealers
        WHERE (region IS NOT NULL AND TRIM(region) != '')
           OR (zone IS NOT NULL AND TRIM(zone) != '')
    """).fetchall()
    regions = set()
    zones = set()
    region_zones = {}
    for region, zone in rows:
        region = (region or "").strip()
        zone = (zone or "").strip()
        if region:
            regions.add(region)
        if zone:
            zones.add(zone)
        if region and zone:
            region_zones.setdefault(region, set()).add(zone)
    return {
        "regions": sorted(regions),
        "zones": sorted(zones),
        "region_zones": {region: sorted(values) for region, values in region_zones.items()},
        "store_statuses": _store_status_filter_options(),
    }


def _overdue_region_zone_filters(conn, start_date, end_date):
    rows = conn.execute("""
        SELECT DISTINCT region, zone
        FROM mart_dealer_overdue_leads
        WHERE assign_date >= CAST(? AS DATE) AND assign_date <= CAST(? AS DATE)
          AND ((region IS NOT NULL AND TRIM(region) != '')
            OR (zone IS NOT NULL AND TRIM(zone) != ''))
    """, [start_date, end_date]).fetchall()
    regions = set()
    zones = set()
    region_zones = {}
    for region, zone in rows:
        region = (region or "").strip()
        zone = (zone or "").strip()
        if region:
            regions.add(region)
        if zone:
            zones.add(zone)
        if region and zone:
            region_zones.setdefault(region, set()).add(zone)
    return {
        "regions": sorted(regions),
        "zones": sorted(zones),
        "region_zones": {region: sorted(values) for region, values in region_zones.items()},
    }


@app.route('/api/store-profile/filters', methods=['GET'])
@require_permission('store_profile.query')
def store_profile_filters():
    return jsonify({"success": True, "data": _store_filter_payload()})


@app.route('/api/store-profile/search', methods=['GET'])
@require_permission('store_profile.query')
def store_profile_search():
    keyword = (request.args.get("q") or "").strip().lower()
    result = []
    for row in _raw_store_rows():
        code = str(row.get("店编号") or "")
        name = str(row.get("店简称") or "")
        if keyword and keyword not in code.lower() and keyword not in name.lower():
            continue
        result.append({"store_code": code, "store_name": name, "region": row.get("大区") or "", "zone": row.get("战区") or ""})
        if len(result) >= 20:
            break
    return jsonify({"success": True, "data": result})


def _query_store_rows():
    page = max(int(request.args.get("page", 1)), 1)
    page_size = min(max(int(request.args.get("page_size", 50)), 1), 200)
    region = request.args.get("region", "")
    zone = request.args.get("zone", "")
    source_status = request.args.get("source_store_status", "")
    store_status = request.args.get("store_status", "")
    governance_status = request.args.get("governance_status", "")
    rating = request.args.get("store_rating", "")
    follow_status = request.args.get("follow_status", "")
    keyword = (request.args.get("search", "") or "").strip().lower()
    start_date = _normalize_date(request.args.get("start_date", ""))
    end_date = _normalize_date(request.args.get("end_date", ""))

    with _governance_conn() as conn:
        configs = _governance_config_map(conn)
        follow_stats = _follow_stats_map(conn)
        rows = []
        for store in _raw_store_rows():
            code = str(store.get("店编号") or "")
            name = str(store.get("店简称") or "")
            config = configs.get(code, {})
            follow = follow_stats.get(code, {})
            if region and store.get("大区") != region:
                continue
            if zone and store.get("战区") != zone:
                continue
            if source_status and store.get("门店状态") != source_status:
                continue
            managed_store_status = config.get("门店状态") or ""
            if store_status and managed_store_status != store_status:
                continue
            if governance_status and (config.get("治理状态") or "") != governance_status:
                continue
            if rating and (config.get("门店评级") or "") != rating:
                continue
            if follow_status == "followed" and not follow:
                continue
            if follow_status == "unfollowed" and follow:
                continue
            if keyword and keyword not in code.lower() and keyword not in name.lower():
                continue
            trend_rows = _daily_stats_for_store(conn, code, start_date, end_date)
            trend = {}
            if trend_rows:
                avg_leads = sum(row["local_lead_count"] for row in trend_rows) / len(trend_rows)
                avg_visits = sum(row["visit_count"] for row in trend_rows) / len(trend_rows)
                trend = {
                    "avg_local_lead_count": round(avg_leads, 1),
                    "avg_visit_count": round(avg_visits, 1),
                    "avg_visit_rate": round(avg_visits * 100 / avg_leads, 2) if avg_leads > 0 else 0,
                }
            rows.append(_store_profile_payload(store, config, follow, trend))
    total = len(rows)
    start = (page - 1) * page_size
    return rows[start:start + page_size], {"total": total, "page": page, "page_size": page_size, "total_pages": (total + page_size - 1) // page_size if total else 0}


@app.route('/api/store-profile/stores', methods=['GET'])
@require_permission('store_profile.query')
def store_profile_stores():
    items, pagination = _query_store_rows()
    return jsonify({"success": True, "data": items, "pagination": pagination})


@app.route('/api/store-profile/summary', methods=['GET'])
@require_permission('store_profile.summary')
def store_profile_summary():
    with _governance_conn() as conn:
        follow_stats = _follow_stats_map(conn)
        total_records = conn.execute("SELECT COUNT(*) FROM 跟进记录").fetchone()[0] if _table_exists(conn, "跟进记录") else 0
        reason_stats = []
        if _table_exists(conn, "跟进记录"):
            reason_stats = _dict_rows(conn.execute("""
                SELECT 跟进原因 AS reason, COUNT(*) AS count, COUNT(DISTINCT 店编号) AS store_count
                FROM 跟进记录
                WHERE 跟进原因 IS NOT NULL AND 跟进原因 != ''
                GROUP BY 跟进原因
                ORDER BY count DESC
                LIMIT 10
            """))
    followed_store_count = len(follow_stats)
    for row in reason_stats:
        row["rate"] = round(row["count"] * 100 / total_records, 2) if total_records else 0
    return jsonify({
        "success": True,
        "data": {
            "followed_store_count": followed_store_count,
            "follow_record_count": total_records,
            "avg_follow_count": round(total_records / followed_store_count, 2) if followed_store_count else 0,
            "high_freq_store_count": len([row for row in follow_stats.values() if int(row.get("follow_count") or 0) >= 3]),
            "reason_stats": reason_stats,
        },
    })


@app.route('/api/store-profile/frequent-stores', methods=['GET'])
@require_permission('store_profile.frequent.view')
def store_profile_frequent_stores():
    min_times = max(int(request.args.get("min_times", 3)), 1)
    stores = _store_base_map()
    with _governance_conn() as conn:
        configs = _governance_config_map(conn)
        follow_stats = _follow_stats_map(conn)
    rows = []
    for code, follow in follow_stats.items():
        if int(follow.get("follow_count") or 0) >= min_times and code in stores:
            rows.append(_store_profile_payload(stores[code], configs.get(code), follow))
    rows.sort(key=lambda row: row["follow_count"], reverse=True)
    return jsonify({"success": True, "data": rows})


@app.route('/api/store-profile/reason-analysis', methods=['GET'])
@require_permission('store_profile.reason_analysis.view')
def store_profile_reason_analysis():
    with _governance_conn() as conn:
        total_records = conn.execute("SELECT COUNT(*) FROM 跟进记录").fetchone()[0] if _table_exists(conn, "跟进记录") else 0
        reasons = []
        if _table_exists(conn, "跟进记录"):
            reasons = _dict_rows(conn.execute("""
                SELECT 跟进原因 AS reason, COUNT(*) AS count, COUNT(DISTINCT 店编号) AS store_count
                FROM 跟进记录
                WHERE 跟进原因 IS NOT NULL AND 跟进原因 != ''
                GROUP BY 跟进原因
                ORDER BY count DESC
            """))
    for row in reasons:
        row["rate"] = round(row["count"] * 100 / total_records, 2) if total_records else 0
    return jsonify({"success": True, "data": {"reason_distribution": reasons}})


@app.route('/api/store-profile/<store_code>/basic-info', methods=['GET'])
@require_permission('store_profile.detail.view')
def store_profile_basic_info(store_code):
    store = _store_base_map().get(store_code)
    if not store:
        return jsonify({"success": False, "message": "门店不存在"}), 404
    with _governance_conn() as conn:
        config = _governance_config_map(conn).get(store_code, {})
        follow = _follow_stats_map(conn).get(store_code, {})
    data = _store_profile_payload(store, config, follow)
    data["monthly_summary"] = _monthly_summary_for_store(store_code)
    return jsonify({"success": True, "data": data})


@app.route('/api/store-profile/<store_code>/daily-stats', methods=['GET'])
@require_permission('store_profile.detail.view')
def store_profile_daily_stats(store_code):
    with _governance_conn() as conn:
        data = _daily_stats_for_store(conn, store_code, _normalize_date(request.args.get("start_date", "")), _normalize_date(request.args.get("end_date", "")))
    return jsonify({"success": True, "data": data})


@app.route('/api/store-profile/<store_code>/follow-history', methods=['GET'])
@require_permission('store_profile.detail.view')
def store_profile_follow_history(store_code):
    with _governance_conn() as conn:
        if not _table_exists(conn, "跟进记录"):
            return jsonify({"success": True, "data": []})
        rows = _dict_rows(conn.execute("""
            SELECT r.记录ID AS record_id, r.任务ID AS task_id, r.日报数据日期 AS report_date,
                   r.跟进原因 AS reason, r.备注 AS remark, r.操作人 AS operator,
                   r.创建时间 AS created_at, r.跟进时间 AS follow_time,
                   t.任务名称 AS task_name, t.周开始日期 AS task_week_start_date, t.状态 AS task_status,
                   t.任务维度 AS task_dimension
            FROM 跟进记录 r
            LEFT JOIN 跟进任务 t ON r.任务ID = t.任务ID
            WHERE r.店编号 = ?
            ORDER BY COALESCE(r.跟进时间, r.创建时间) DESC
        """, (store_code,)))
    weekly_rows = {}
    for row in rows:
        row["normalized_report_date"] = _normalize_date(row.get("report_date"))
        completed_at = row.get("follow_time") or row.get("created_at") or ""
        row["completed_at"] = completed_at
        week_start, week_end, week_label = _week_key_for_follow_record(row)
        if not week_start:
            week_start = row.get("task_week_start_date") or ""
            week_end = _shift_date_str(week_start, 6) if week_start else ""
            week_label = f"{week_start} 至 {week_end}" if week_start and week_end else week_start
        row["week_start_date"] = week_start
        row["week_end_date"] = week_end
        row["week_label"] = week_label
        current = weekly_rows.get(week_start)
        if current is None or str(completed_at or "") > str(current.get("completed_at") or ""):
            weekly_rows[week_start] = row
    result = sorted(weekly_rows.values(), key=lambda item: item.get("completed_at") or "", reverse=True)
    return jsonify({"success": True, "data": result})


@app.route('/api/store-profile/<store_code>/reason-analysis', methods=['GET'])
@require_permission('store_profile.detail.view')
def store_profile_store_reason_analysis(store_code):
    with _governance_conn() as conn:
        if not _table_exists(conn, "跟进记录"):
            return jsonify({"success": True, "data": []})
        category_map = _reason_category_map(conn)
        raw_rows = _dict_rows(conn.execute("""
            SELECT 跟进原因 AS reason, COALESCE(跟进时间, 创建时间) AS latest_time
            FROM 跟进记录
            WHERE 店编号 = ? AND 跟进原因 IS NOT NULL AND 跟进原因 != ''
        """, (store_code,)))
    stats = {}
    for row in raw_rows:
        latest_time = row.get("latest_time") or ""
        for reason in _split_follow_reasons(row.get("reason")):
            item = stats.setdefault(reason, {
                "reason": reason,
                "reason_category": category_map.get(reason, "未分类"),
                "count": 0,
                "latest_time": "",
            })
            item["count"] += 1
            if str(latest_time) > str(item.get("latest_time") or ""):
                item["latest_time"] = latest_time
    rows = sorted(stats.values(), key=lambda item: (int(item.get("count") or 0), str(item.get("latest_time") or "")), reverse=True)
    return jsonify({"success": True, "data": rows})


@app.route('/api/store-management/filters', methods=['GET'])
@require_permission('store_management.query')
def store_management_filters():
    return jsonify({"success": True, "data": _store_filter_payload()})


@app.route('/api/store-management/stores', methods=['GET'])
@require_permission('store_management.query')
def store_management_stores():
    items, pagination = _query_store_rows()
    return jsonify({"success": True, "data": items, "pagination": pagination})


@app.route('/api/store-management/stores/<store_code>', methods=['PATCH'])
@require_permission('store_management.edit')
def store_management_update_store(store_code):
    payload = request.get_json(silent=True) or {}
    with _governance_conn() as conn:
        before = conn.execute("SELECT * FROM 门店管理配置表 WHERE 店编号 = ?", (store_code,)).fetchone()
        before_data = dict(before) if before else None
        existing = before_data or {}
        conn.execute("""
            INSERT INTO 门店管理配置表 (店编号, 门店状态, 治理状态, 门店评级, 状态备注, 管理员备注, 更新时间)
            VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(店编号) DO UPDATE SET
                门店状态 = excluded.门店状态,
                治理状态 = excluded.治理状态,
                门店评级 = excluded.门店评级,
                状态备注 = excluded.状态备注,
                管理员备注 = excluded.管理员备注,
                更新时间 = datetime('now')
        """, (
            store_code,
            payload.get("store_status", existing.get("门店状态") or "") or "",
            payload.get("governance_status", existing.get("治理状态") or "") or "",
            payload.get("store_rating", existing.get("门店评级") or "") or "",
            payload.get("status_note", existing.get("状态备注") or "") or "",
            payload.get("admin_note", existing.get("管理员备注") or "") or "",
        ))
        conn.commit()
    record_audit_log(auth_db, "门店管理", "编辑治理信息", "store_governance", store_code, before_data=before_data, after_data=payload)
    return jsonify({"success": True, "message": "保存成功"})


@app.route('/api/store-management/statuses', methods=['GET'])
@require_permission('store_management.config.view')
def store_management_statuses():
    with _governance_conn() as conn:
        return jsonify({"success": True, "data": _status_options(conn, request.args.get("type", "治理状态"))})


@app.route('/api/store-management/statuses', methods=['POST'])
@require_permission('store_management.config.manage')
def store_management_create_status():
    payload = request.get_json(silent=True) or {}
    config_type = payload.get("config_type") or "治理状态"
    name = (payload.get("name") or "").strip()
    color = payload.get("color") or "#64748b"
    enabled = 1 if payload.get("enabled", True) else 0
    if not name:
        return jsonify({"success": False, "message": "名称不能为空"}), 400
    with _governance_conn() as conn:
        max_sort = conn.execute("SELECT MAX(排序) FROM 门店状态配置表 WHERE 配置类型 = ?", (config_type,)).fetchone()[0] or 0
        cursor = conn.execute("""
            INSERT INTO 门店状态配置表 (状态名称, 状态颜色, 排序, 配置类型, 启用状态, 说明)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (name, color, max_sort + 1, config_type, enabled, payload.get("description") or ""))
        conn.commit()
        status_id = cursor.lastrowid
    record_audit_log(auth_db, "门店管理", f"新增{config_type}", "store_status_config", status_id, after_data=payload)
    return jsonify({"success": True, "data": {"id": status_id}})


@app.route('/api/store-management/statuses/<int:status_id>', methods=['PATCH'])
@require_permission('store_management.config.manage')
def store_management_update_status(status_id):
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"success": False, "message": "名称不能为空"}), 400
    with _governance_conn() as conn:
        before = conn.execute("SELECT * FROM 门店状态配置表 WHERE 状态ID = ?", (status_id,)).fetchone()
        if not before:
            return jsonify({"success": False, "message": "配置不存在"}), 404
        enabled = 1 if payload.get("enabled", before["启用状态"] if "启用状态" in before.keys() else 1) else 0
        conn.execute("""
            UPDATE 门店状态配置表
            SET 状态名称 = ?, 状态颜色 = ?, 排序 = COALESCE(?, 排序), 启用状态 = ?, 说明 = ?
            WHERE 状态ID = ?
        """, (name, payload.get("color") or "#64748b", payload.get("sort_order"), enabled, payload.get("description") or "", status_id))
        conn.commit()
    record_audit_log(auth_db, "门店管理", "编辑配置", "store_status_config", status_id, before_data=dict(before), after_data=payload)
    return jsonify({"success": True})


@app.route('/api/store-management/statuses/<int:status_id>', methods=['DELETE'])
@require_permission('store_management.config.manage')
def store_management_delete_status(status_id):
    with _governance_conn() as conn:
        before = conn.execute("SELECT * FROM 门店状态配置表 WHERE 状态ID = ?", (status_id,)).fetchone()
        if not before:
            return jsonify({"success": False, "message": "配置不存在"}), 404
        conn.execute("DELETE FROM 门店状态配置表 WHERE 状态ID = ?", (status_id,))
        conn.commit()
    record_audit_log(auth_db, "门店管理", "删除配置", "store_status_config", status_id, before_data=dict(before))
    return jsonify({"success": True})


@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'success': True,
        'status': 'ok',
        'version': '2.0.0',
        'data': {
            'layers': {
                'raw': 'connected',
                'mart': 'connected',
                'metric': 'connected'
            }
        }
    })


@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    """获取仪表盘数据
    参数: period - "day" (昨日数据) or "month" (当月累计数据)
    """
    try:
        period = request.args.get('period', 'day')
        data = duck_db.get_dashboard_data(period=period)
        _decorate_rows_with_store_status(data.get('dealer_ranking') or [])
        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/visit_stats/export', methods=['GET'])
def export_visit_stats():
    try:
        import openpyxl
        
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        region = request.args.get('region', '')
        zone = request.args.get('zone', '')
        dealer_code = request.args.get('dealer_code', '')
        store_status = request.args.get('store_status', '')
        
        conn = duck_db.get_connection()
        sqlite_path = str(raw_db.db_path)
        
        date_filter = ""
        params = []
        
        if date_from:
            date_filter += " AND CAST(visit_time AS DATE) >= ?"
            params.append(date_from)
        
        if date_to:
            date_filter += " AND CAST(visit_time AS DATE) <= ?"
            params.append(date_to)
        
        dealer_filter = ""
        dealer_params = []
        
        if region:
            dealer_filter += " AND d.region = ?"
            dealer_params.append(region)
        
        if zone:
            dealer_filter += " AND d.zone = ?"
            dealer_params.append(zone)
        
        if dealer_code:
            dealer_filter += " AND dv.dealer_id = ?"
            dealer_params.append(dealer_code)
        store_status_ids = _store_status_filter_ids(store_status)
        store_status_filter, store_status_params = _dealer_id_filter_sql("d", store_status_ids)
        dealer_filter += store_status_filter
        dealer_params.extend(store_status_params)
        
        sqlite_path = str(raw_db.db_path)
        
        sql = f"""
            WITH dealer_visits AS (
                SELECT 
                    f.dealer_id,
                    COUNT(*) as visit_count,
                    COUNT(DISTINCT m.lead_id || '_' || CAST(m.visit_time AS DATE)) as lead_count,
                    COUNT(CASE WHEN m.channel_1 = '线上' THEN 1 END) as online_count,
                    COUNT(CASE WHEN m.channel_1 = '线下' THEN 1 END) as offline_count,
                    COUNT(DISTINCT CASE WHEN m.channel_1 = '线上' THEN m.lead_id || '_' || CAST(m.visit_time AS DATE) END) as online_lead_count,
                    COUNT(DISTINCT CASE WHEN m.channel_1 = '线下' THEN m.lead_id || '_' || CAST(m.visit_time AS DATE) END) as offline_lead_count
                FROM (
                    SELECT DISTINCT dealer_id, CAST(visit_time AS DATE) as visit_date
                    FROM mart_customer_visit
                    WHERE 1=1 {date_filter}
                ) f
                JOIN mart_customer_visit m ON f.dealer_id = m.dealer_id AND CAST(m.visit_time AS DATE) = f.visit_date
                GROUP BY f.dealer_id
            )
            SELECT 
                d.region,
                d.zone,
                d.dealer_id,
                d.dealer_name,
                COALESCE(dv.visit_count, 0) as visit_count,
                COALESCE(dv.lead_count, 0) as lead_count,
                COALESCE(dv.online_count, 0) as online_count,
                COALESCE(dv.online_lead_count, 0) as online_lead_count,
                COALESCE(dv.offline_count, 0) as offline_count,
                COALESCE(dv.offline_lead_count, 0) as offline_lead_count
            FROM mart_dealers d
            LEFT JOIN dealer_visits dv ON d.dealer_id = dv.dealer_id
            WHERE 1=1 {dealer_filter}
            ORDER BY d.source_rowid
        """
        
        results = conn.execute(sql, params + dealer_params).fetchall()
        
        columns = _xlsx_headers_with_store_status(
            ['大区', '战区', '门店编号', '门店名称', '进店次数', '进店客流', '线上进店数', '线上进店客流', '线下进店数', '线下进店客流'],
            4,
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "客流统计"
        
        ws.append(columns)
        
        for row in _xlsx_rows_with_store_status(results, dealer_index=2, insert_index=4):
            row_data = [val if val is not None else '' for val in row]
            ws.append(row_data)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"客流统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/visit_stats', methods=['GET'])
def get_visit_stats():
    try:
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        region = request.args.get('region', '')
        zone = request.args.get('zone', '')
        dealer_code = request.args.get('dealer_code', '')
        store_status = request.args.get('store_status', '')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 100))
        
        conn = duck_db.get_connection()
        
        date_filter = ""
        params = []
        
        if date_from:
            date_filter += " AND CAST(visit_time AS DATE) >= ?"
            params.append(date_from)
        
        if date_to:
            date_filter += " AND CAST(visit_time AS DATE) <= ?"
            params.append(date_to)
        
        dealer_filter = ""
        dealer_params = []
        
        if region:
            dealer_filter += " AND d.region = ?"
            dealer_params.append(region)
        
        if zone:
            dealer_filter += " AND d.zone = ?"
            dealer_params.append(zone)
        
        if dealer_code:
            dealer_filter += " AND dv.dealer_id = ?"
            dealer_params.append(dealer_code)
        store_status_ids = _store_status_filter_ids(store_status)
        store_status_filter, store_status_params = _dealer_id_filter_sql("d", store_status_ids)
        dealer_filter += store_status_filter
        dealer_params.extend(store_status_params)
        
        sqlite_path = str(raw_db.db_path)
        
        sql = f"""
            WITH dealer_visits AS (
                SELECT 
                    f.dealer_id,
                    COUNT(*) as visit_count,
                    COUNT(DISTINCT m.lead_id || '_' || CAST(m.visit_time AS DATE)) as lead_count,
                    COUNT(CASE WHEN m.channel_1 = '线上' THEN 1 END) as online_count,
                    COUNT(CASE WHEN m.channel_1 = '线下' THEN 1 END) as offline_count,
                    COUNT(DISTINCT CASE WHEN m.channel_1 = '线上' THEN m.lead_id || '_' || CAST(m.visit_time AS DATE) END) as online_lead_count,
                    COUNT(DISTINCT CASE WHEN m.channel_1 = '线下' THEN m.lead_id || '_' || CAST(m.visit_time AS DATE) END) as offline_lead_count
                FROM (
                    SELECT DISTINCT dealer_id, CAST(visit_time AS DATE) as visit_date
                    FROM mart_customer_visit
                    WHERE 1=1 {date_filter}
                ) f
                JOIN mart_customer_visit m ON f.dealer_id = m.dealer_id AND CAST(m.visit_time AS DATE) = f.visit_date
                GROUP BY f.dealer_id
            )
            SELECT 
                d.region,
                d.zone,
                d.dealer_id,
                d.dealer_name,
                COALESCE(dv.visit_count, 0) as visit_count,
                COALESCE(dv.lead_count, 0) as lead_count,
                COALESCE(dv.online_count, 0) as online_count,
                COALESCE(dv.online_lead_count, 0) as online_lead_count,
                COALESCE(dv.offline_count, 0) as offline_count,
                COALESCE(dv.offline_lead_count, 0) as offline_lead_count
            FROM mart_dealers d
            LEFT JOIN dealer_visits dv ON d.dealer_id = dv.dealer_id
            WHERE 1=1 {dealer_filter}
            ORDER BY d.source_rowid
        """
        
        results = conn.execute(sql, params + dealer_params).fetchall()
        
        total_grand = {
            'total_visits': sum(row[4] for row in results),
            'unique_lead_visits': sum(row[5] for row in results),
            'online_visits': sum(row[6] for row in results),
            'online_lead_visits': sum(row[7] for row in results),
            'offline_visits': sum(row[8] for row in results),
            'offline_lead_visits': sum(row[9] for row in results),
        }
        
        data = []
        for row in results:
            data.append({
                'region': row[0] or '',
                'zone': row[1] or '',
                'dealer_id': row[2],
                'dealer_name': row[3] or '',
                'total_visits': row[4] or 0,
                'unique_lead_visits': row[5] or 0,
                'online_visits': row[6] or 0,
                'online_lead_visits': row[7] or 0,
                'offline_visits': row[8] or 0,
                'offline_lead_visits': row[9] or 0
            })
        _decorate_rows_with_store_status(data)
        
        total = len(data)
        offset = (page - 1) * page_size
        paged_data = data[offset:offset + page_size]
        region_zone_filters = _mart_dealer_region_zone_filters(conn)
        
        return jsonify({
            'success': True,
            'data': paged_data,
            'pagination': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size if page_size > 0 else 1
            },
            'grand_total': total_grand,
            'filters': region_zone_filters
        })
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/metrics', methods=['GET'])
def get_metrics():
    """获取指标列表"""
    metrics = [
        {'name': 'lead_count', 'display_name': '线索数量', 'category': 'kpi', 'description': '统计周期内的线索总数'},
        {'name': 'conversion_rate', 'display_name': '转化率', 'category': 'kpi', 'description': '转化线索占比'},
        {'name': 'follow_in_30min_rate', 'display_name': '30分钟跟进率', 'category': 'kpi', 'description': '30分钟内跟进线索占比'},
        {'name': 'to_shop_rate', 'display_name': '到店率', 'category': 'kpi', 'description': '到店线索占比'}
    ]
    return jsonify({'success': True, 'data': metrics})


@app.route('/api/dealers', methods=['GET'])
def get_dealers():
    """获取经销商列表"""
    try:
        dealers = raw_db.get_dealers()
        return jsonify({
            'success': True,
            'data': [dict(d) for d in dealers]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/refresh/trigger', methods=['POST'])
def trigger_refresh():
    """触发数据刷新
    mode: 'full' (全量重算) | 'incremental' (增量同步新数据) | 'recompute' (仅重算指标) | 'refresh_personnel' (刷新人员信息)
    """
    try:
        data = request.get_json(silent=True) or {}
        mode = data.get('mode', 'full')
        import gc, time

        with duckdb_maintenance_lock("refreshing DuckDB"):
            duck_db.close()
            gc.collect()
            time.sleep(0.3)

            try:
                if mode == 'full':
                    print("Starting full refresh...")
                    _rebuild_duckdb_from_sqlite()
                    stats = duck_db.get_count_stats()
                    print(f"Full refresh completed: {stats}")

                elif mode == 'incremental':
                    print("Starting incremental sync...")
                    new_count = duck_db.load_incremental()
                    if new_count > 0:
                        duck_db.compute_all_metrics()
                        print(f"Incremental sync: {new_count} new leads added")
                    else:
                        print("Incremental sync: no new data found")
                    stats = duck_db.get_count_stats()
                elif mode == 'recompute':
                    date_str = data.get('date')
                    if date_str:
                        duck_db.compute_all_metrics(date_str=date_str)
                    else:
                        duck_db.compute_all_metrics()
                    stats = duck_db.get_count_stats()

                elif mode == 'refresh_personnel':
                    print("Refreshing personnel info...")
                    duck_db.refresh_personnel_info()
                    stats = duck_db.get_count_stats()

                else:
                    return jsonify({
                        'success': False,
                        'message': f'Unknown mode: {mode}'
                    }), 400
            except Exception as exc:
                if not archive_unreadable_duckdb(exc):
                    raise
                print("DuckDB hit a corruption/fatal index error, rebuilding from raw SQLite...")
                duck_db.close()
                _reset_duck_db()
                _rebuild_duckdb_from_sqlite()
                stats = duck_db.get_count_stats()

            duck_db.close()
            gc.collect()

            return jsonify({
                'success': True,
                'data': {
                    'task_id': 'refresh_' + datetime.now().strftime('%Y%m%d%H%M%S'),
                    'status': 'completed',
                    'mode': mode,
                    'stats': stats
                }
            })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/explore/leads', methods=['GET'])
def explore_leads():
    """探索线索数据"""
    try:
        conn = duck_db.get_connection()
        
        filters = []
        params = []
        
        dealer_id = request.args.get('dealer_id')
        if dealer_id:
            filters.append("dealer_id = ?")
            params.append(dealer_id)
        
        region = request.args.get('region')
        if region:
            filters.append("region = ?")
            params.append(region)
        
        channel_1 = request.args.get('channel_1')
        if channel_1:
            filters.append("channel_1 = ?")
            params.append(channel_1)
        
        is_converted = request.args.get('is_converted')
        if is_converted is not None:
            filters.append("is_converted = ?")
            params.append(is_converted.lower() == 'true')
        
        where_clause = " AND ".join(filters) if filters else "1=1"
        
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        offset = (page - 1) * page_size
        
        count_query = f"SELECT COUNT(*) FROM mart_leads WHERE " + where_clause
        total = conn.execute(count_query, params).fetchone()[0]
        
        query = f"""
            SELECT lead_id, dealer_id, dealer_name, region, channel_1,
                   assign_date, first_follow_date, is_converted, conversion_date
            FROM mart_leads
            WHERE {where_clause}
            ORDER BY assign_date DESC
            LIMIT ? OFFSET ?
        """
        results = conn.execute(query, params + [page_size, offset]).fetchall()
        
        return jsonify({
            'success': True,
            'data': {
                'list': [dict(zip([d[0] for d in results.description], r)) for r in results],
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total
                }
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/explore/aggregate', methods=['POST'])
def aggregate_data():
    """数据聚合查询"""
    try:
        data = request.get_json(silent=True) or {}
        group_by = data.get('group_by', ['channel_1'])
        metrics = data.get('metrics', ['lead_count'])
        
        valid_fields = ['channel_1', 'region', 'dealer_id', 'assign_date']
        group_by = [f for f in group_by if f in valid_fields]
        
        if not group_by:
            return jsonify({'success': False, 'message': 'No valid group fields'}), 400
        
        conn = duck_db.get_connection()
        
        group_clause = ", ".join(group_by)
        select_clause = ", ".join(group_by)
        
        metric_calcs = []
        for m in metrics:
            if m == 'lead_count':
                metric_calcs.append("COUNT(*) AS lead_count")
            elif m == 'conversion_count':
                metric_calcs.append("SUM(CASE WHEN is_converted THEN 1 ELSE 0 END) AS conversion_count")
            elif m == 'conversion_rate':
                metric_calcs.append("AVG(CASE WHEN is_converted THEN 1 ELSE 0 END) * 100 AS conversion_rate")
        
        if metric_calcs:
            select_clause += ", " + ", ".join(metric_calcs)
        
        filters = data.get('filters', {})
        where_parts = []
        params = []
        
        if 'start_date' in filters:
            where_parts.append("assign_date >= ?")
            params.append(filters['start_date'])
        if 'end_date' in filters:
            where_parts.append("assign_date <= ?")
            params.append(filters['end_date'])
        
        where_clause = " AND ".join(where_parts) if where_parts else "1=1"
        
        query = f"SELECT {select_clause} FROM mart_leads WHERE {where_clause} GROUP BY {group_clause}"
        results = conn.execute(query, params).fetchall()
        
        return jsonify({
            'success': True,
            'data': [dict(zip([d[0] for d in results.description], r)) for r in results]
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/follow-up/distribution', methods=['GET'])
def get_follow_up_distribution():
    """获取跟进次数分布数据
    
    支持自定义时间范围：start_date, end_date (格式: YYYY-MM-DD)
    默认：当月累计至前一天18:00前
    线索标准：一级渠道为线上且四级渠道不包含"反写"
    跟进次数标准：线索表里的"总跟进次数"字段，空值代表未跟进
    """
    try:
        from datetime import datetime, timedelta, date
        
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        today = date.today()
        
        if start_date_str:
            try:
                start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                start_datetime = datetime(start_dt.year, start_dt.month, start_dt.day, 0, 0, 0)
            except ValueError:
                return jsonify({'success': False, 'message': 'start_date 格式错误，应为 YYYY-MM-DD'}), 400
        else:
            month_start = date(today.year, today.month, 1)
            start_datetime = datetime(month_start.year, month_start.month, month_start.day, 0, 0, 0)
        
        if end_date_str:
            try:
                end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
                yesterday = today - timedelta(days=1)
                if date(end_dt.year, end_dt.month, end_dt.day) >= yesterday:
                    end_datetime = datetime(end_dt.year, end_dt.month, end_dt.day, 18, 0, 0)
                else:
                    end_datetime = datetime(end_dt.year, end_dt.month, end_dt.day, 23, 59, 59)
            except ValueError:
                return jsonify({'success': False, 'message': 'end_date 格式错误，应为 YYYY-MM-DD'}), 400
        else:
            yesterday = today - timedelta(days=1)
            end_datetime = datetime(yesterday.year, yesterday.month, yesterday.day, 18, 0, 0)
        
        # 获取门店列表
        dealers = raw_db.get_dealers()
        dealer_list = [dict(d) for d in dealers]
        
        # 获取跟进次数分布数据
        with raw_db.get_connection() as conn:
            query = """
                SELECT 
                    d.店编号 as dealer_id,
                    d.店简称 as dealer_name,
                    d.大区 as region,
                    d.战区 as zone,
                    SUM(CASE WHEN CAST(COALESCE(NULLIF(l.总跟进次数, ''), '0') AS INTEGER) = 0 THEN 1 ELSE 0 END) as follow_0,
                    SUM(CASE WHEN CAST(COALESCE(NULLIF(l.总跟进次数, ''), '0') AS INTEGER) = 1 THEN 1 ELSE 0 END) as follow_1,
                    SUM(CASE WHEN CAST(COALESCE(NULLIF(l.总跟进次数, ''), '0') AS INTEGER) = 2 THEN 1 ELSE 0 END) as follow_2,
                    SUM(CASE WHEN CAST(COALESCE(NULLIF(l.总跟进次数, ''), '0') AS INTEGER) = 3 THEN 1 ELSE 0 END) as follow_3,
                    SUM(CASE WHEN CAST(COALESCE(NULLIF(l.总跟进次数, ''), '0') AS INTEGER) >= 4 THEN 1 ELSE 0 END) as follow_4_plus,
                    COUNT(*) as total
                FROM 门店表 d
                LEFT JOIN 线索表 l ON d.店编号 = l.门店
                WHERE 
                    l.一级渠道 = '线上'
                    AND (l.四级渠道 IS NULL OR l.四级渠道 NOT LIKE '%反写%')
                    AND l.跟进截止时间 IS NOT NULL AND l.跟进截止时间 != ''
                    AND l.最终下发时间 >= ?
                    AND l.最终下发时间 <= ?
                GROUP BY d.店编号, d.店简称, d.大区, d.战区
                ORDER BY d.rowid
            """
            
            results = conn.execute(query, [
                start_datetime.strftime("%Y-%m-%d %H:%M:%S"),
                end_datetime.strftime("%Y-%m-%d %H:%M:%S")
            ]).fetchall()
            
            distribution_data = []
            for row in results:
                row_dict = {
                    'dealer_id': row['dealer_id'],
                    'dealer_name': row['dealer_name'],
                    'region': row['region'] or '',
                    'zone': row['zone'] or '',
                    'follow_0': row['follow_0'] or 0,
                    'follow_1': row['follow_1'] or 0,
                    'follow_2': row['follow_2'] or 0,
                    'follow_3': row['follow_3'] or 0,
                    'follow_4_plus': row['follow_4_plus'] or 0,
                    'total': row['total'] or 0,
                    'follow_0_rate': round(row['follow_0'] * 100.0 / (row['total'] or 1), 1),
                    'follow_1_rate': round(row['follow_1'] * 100.0 / (row['total'] or 1), 1),
                    'follow_2_rate': round(row['follow_2'] * 100.0 / (row['total'] or 1), 1),
                    'follow_3_rate': round(row['follow_3'] * 100.0 / (row['total'] or 1), 1),
                    'follow_4_plus_rate': round(row['follow_4_plus'] * 100.0 / (row['total'] or 1), 1)
                }
                distribution_data.append(row_dict)
        
        return jsonify({
            'success': True,
            'data': {
                'dealers': dealer_list,
                'distribution': distribution_data,
                'time_range': {
                    'start_date': start_datetime.strftime("%Y-%m-%d"),
                    'end_date': end_datetime.strftime("%Y-%m-%d"),
                    'description': f"统计范围：{start_datetime.strftime('%Y年%m月%d日')} 至 {end_datetime.strftime('%Y年%m月%d日%H:%M')}"
                }
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/customer_visit/detail', methods=['GET'])
def get_customer_visit_detail():
    try:
        dealer_code = request.args.get('dealer_code', '')
        channel_1 = request.args.get('channel_1', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        phone = request.args.get('phone', '')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 100))
        
        conn = duck_db.get_connection()
        
        where_clauses = ["m.dealer_id IS NOT NULL"]
        params = []
        
        if dealer_code:
            where_clauses.append("m.dealer_id = ?")
            params.append(dealer_code)
        
        if channel_1:
            if channel_1 == '__EMPTY__':
                where_clauses.append("(m.channel_1 IS NULL OR m.channel_1 = '')")
            else:
                where_clauses.append("m.channel_1 = ?")
                params.append(channel_1)
        
        if date_from:
            where_clauses.append("m.visit_time >= ?")
            params.append(date_from)
        
        if date_to:
            where_clauses.append("m.visit_time <= ?")
            params.append(date_to + ' 23:59:59')
        
        if phone:
            where_clauses.append("COALESCE(m.phone, '') = ?")
            params.append(phone)
        
        where_sql = " AND ".join(where_clauses)
        
        count_sql = f"""
            SELECT COUNT(*) as total
            FROM mart_customer_visit m
            WHERE {where_sql}
        """
        
        total_result = conn.execute(count_sql, params).fetchone()
        total = total_result[0] if total_result else 0
        
        offset = (page - 1) * page_size
        
        sql = f"""
            SELECT 
                COALESCE(m.region, '') as 大区,
                COALESCE(m.zone, '') as 战区,
                m.dealer_id,
                COALESCE(m.dealer_short_name, '') as 店简称,
                m.lead_id,
                COALESCE(m.channel_1, '') as channel_1,
                COALESCE(m.channel_2, '') as channel_2,
                COALESCE(m.channel_3, '') as channel_3,
                COALESCE(m.channel_4, '') as channel_4,
                m.visit_time,
                COALESCE(m.follower_id, '') as follower_id,
                COALESCE(m.follower_name, '') as follower_name,
                COALESCE(m.follower_position, '') as follower_position,
                m.followup_created_time,
                COALESCE(m.phone, '') as phone
            FROM mart_customer_visit m
            WHERE {where_sql}
            ORDER BY m.visit_time DESC
            LIMIT {page_size} OFFSET {offset}
        """
        
        results = conn.execute(sql, params).fetchall()
        
        data = []
        for row in results:
            data.append({
                '大区': row[0] or '',
                '战区': row[1] or '',
                '店编号': row[2],
                '店简称': row[3] or '',
                '门店线索id': row[4],
                '一级渠道': row[5] or '',
                '二级渠道': row[6] or '',
                '三级渠道': row[7] or '',
                '四级渠道': row[8] or '',
                '客户进店时间': str(row[9]) if row[9] else '',
                '跟进人': row[10] or '',
                '顾问姓名': row[11] or '',
                '顾问岗位': row[12] or '',
                '创建时间': str(row[13]) if row[13] else '',
                '手机号': row[14] or ''
            })
        
        return jsonify({
            'success': True,
            'data': data,
            'pagination': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size
            }
        })
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/customer_visit/stats', methods=['GET'])
def get_customer_visit_stats():
    try:
        conn = duck_db.get_connection()
        sql = """
            SELECT 
                COUNT(*) as total_visits,
                COUNT(DISTINCT m.lead_id) as unique_leads,
                COUNT(DISTINCT m.dealer_id) as dealer_count,
                COUNT(DISTINCT m.follower_name) as consultant_count
            FROM mart_customer_visit m
        """
        
        result = conn.execute(sql).fetchone()
        
        return jsonify({
            'success': True,
            'data': {
                'total_visits': result[0] if result else 0,
                'unique_leads': result[1] if result else 0,
                'dealer_count': result[2] if result else 0,
                'consultant_count': result[3] if result else 0
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


def _outbound_call_filters():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    global_keyword = request.args.get('global_keyword', '').strip()
    call_person = request.args.get('call_person', '').strip()
    call_number = request.args.get('call_number', '').strip()
    dealer = request.args.get('dealer', '').strip()
    answer_group = request.args.get('answer_group', '').strip()
    min_duration = request.args.get('min_duration', '').strip()
    max_duration = request.args.get('max_duration', '').strip()

    where_clauses = ["m.call_start_time IS NOT NULL"]
    params = []

    if date_from:
        where_clauses.append("m.call_start_time >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("m.call_start_time <= ?")
        params.append(date_to + ' 23:59:59')
    if global_keyword:
        keyword = f"%{global_keyword}%"
        where_clauses.append("""(
            COALESCE(m.region, '') LIKE ?
            OR COALESCE(m.zone, '') LIKE ?
            OR COALESCE(m.dealer_id, '') LIKE ?
            OR COALESCE(m.dealer_name, '') LIKE ?
            OR COALESCE(m.staff_id, '') LIKE ?
            OR COALESCE(m.consultant_name, '') LIKE ?
            OR COALESCE(m.consultant_role, '') LIKE ?
            OR COALESCE(m.seat_id, '') LIKE ?
            OR COALESCE(m.seat_phone, '') LIKE ?
            OR COALESCE(m.caller_name, '') LIKE ?
            OR COALESCE(m.call_number, '') LIKE ?
            OR COALESCE(m.answer_status, '') LIKE ?
            OR COALESCE(m.queue_name, '') LIKE ?
        )""")
        params.extend([keyword] * 13)
    if call_person:
        keyword = f"%{call_person}%"
        where_clauses.append("""(
            COALESCE(m.caller_name, '') LIKE ?
            OR COALESCE(m.consultant_name, '') LIKE ?
            OR COALESCE(m.staff_id, '') LIKE ?
            OR COALESCE(m.seat_id, '') LIKE ?
            OR COALESCE(m.seat_phone, '') LIKE ?
        )""")
        params.extend([keyword] * 5)
    if call_number:
        where_clauses.append("COALESCE(m.call_number, '') LIKE ?")
        params.append(f"%{call_number}%")
    if dealer:
        keyword = f"%{dealer}%"
        where_clauses.append("""(
            COALESCE(m.dealer_id, '') LIKE ?
            OR COALESCE(m.dealer_name, '') LIKE ?
            OR COALESCE(m.region, '') LIKE ?
            OR COALESCE(m.zone, '') LIKE ?
        )""")
        params.extend([keyword] * 4)
    if answer_group == 'answered':
        where_clauses.append("m.answer_group = '接通'")
    elif answer_group == 'unanswered':
        where_clauses.append("m.answer_group = '未接听'")
    if min_duration:
        where_clauses.append("m.talk_duration_sec >= ?")
        params.append(int(min_duration))
    if max_duration:
        where_clauses.append("m.talk_duration_sec <= ?")
        params.append(int(max_duration))

    return " AND ".join(where_clauses), params


OUTBOUND_DETAIL_EXPORT_MAX_ROWS = 50000
OUTBOUND_DETAIL_EXPORT_MAX_DAYS = 7


def _outbound_detail_export_guard(conn, where_sql, params):
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    if date_from and date_to:
        try:
            start_date = datetime.strptime(date_from[:10], '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to[:10], '%Y-%m-%d').date()
            if (end_date - start_date).days + 1 > OUTBOUND_DETAIL_EXPORT_MAX_DAYS:
                return (
                    jsonify({
                        'success': False,
                        'message': f'外呼明细导出日期跨度不能超过 {OUTBOUND_DETAIL_EXPORT_MAX_DAYS} 天，请缩小外呼日期范围后重试。'
                    }),
                    400
                )
        except ValueError:
            return jsonify({'success': False, 'message': '外呼开始或结束日期格式不正确'}), 400

    total = conn.execute(f"""
        SELECT COUNT(*)
        FROM mart_outbound_call_detail m
        WHERE {where_sql}
    """, params).fetchone()[0]
    if total > OUTBOUND_DETAIL_EXPORT_MAX_ROWS:
        return (
            jsonify({
                'success': False,
                'message': f'当前筛选结果共 {total} 条，超过单次导出上限 {OUTBOUND_DETAIL_EXPORT_MAX_ROWS} 条，请缩小日期范围或增加门店、顾问、外呼号码等筛选条件后重试。'
            }),
            400
        )
    return None


def _outbound_stats_filters(alias='m', default_call_type='outbound'):
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    region = request.args.get('region', '').strip()
    zone = request.args.get('zone', '').strip()
    dealer = request.args.get('dealer', '').strip()
    call_person = request.args.get('call_person', '').strip()
    answer_group = request.args.get('answer_group', '').strip()
    call_type = request.args.get('call_type', default_call_type).strip()

    where_clauses = [f"{alias}.call_start_time IS NOT NULL"]
    params = []

    if date_from:
        where_clauses.append(f"{alias}.call_start_time >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append(f"{alias}.call_start_time <= ?")
        params.append(date_to + ' 23:59:59')
    if region:
        where_clauses.append(f"COALESCE({alias}.region, '') = ?")
        params.append(region)
    if zone:
        where_clauses.append(f"COALESCE({alias}.zone, '') = ?")
        params.append(zone)
    if dealer:
        keyword = f"%{dealer}%"
        where_clauses.append(f"""(
            COALESCE({alias}.dealer_id, '') LIKE ?
            OR COALESCE({alias}.dealer_name, '') LIKE ?
        )""")
        params.extend([keyword, keyword])
    if call_person:
        keyword = f"%{call_person}%"
        where_clauses.append(f"""(
            COALESCE({alias}.caller_name, '') LIKE ?
            OR COALESCE({alias}.consultant_name, '') LIKE ?
            OR COALESCE({alias}.staff_id, '') LIKE ?
            OR COALESCE({alias}.seat_id, '') LIKE ?
            OR COALESCE({alias}.seat_phone, '') LIKE ?
        )""")
        params.extend([keyword] * 5)
    if answer_group == 'answered':
        where_clauses.append(f"{alias}.answer_group = '接通'")
    elif answer_group == 'unanswered':
        where_clauses.append(f"{alias}.answer_group = '未接听'")
    if call_type == 'inbound':
        where_clauses.append(f"COALESCE({alias}.call_type, '') LIKE '%呼入%'")
    elif call_type == 'all':
        pass
    else:
        where_clauses.append(f"COALESCE({alias}.call_type, '') NOT LIKE '%呼入%'")

    return " AND ".join(where_clauses), params


def _outbound_stats_daily_filters(alias='m', default_call_type='outbound'):
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    region = request.args.get('region', '').strip()
    zone = request.args.get('zone', '').strip()
    dealer = request.args.get('dealer', '').strip()
    call_person = request.args.get('call_person', '').strip()
    call_type = request.args.get('call_type', default_call_type).strip()

    where_clauses = [f"{alias}.stat_date IS NOT NULL"]
    params = []

    if date_from:
        where_clauses.append(f"{alias}.stat_date >= CAST(? AS DATE)")
        params.append(date_from)
    if date_to:
        where_clauses.append(f"{alias}.stat_date <= CAST(? AS DATE)")
        params.append(date_to)
    if region:
        where_clauses.append(f"COALESCE({alias}.region, '') = ?")
        params.append(region)
    if zone:
        where_clauses.append(f"COALESCE({alias}.zone, '') = ?")
        params.append(zone)
    if dealer:
        keyword = f"%{dealer}%"
        where_clauses.append(f"""(
            COALESCE({alias}.dealer_id, '') LIKE ?
            OR COALESCE({alias}.dealer_name, '') LIKE ?
        )""")
        params.extend([keyword, keyword])
    if call_person:
        keyword = f"%{call_person}%"
        where_clauses.append(f"""(
            COALESCE({alias}.consultant_name, '') LIKE ?
            OR COALESCE({alias}.staff_id, '') LIKE ?
            OR COALESCE({alias}.seat_id, '') LIKE ?
            OR COALESCE({alias}.seat_phone, '') LIKE ?
        )""")
        params.extend([keyword] * 4)
    if call_type == 'inbound':
        where_clauses.append(f"COALESCE({alias}.call_type_group, '') = '呼入'")
    elif call_type == 'all':
        pass
    else:
        where_clauses.append(f"COALESCE({alias}.call_type_group, '') = '外呼'")

    return " AND ".join(where_clauses), params


def _outbound_stats_has_answer_filter():
    return request.args.get('answer_group', '').strip() in ('answered', 'unanswered')


def _outbound_dealer_filters(alias='d'):
    region = request.args.get('region', '').strip()
    zone = request.args.get('zone', '').strip()
    dealer = request.args.get('dealer', '').strip()
    where_clauses = ["1=1"]
    params = []
    if region:
        where_clauses.append(f"COALESCE({alias}.region, '') = ?")
        params.append(region)
    if zone:
        where_clauses.append(f"COALESCE({alias}.zone, '') = ?")
        params.append(zone)
    if dealer:
        keyword = f"%{dealer}%"
        where_clauses.append(f"""(
            COALESCE({alias}.dealer_id, '') LIKE ?
            OR COALESCE({alias}.dealer_name, '') LIKE ?
        )""")
        params.extend([keyword, keyword])
    return " AND ".join(where_clauses), params


def _rate(numerator, denominator):
    return round(float(numerator or 0) * 100.0 / float(denominator or 0), 1) if denominator else 0


def _format_duration(seconds):
    try:
        seconds = max(int(seconds or 0), 0)
    except (TypeError, ValueError):
        seconds = 0
    minutes, rest = divmod(seconds, 60)
    if minutes:
        return f"{minutes}分{rest:02d}秒"
    return f"{rest}秒"


def _outbound_stats_dict(row):
    total = row.get('total_calls') or 0
    answered = row.get('answered_calls') or 0
    effective = row.get('effective_calls') or 0
    return {
        **dict(row),
        'answered_rate': _rate(answered, total),
        'effective_rate': _rate(effective, total),
        'effective_30s_rate': _rate(row.get('effective_30s_calls') or 0, effective),
        'effective_60s_rate': _rate(row.get('effective_60s_calls') or 0, effective),
        'avg_talk_duration_text': _format_duration(row.get('avg_talk_duration_sec') or 0),
        'recording_rate': _rate(row.get('recording_calls') or 0, total),
        'short_talk_rate': _rate(row.get('short_talk_calls') or 0, answered),
    }


def _outbound_stats_zero_row(extra=None):
    data = {
        'total_calls': 0,
        'unique_numbers': 0,
        'answered_calls': 0,
        'effective_calls': 0,
        'effective_30s_calls': 0,
        'effective_60s_calls': 0,
        'total_talk_duration_sec': 0,
        'avg_talk_duration_sec': 0,
        'active_staff_count': 0,
        'recording_calls': 0,
        'no_recording_calls': 0,
        'short_talk_calls': 0,
        'unmatched_staff_calls': 0,
        'per_staff_calls': 0,
        'high_freq_unanswered_numbers': 0,
    }
    if extra:
        data.update(extra)
    return data


def _outbound_row_dict(row):
    return {
        '大区': row[0] or '',
        '战区': row[1] or '',
        '门店编码': row[2] or '',
        '门店名称': row[3] or '',
        '顾问ID': row[4] or '',
        '顾问姓名': row[5] or '',
        '顾问岗位': row[6] or '',
        '座席工号': row[7] or '',
        '座席电话': row[8] or '',
        '外呼人': row[9] or '',
        '开始时间': row[10].strftime('%Y-%m-%d %H:%M:%S') if row[10] else '',
        '结束时间': row[11].strftime('%Y-%m-%d %H:%M:%S') if row[11] else '',
        '外呼轮次': row[12] or 0,
        '外呼号码': row[13] or '',
        '接听状态': row[14] or '',
        '通话时长': row[15] or 0,
        '振铃时长': row[16] or 0,
        '是否有录音': '是' if row[17] else '否',
        '通话时长文本': _format_duration(row[15]),
        '振铃时长文本': _format_duration(row[16]),
    }


def _outbound_call_available_range():
    duck_db.ensure_outbound_call_detail()
    conn = duck_db.get_connection()
    row = conn.execute("""
        SELECT MIN(CAST(call_start_time AS DATE)), MAX(CAST(call_start_time AS DATE))
        FROM mart_outbound_call_detail
        WHERE call_start_time IS NOT NULL
    """).fetchone()
    min_date = row[0].strftime('%Y-%m-%d') if row and row[0] else ''
    max_date = row[1].strftime('%Y-%m-%d') if row and row[1] else ''
    default_from = ''
    default_to = max_date
    if row and row[1]:
        latest = row[1]
        default_from = latest.replace(day=1).strftime('%Y-%m-%d')
    return {
        'min_date': min_date,
        'max_date': max_date,
        'default_date_from': default_from,
        'default_date_to': default_to,
    }


def _outbound_stats_summary_data():
    duck_db.ensure_outbound_call_stats()
    if _outbound_stats_has_answer_filter():
        where_sql, params = _outbound_stats_filters()
        conn = duck_db.get_connection()
        row = conn.execute(f"""
            SELECT
                COUNT(*) AS total_calls,
                COUNT(DISTINCT NULLIF(m.call_number, '')) AS unique_numbers,
                SUM(CASE WHEN m.answer_group = '接通' THEN 1 ELSE 0 END) AS answered_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 10 THEN 1 ELSE 0 END) AS effective_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 30 THEN 1 ELSE 0 END) AS effective_30s_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 60 THEN 1 ELSE 0 END) AS effective_60s_calls,
                SUM(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE 0 END) AS total_talk_duration_sec,
                AVG(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE NULL END) AS avg_talk_duration_sec,
                COUNT(DISTINCT CASE WHEN COALESCE(m.staff_id, '') != '' THEN m.staff_id ELSE NULL END) AS active_staff_count,
                SUM(CASE WHEN m.has_recording THEN 1 ELSE 0 END) AS recording_calls,
                SUM(CASE WHEN m.has_recording THEN 0 ELSE 1 END) AS no_recording_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec < 10 THEN 1 ELSE 0 END) AS short_talk_calls,
                SUM(CASE WHEN COALESCE(m.staff_match_type, '') = 'unmatched' THEN 1 ELSE 0 END) AS unmatched_staff_calls
            FROM mart_outbound_call_detail m
            WHERE {where_sql}
        """, params).fetchone()
    else:
        where_sql, params = _outbound_stats_daily_filters()
        conn = duck_db.get_connection()
        row = conn.execute(f"""
            SELECT
                SUM(m.total_calls) AS total_calls,
                0 AS unique_numbers,
                SUM(m.answered_calls) AS answered_calls,
                SUM(m.effective_calls) AS effective_calls,
                SUM(m.effective_30s_calls) AS effective_30s_calls,
                SUM(m.effective_60s_calls) AS effective_60s_calls,
                SUM(m.total_talk_duration_sec) AS total_talk_duration_sec,
                CASE WHEN SUM(m.answered_calls) > 0 THEN SUM(m.total_talk_duration_sec) * 1.0 / SUM(m.answered_calls) ELSE 0 END AS avg_talk_duration_sec,
                COUNT(DISTINCT CASE WHEN m.total_calls > 0 AND COALESCE(m.staff_id, '') != '' THEN m.staff_id ELSE NULL END) AS active_staff_count,
                SUM(m.recording_calls) AS recording_calls,
                SUM(m.no_recording_calls) AS no_recording_calls,
                SUM(m.short_talk_calls) AS short_talk_calls,
                SUM(m.unmatched_staff_calls) AS unmatched_staff_calls
            FROM mart_outbound_call_stats_daily m
            WHERE {where_sql}
        """, params).fetchone()
        detail_where_sql, detail_params = _outbound_stats_filters('m')
        unique_numbers = conn.execute(f"""
            SELECT COUNT(DISTINCT NULLIF(m.call_number, ''))
            FROM mart_outbound_call_detail m
            WHERE {detail_where_sql}
              AND NULLIF(m.call_number, '') IS NOT NULL
        """, detail_params).fetchone()[0]
    keys = [
        'total_calls', 'unique_numbers', 'answered_calls', 'effective_calls',
        'effective_30s_calls', 'effective_60s_calls', 'total_talk_duration_sec',
        'avg_talk_duration_sec', 'active_staff_count', 'recording_calls',
        'no_recording_calls', 'short_talk_calls', 'unmatched_staff_calls'
    ]
    data = dict(zip(keys, row or [0] * len(keys)))
    if not _outbound_stats_has_answer_filter():
        data['unique_numbers'] = unique_numbers
    data['avg_talk_duration_sec'] = int(round(data.get('avg_talk_duration_sec') or 0))
    data['per_staff_calls'] = round((data.get('total_calls') or 0) / data['active_staff_count'], 1) if data.get('active_staff_count') else 0
    data['high_freq_unanswered_numbers'] = _outbound_stats_high_freq_count(conn)
    return _outbound_stats_dict(data)


def _outbound_stats_high_freq_count(conn=None):
    conn = conn or duck_db.get_connection()
    where_sql, params = _outbound_stats_filters('m')
    return conn.execute(f"""
        SELECT COUNT(*)
        FROM (
            SELECT m.dealer_id, m.call_number
            FROM mart_outbound_call_detail m
            WHERE {where_sql}
              AND NULLIF(m.call_number, '') IS NOT NULL
            GROUP BY m.dealer_id, m.call_number
            HAVING COUNT(*) >= 5
               AND SUM(CASE WHEN m.answer_group = '接通' THEN 1 ELSE 0 END) = 0
        )
    """, params).fetchone()[0]


def _outbound_stats_store_rows():
    duck_db.ensure_outbound_call_stats()
    use_detail = _outbound_stats_has_answer_filter()
    where_sql, params = (_outbound_stats_filters('m') if use_detail else _outbound_stats_daily_filters('m'))
    detail_where_sql, detail_params = _outbound_stats_filters('m')
    dealer_sql, dealer_params = _outbound_dealer_filters('d')
    conn = duck_db.get_connection()
    source_table = 'mart_outbound_call_detail' if use_detail else 'mart_outbound_call_stats_daily'
    total_expr = "COUNT(*)" if use_detail else "SUM(m.total_calls)"
    unique_expr = "COUNT(DISTINCT NULLIF(m.call_number, ''))" if use_detail else "0"
    answered_expr = "SUM(CASE WHEN m.answer_group = '接通' THEN 1 ELSE 0 END)" if use_detail else "SUM(m.answered_calls)"
    effective_expr = "SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 10 THEN 1 ELSE 0 END)" if use_detail else "SUM(m.effective_calls)"
    effective_30s_expr = "SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 30 THEN 1 ELSE 0 END)" if use_detail else "SUM(m.effective_30s_calls)"
    effective_60s_expr = "SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 60 THEN 1 ELSE 0 END)" if use_detail else "SUM(m.effective_60s_calls)"
    total_talk_expr = "SUM(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE 0 END)" if use_detail else "SUM(m.total_talk_duration_sec)"
    avg_talk_expr = "AVG(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE NULL END)" if use_detail else "CASE WHEN SUM(m.answered_calls) > 0 THEN SUM(m.total_talk_duration_sec) * 1.0 / SUM(m.answered_calls) ELSE 0 END"
    active_staff_expr = "COUNT(DISTINCT CASE WHEN COALESCE(m.staff_id, '') != '' THEN m.staff_id ELSE NULL END)" if use_detail else "COUNT(DISTINCT CASE WHEN m.total_calls > 0 AND COALESCE(m.staff_id, '') != '' THEN m.staff_id ELSE NULL END)"
    recording_expr = "SUM(CASE WHEN m.has_recording THEN 1 ELSE 0 END)" if use_detail else "SUM(m.recording_calls)"
    no_recording_expr = "SUM(CASE WHEN m.has_recording THEN 0 ELSE 1 END)" if use_detail else "SUM(m.no_recording_calls)"
    short_talk_expr = "SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec < 10 THEN 1 ELSE 0 END)" if use_detail else "SUM(m.short_talk_calls)"
    unmatched_expr = "SUM(CASE WHEN COALESCE(m.staff_match_type, '') = 'unmatched' THEN 1 ELSE 0 END)" if use_detail else "SUM(m.unmatched_staff_calls)"
    rows = conn.execute(f"""
        WITH stats AS (
            SELECT
                m.dealer_id,
                {total_expr} AS total_calls,
                {unique_expr} AS unique_numbers,
                {answered_expr} AS answered_calls,
                {effective_expr} AS effective_calls,
                {effective_30s_expr} AS effective_30s_calls,
                {effective_60s_expr} AS effective_60s_calls,
                {total_talk_expr} AS total_talk_duration_sec,
                {avg_talk_expr} AS avg_talk_duration_sec,
                {active_staff_expr} AS active_staff_count,
                {recording_expr} AS recording_calls,
                {no_recording_expr} AS no_recording_calls,
                {short_talk_expr} AS short_talk_calls,
                {unmatched_expr} AS unmatched_staff_calls
            FROM {source_table} m
            WHERE {where_sql}
            GROUP BY m.dealer_id
        ),
        unique_numbers AS (
            SELECT m.dealer_id, COUNT(DISTINCT NULLIF(m.call_number, '')) AS unique_numbers
            FROM mart_outbound_call_detail m
            WHERE {detail_where_sql}
            GROUP BY m.dealer_id
        ),
        rounds AS (
            SELECT dealer_id, AVG(call_count) AS avg_call_round
            FROM (
                SELECT m.dealer_id, m.call_number, COUNT(*) AS call_count
                FROM mart_outbound_call_detail m
                WHERE {detail_where_sql}
                  AND NULLIF(m.call_number, '') IS NOT NULL
                GROUP BY m.dealer_id, m.call_number
            )
            GROUP BY dealer_id
        ),
        risk AS (
            SELECT dealer_id, COUNT(*) AS high_freq_unanswered_numbers
            FROM (
                SELECT m.dealer_id, m.call_number
                FROM mart_outbound_call_detail m
                WHERE {detail_where_sql}
                  AND NULLIF(m.call_number, '') IS NOT NULL
                GROUP BY m.dealer_id, m.call_number
                HAVING COUNT(*) >= 5 AND SUM(CASE WHEN m.answer_group = '接通' THEN 1 ELSE 0 END) = 0
            )
            GROUP BY dealer_id
        )
        SELECT
            COALESCE(d.region, '') AS region,
            COALESCE(d.zone, '') AS zone,
            COALESCE(d.dealer_id, '') AS dealer_id,
            COALESCE(d.dealer_name, '') AS dealer_name,
            COALESCE(s.total_calls, 0) AS total_calls,
            COALESCE(u.unique_numbers, s.unique_numbers, 0) AS unique_numbers,
            COALESCE(s.answered_calls, 0) AS answered_calls,
            COALESCE(s.effective_calls, 0) AS effective_calls,
            COALESCE(s.effective_30s_calls, 0) AS effective_30s_calls,
            COALESCE(s.effective_60s_calls, 0) AS effective_60s_calls,
            COALESCE(s.total_talk_duration_sec, 0) AS total_talk_duration_sec,
            COALESCE(s.avg_talk_duration_sec, 0) AS avg_talk_duration_sec,
            COALESCE(s.active_staff_count, 0) AS active_staff_count,
            CASE WHEN COALESCE(s.active_staff_count, 0) > 0 THEN COALESCE(s.total_calls, 0) * 1.0 / s.active_staff_count ELSE 0 END AS per_staff_calls,
            COALESCE(r.avg_call_round, 0) AS avg_call_round,
            COALESCE(k.high_freq_unanswered_numbers, 0) AS high_freq_unanswered_numbers,
            COALESCE(s.recording_calls, 0) AS recording_calls,
            COALESCE(s.no_recording_calls, 0) AS no_recording_calls,
            COALESCE(s.short_talk_calls, 0) AS short_talk_calls,
            COALESCE(s.unmatched_staff_calls, 0) AS unmatched_staff_calls
        FROM mart_dealers d
        LEFT JOIN stats s ON d.dealer_id = s.dealer_id
        LEFT JOIN unique_numbers u ON d.dealer_id = u.dealer_id
        LEFT JOIN rounds r ON d.dealer_id = r.dealer_id
        LEFT JOIN risk k ON d.dealer_id = k.dealer_id
        WHERE {dealer_sql}
        ORDER BY COALESCE(s.total_calls, 0) DESC, d.dealer_id
    """, params + detail_params + detail_params + detail_params + dealer_params).fetchall()
    columns = [desc[0] for desc in conn.description]
    data = []
    for row in rows:
        item = dict(zip(columns, row))
        item['avg_talk_duration_sec'] = int(round(item.get('avg_talk_duration_sec') or 0))
        item['per_staff_calls'] = round(item.get('per_staff_calls') or 0, 1)
        item['avg_call_round'] = round(item.get('avg_call_round') or 0, 1)
        data.append(_outbound_stats_dict(item))
    return data


@app.route('/api/outbound-call/options', methods=['GET'])
def get_outbound_call_options():
    try:
        return jsonify({'success': True, 'data': _outbound_call_available_range()})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call/summary', methods=['GET'])
def get_outbound_call_summary():
    try:
        duck_db.ensure_outbound_call_detail()
        where_sql, params = _outbound_call_filters()
        conn = duck_db.get_connection()
        result = conn.execute(f"""
            SELECT
                COUNT(*) AS total_calls,
                SUM(CASE WHEN answer_group = '接通' THEN 1 ELSE 0 END) AS answered_calls,
                AVG(CASE WHEN answer_group = '接通' THEN talk_duration_sec ELSE NULL END) AS avg_talk_duration_sec
            FROM mart_outbound_call_detail m
            WHERE {where_sql}
        """, params).fetchone()

        total = result[0] if result else 0
        answered = result[1] if result and result[1] is not None else 0
        avg_talk = int(round(result[2] or 0)) if result else 0

        return jsonify({
            'success': True,
            'data': {
                'total_calls': total,
                'answered_calls': answered,
                'answered_rate': round(answered * 100.0 / total, 1) if total else 0,
                'avg_talk_duration_sec': avg_talk,
                'avg_talk_duration_text': _format_duration(avg_talk),
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call/detail', methods=['GET'])
def get_outbound_call_detail():
    try:
        duck_db.ensure_outbound_call_detail()
        page = max(int(request.args.get('page', 1)), 1)
        page_size = min(max(int(request.args.get('page_size', 100)), 1), 500)
        where_sql, params = _outbound_call_filters()
        conn = duck_db.get_connection()

        total = conn.execute(f"""
            SELECT COUNT(*)
            FROM mart_outbound_call_detail m
            WHERE {where_sql}
        """, params).fetchone()[0]

        offset = (page - 1) * page_size
        rows = conn.execute(f"""
            SELECT
                COALESCE(m.region, '') AS region,
                COALESCE(m.zone, '') AS zone,
                COALESCE(m.dealer_id, '') AS dealer_id,
                COALESCE(m.dealer_name, '') AS dealer_name,
                COALESCE(m.staff_id, '') AS staff_id,
                COALESCE(m.consultant_name, '') AS consultant_name,
                COALESCE(m.consultant_role, '') AS consultant_role,
                COALESCE(m.seat_id, '') AS seat_id,
                COALESCE(m.seat_phone, '') AS seat_phone,
                COALESCE(m.caller_name, '') AS caller_name,
                m.call_start_time,
                m.call_end_time,
                m.call_round,
                COALESCE(m.call_number, '') AS call_number,
                COALESCE(m.answer_status, '') AS answer_status,
                m.talk_duration_sec,
                m.ring_duration_sec,
                m.has_recording
            FROM mart_outbound_call_detail m
            WHERE {where_sql}
            ORDER BY m.call_start_time DESC, m.outbound_call_id DESC
            LIMIT {page_size} OFFSET {offset}
        """, params).fetchall()

        return jsonify({
            'success': True,
            'data': [_outbound_row_dict(row) for row in rows],
            'pagination': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'total_pages': (total + page_size - 1) // page_size
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call/export', methods=['GET'])
def export_outbound_call_detail():
    try:
        import openpyxl

        duck_db.ensure_outbound_call_detail()
        where_sql, params = _outbound_call_filters()
        conn = duck_db.get_connection()
        guard_response = _outbound_detail_export_guard(conn, where_sql, params)
        if guard_response:
            return guard_response

        rows = conn.execute(f"""
            SELECT
                COALESCE(m.region, '') AS region,
                COALESCE(m.zone, '') AS zone,
                COALESCE(m.dealer_id, '') AS dealer_id,
                COALESCE(m.dealer_name, '') AS dealer_name,
                COALESCE(m.staff_id, '') AS staff_id,
                COALESCE(m.consultant_name, '') AS consultant_name,
                COALESCE(m.consultant_role, '') AS consultant_role,
                COALESCE(m.seat_id, '') AS seat_id,
                COALESCE(m.seat_phone, '') AS seat_phone,
                COALESCE(m.caller_name, '') AS caller_name,
                m.call_start_time,
                m.call_end_time,
                m.call_round,
                COALESCE(m.call_number, '') AS call_number,
                COALESCE(m.answer_status, '') AS answer_status,
                m.talk_duration_sec,
                m.ring_duration_sec,
                m.has_recording
            FROM mart_outbound_call_detail m
            WHERE {where_sql}
            ORDER BY m.call_start_time DESC, m.outbound_call_id DESC
        """, params)

        columns = [
            '大区', '战区', '门店编码', '门店名称',
            '顾问ID', '顾问姓名', '顾问岗位',
            '座席工号', '座席电话', '外呼人',
            '开始时间', '结束时间', '外呼轮次', '外呼号码',
            '接听状态', '通话时长', '振铃时长', '是否有录音'
        ]
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet('外呼明细')
        ws.title = '外呼明细'
        ws.append(columns)
        while True:
            batch = rows.fetchmany(5000)
            if not batch:
                break
            for row in batch:
                item = _outbound_row_dict(row)
                ws.append([
                    item['大区'], item['战区'], item['门店编码'], item['门店名称'],
                    item['顾问ID'], item['顾问姓名'], item['顾问岗位'],
                    item['座席工号'], item['座席电话'], item['外呼人'],
                    item['开始时间'], item['结束时间'], item['外呼轮次'], item['外呼号码'],
                    item['接听状态'], item['通话时长文本'], item['振铃时长文本'], item['是否有录音']
                ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"外呼明细_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/options', methods=['GET'])
def get_outbound_call_stats_options():
    try:
        duck_db.ensure_outbound_call_stats()
        conn = duck_db.get_connection()
        date_range = _outbound_call_available_range()
        regions = [row[0] for row in conn.execute("""
            SELECT DISTINCT region FROM mart_dealers
            WHERE COALESCE(region, '') != ''
            ORDER BY region
        """).fetchall()]
        zones = [row[0] for row in conn.execute("""
            SELECT DISTINCT zone FROM mart_dealers
            WHERE COALESCE(zone, '') != ''
            ORDER BY zone
        """).fetchall()]
        return jsonify({'success': True, 'data': {'regions': regions, 'zones': zones, **date_range}})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/summary', methods=['GET'])
def get_outbound_call_stats_summary():
    try:
        return jsonify({'success': True, 'data': _outbound_stats_summary_data()})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/trend', methods=['GET'])
def get_outbound_call_stats_trend():
    try:
        duck_db.ensure_outbound_call_stats()
        conn = duck_db.get_connection()
        if _outbound_stats_has_answer_filter():
            where_sql, params = _outbound_stats_filters()
            rows = conn.execute(f"""
                SELECT
                    CAST(m.call_start_time AS DATE) AS stat_date,
                    COUNT(*) AS total_calls,
                    COUNT(DISTINCT NULLIF(m.call_number, '')) AS unique_numbers,
                    SUM(CASE WHEN m.answer_group = '接通' THEN 1 ELSE 0 END) AS answered_calls,
                    SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 10 THEN 1 ELSE 0 END) AS effective_calls,
                    SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 30 THEN 1 ELSE 0 END) AS effective_30s_calls,
                    SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 60 THEN 1 ELSE 0 END) AS effective_60s_calls,
                    SUM(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE 0 END) AS total_talk_duration_sec,
                    AVG(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE NULL END) AS avg_talk_duration_sec,
                    SUM(CASE WHEN m.has_recording THEN 1 ELSE 0 END) AS recording_calls,
                    SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec < 10 THEN 1 ELSE 0 END) AS short_talk_calls
                FROM mart_outbound_call_detail m
                WHERE {where_sql}
                GROUP BY stat_date
                ORDER BY stat_date
            """, params).fetchall()
        else:
            where_sql, params = _outbound_stats_daily_filters()
            rows = conn.execute(f"""
                SELECT
                    m.stat_date,
                    SUM(m.total_calls) AS total_calls,
                    0 AS unique_numbers,
                    SUM(m.answered_calls) AS answered_calls,
                    SUM(m.effective_calls) AS effective_calls,
                    SUM(m.effective_30s_calls) AS effective_30s_calls,
                    SUM(m.effective_60s_calls) AS effective_60s_calls,
                    SUM(m.total_talk_duration_sec) AS total_talk_duration_sec,
                    CASE WHEN SUM(m.answered_calls) > 0 THEN SUM(m.total_talk_duration_sec) * 1.0 / SUM(m.answered_calls) ELSE 0 END AS avg_talk_duration_sec,
                    SUM(m.recording_calls) AS recording_calls,
                    SUM(m.short_talk_calls) AS short_talk_calls
                FROM mart_outbound_call_stats_daily m
                WHERE {where_sql}
                GROUP BY m.stat_date
                ORDER BY m.stat_date
            """, params).fetchall()
        columns = [desc[0] for desc in conn.description]
        data = []
        for row in rows:
            item = dict(zip(columns, row))
            item['stat_date'] = item['stat_date'].strftime('%Y-%m-%d') if item.get('stat_date') else ''
            item['avg_talk_duration_sec'] = int(round(item.get('avg_talk_duration_sec') or 0))
            data.append(_outbound_stats_dict(item))
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/stores', methods=['GET'])
def get_outbound_call_stats_stores():
    try:
        return jsonify({'success': True, 'data': _outbound_stats_store_rows()})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/consultants', methods=['GET'])
def get_outbound_call_stats_consultants():
    try:
        duck_db.ensure_outbound_call_stats()
        where_sql, params = _outbound_stats_filters('m')
        conn = duck_db.get_connection()
        rows = conn.execute(f"""
            SELECT
                COALESCE(m.region, '') AS region,
                COALESCE(m.zone, '') AS zone,
                COALESCE(m.dealer_id, '') AS dealer_id,
                COALESCE(m.dealer_name, '') AS dealer_name,
                COALESCE(m.staff_id, '') AS staff_id,
                COALESCE(m.consultant_name, '') AS consultant_name,
                COALESCE(m.consultant_role, '') AS consultant_role,
                COALESCE(m.seat_id, '') AS seat_id,
                COALESCE(m.seat_phone, '') AS seat_phone,
                COUNT(*) AS total_calls,
                COUNT(DISTINCT NULLIF(m.call_number, '')) AS unique_numbers,
                SUM(CASE WHEN m.answer_group = '接通' THEN 1 ELSE 0 END) AS answered_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 10 THEN 1 ELSE 0 END) AS effective_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 30 THEN 1 ELSE 0 END) AS effective_30s_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec >= 60 THEN 1 ELSE 0 END) AS effective_60s_calls,
                SUM(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE 0 END) AS total_talk_duration_sec,
                AVG(CASE WHEN m.answer_group = '接通' THEN m.talk_duration_sec ELSE NULL END) AS avg_talk_duration_sec,
                SUM(CASE WHEN m.has_recording THEN 1 ELSE 0 END) AS recording_calls,
                SUM(CASE WHEN m.has_recording THEN 0 ELSE 1 END) AS no_recording_calls,
                SUM(CASE WHEN m.answer_group = '接通' AND m.talk_duration_sec < 10 THEN 1 ELSE 0 END) AS short_talk_calls,
                SUM(CASE WHEN COALESCE(m.staff_match_type, '') = 'unmatched' THEN 1 ELSE 0 END) AS unmatched_staff_calls
            FROM mart_outbound_call_detail m
            WHERE {where_sql}
            GROUP BY region, zone, dealer_id, dealer_name, staff_id, consultant_name, consultant_role, seat_id, seat_phone
            ORDER BY total_calls DESC
            LIMIT 300
        """, params).fetchall()
        columns = [desc[0] for desc in conn.description]
        data = []
        for row in rows:
            item = dict(zip(columns, row))
            item['avg_talk_duration_sec'] = int(round(item.get('avg_talk_duration_sec') or 0))
            data.append(_outbound_stats_dict(item))
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/numbers', methods=['GET'])
def get_outbound_call_stats_numbers():
    try:
        duck_db.ensure_outbound_call_stats()
        where_sql, params = _outbound_stats_filters('m')
        conn = duck_db.get_connection()
        rows = conn.execute(f"""
            WITH numbered AS (
                SELECT
                    m.*,
                    ROW_NUMBER() OVER (
                        PARTITION BY COALESCE(m.dealer_id, ''), COALESCE(m.call_number, '')
                        ORDER BY m.call_start_time DESC, m.outbound_call_id DESC
                    ) AS latest_rn
                FROM mart_outbound_call_detail m
                WHERE {where_sql}
                  AND NULLIF(m.call_number, '') IS NOT NULL
            )
            SELECT
                COALESCE(dealer_id, '') AS dealer_id,
                COALESCE(dealer_name, '') AS dealer_name,
                COALESCE(region, '') AS region,
                COALESCE(zone, '') AS zone,
                COALESCE(call_number, '') AS call_number,
                COUNT(*) AS total_calls,
                SUM(CASE WHEN answer_group = '接通' THEN 1 ELSE 0 END) AS answered_calls,
                SUM(CASE WHEN answer_group = '接通' AND talk_duration_sec >= 10 THEN 1 ELSE 0 END) AS effective_calls,
                SUM(CASE WHEN answer_group = '接通' AND talk_duration_sec >= 30 THEN 1 ELSE 0 END) AS effective_30s_calls,
                SUM(CASE WHEN answer_group = '接通' AND talk_duration_sec >= 60 THEN 1 ELSE 0 END) AS effective_60s_calls,
                MIN(call_start_time) AS first_call_time,
                MAX(call_start_time) AS latest_call_time,
                MIN(CASE WHEN answer_group = '接通' THEN call_answer_time ELSE NULL END) AS first_answer_time,
                MAX(talk_duration_sec) AS max_talk_duration_sec,
                COALESCE(MAX(CASE WHEN latest_rn = 1 THEN caller_name ELSE NULL END), '') AS latest_caller_name,
                COALESCE(MAX(CASE WHEN latest_rn = 1 THEN staff_id ELSE NULL END), '') AS latest_staff_id,
                COALESCE(MAX(CASE WHEN latest_rn = 1 THEN seat_id ELSE NULL END), '') AS latest_seat_id,
                COALESCE(MAX(CASE WHEN latest_rn = 1 THEN seat_phone ELSE NULL END), '') AS latest_seat_phone
            FROM numbered
            GROUP BY dealer_id, dealer_name, region, zone, call_number
            ORDER BY
                CASE WHEN COUNT(*) >= 5 AND SUM(CASE WHEN answer_group = '接通' THEN 1 ELSE 0 END) = 0 THEN 0 ELSE 1 END,
                total_calls DESC,
                latest_call_time DESC
            LIMIT 300
        """, params).fetchall()
        columns = [desc[0] for desc in conn.description]
        data = []
        for row in rows:
            item = dict(zip(columns, row))
            item['has_answered'] = (item.get('answered_calls') or 0) > 0
            item['is_high_freq_unanswered'] = (item.get('total_calls') or 0) >= 5 and not item['has_answered']
            item['risk_tag'] = '高频未接通' if item['is_high_freq_unanswered'] else ('多轮后接通' if (item.get('total_calls') or 0) >= 5 else '正常触达')
            for key in ('first_call_time', 'latest_call_time', 'first_answer_time'):
                item[key] = item[key].strftime('%Y-%m-%d %H:%M:%S') if item.get(key) else ''
            item['max_talk_duration_text'] = _format_duration(item.get('max_talk_duration_sec') or 0)
            data.append(_outbound_stats_dict(item))
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/quality', methods=['GET'])
def get_outbound_call_stats_quality():
    try:
        data = _outbound_stats_summary_data()
        low_store_count = 0
        try:
            duck_db.ensure_outbound_call_stats()
            where_sql, params = _outbound_stats_filters('m')
            conn = duck_db.get_connection()
            low_store_count = conn.execute(f"""
                SELECT COUNT(*)
                FROM (
                    SELECT dealer_id,
                           COUNT(*) AS total_calls,
                           SUM(CASE WHEN answer_group = '接通' THEN 1 ELSE 0 END) AS answered_calls
                    FROM mart_outbound_call_detail m
                    WHERE {where_sql}
                    GROUP BY dealer_id
                    HAVING COUNT(*) > 0
                       AND SUM(CASE WHEN answer_group = '接通' THEN 1 ELSE 0 END) * 1.0 / COUNT(*) < 0.3
                )
            """, params).fetchone()[0]
        except Exception:
            low_store_count = 0
        rows = [
            {'type': '高频未接通号码', 'count': data.get('high_freq_unanswered_numbers', 0), 'ratio': _rate(data.get('high_freq_unanswered_numbers', 0), data.get('unique_numbers', 0)), 'rule': '同门店同号码外呼次数 >= 5，且从未接通', 'priority': '高'},
            {'type': '短通话记录', 'count': data.get('short_talk_calls', 0), 'ratio': _rate(data.get('short_talk_calls', 0), data.get('answered_calls', 0)), 'rule': '接通后通话时长 < 10秒', 'priority': '中'},
            {'type': '无录音外呼', 'count': data.get('no_recording_calls', 0), 'ratio': _rate(data.get('no_recording_calls', 0), data.get('total_calls', 0)), 'rule': '录音字段为空', 'priority': '中'},
            {'type': '无匹配顾问记录', 'count': data.get('unmatched_staff_calls', 0), 'ratio': _rate(data.get('unmatched_staff_calls', 0), data.get('total_calls', 0)), 'rule': '座席工号、座席电话均无法匹配人员表', 'priority': '高'},
            {'type': '低接通门店', 'count': low_store_count, 'ratio': low_store_count, 'rule': '门店接通率低于 30%', 'priority': '高'},
        ]
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/outbound-call-stats/export', methods=['GET'])
def export_outbound_call_stats():
    try:
        import openpyxl

        stores = _outbound_stats_store_rows()
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet('门店外呼统计')
        ws.title = '门店外呼统计'
        headers = [
            '大区', '战区', '门店编码', '门店名称', '外呼总数', '去重号码数',
            '接通数', '接通率', '有效通话数', '有效通话率',
            '30s有效通话', '30s通话占比', '60s有效通话', '60s通话占比',
            '人均外呼数', '平均通话时长', '平均外呼轮次',
            '高频未接通号码数', '录音覆盖率'
        ]
        ws.append(headers)
        for item in stores:
            ws.append([
                item.get('region', ''),
                item.get('zone', ''),
                item.get('dealer_id', ''),
                item.get('dealer_name', ''),
                item.get('total_calls', 0),
                item.get('unique_numbers', 0),
                item.get('answered_calls', 0),
                item.get('answered_rate', 0),
                item.get('effective_calls', 0),
                item.get('effective_rate', 0),
                item.get('effective_30s_calls', 0),
                item.get('effective_30s_rate', 0),
                item.get('effective_60s_calls', 0),
                item.get('effective_60s_rate', 0),
                item.get('per_staff_calls', 0),
                item.get('avg_talk_duration_text', ''),
                item.get('avg_call_round', 0),
                item.get('high_freq_unanswered_numbers', 0),
                item.get('recording_rate', 0),
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"外呼统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/customer_visit/export', methods=['GET'])
def export_customer_visit():
    try:
        import openpyxl
        
        dealer_code = request.args.get('dealer_code', '')
        channel_1 = request.args.get('channel_1', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        phone = request.args.get('phone', '')
        
        conn = duck_db.get_connection()
        
        where_clauses = ["m.dealer_id IS NOT NULL"]
        params = []
        
        if dealer_code:
            where_clauses.append("m.dealer_id = ?")
            params.append(dealer_code)
        
        if channel_1:
            if channel_1 == '__EMPTY__':
                where_clauses.append("(m.channel_1 IS NULL OR m.channel_1 = '')")
            else:
                where_clauses.append("m.channel_1 = ?")
                params.append(channel_1)
        
        if date_from:
            where_clauses.append("m.visit_time >= ?")
            params.append(date_from)
        
        if date_to:
            where_clauses.append("m.visit_time <= ?")
            params.append(date_to + ' 23:59:59')
        
        if phone:
            where_clauses.append("COALESCE(m.phone, '') = ?")
            params.append(phone)
        
        where_sql = " AND ".join(where_clauses)
        
        sql = f"""
            SELECT 
                COALESCE(m.region, '') as 大区,
                COALESCE(m.zone, '') as 战区,
                m.dealer_id as 店编号,
                COALESCE(m.dealer_short_name, '') as 店简称,
                m.lead_id as 门店线索id,
                COALESCE(m.channel_1, '') as 一级渠道,
                COALESCE(m.channel_2, '') as 二级渠道,
                COALESCE(m.channel_3, '') as 三级渠道,
                COALESCE(m.channel_4, '') as 四级渠道,
                m.visit_time as 客户进店时间,
                COALESCE(m.follower_id, '') as 跟进人,
                COALESCE(m.follower_name, '') as 顾问姓名,
                COALESCE(m.follower_position, '') as 顾问岗位,
                COALESCE(m.phone, '') as 手机号,
                m.followup_created_time as 创建时间
            FROM mart_customer_visit m
            WHERE {where_sql}
            ORDER BY m.visit_time DESC
        """
        
        result_obj = conn.execute(sql, params)
        columns = [desc[0] for desc in result_obj.description]
        results = result_obj.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "客流明细"
        
        ws.append(columns)
        
        for row in results:
            row_data = []
            for val in row:
                if isinstance(val, datetime):
                    row_data.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    row_data.append(val if val is not None else '')
            ws.append(row_data)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"客流明细_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/tables', methods=['GET'])
def get_query_tables():
    """获取所有可查询的表列表"""
    try:
        category = request.args.get('category')
        tables = metadata_registry.get_all_tables(category=category)
        return jsonify({
            'success': True,
            'data': tables
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/table/<table_name>/schema', methods=['GET'])
def get_table_schema(table_name):
    """获取指定表的字段元数据"""
    try:
        schema = metadata_registry.get_table_schema(table_name)
        if not schema:
            return jsonify({
                'success': False,
                'message': f'Table {table_name} not found or not queryable'
            }), 404

        return jsonify({
            'success': True,
            'data': schema
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/detail', methods=['POST'])
def query_detail():
    """明细查询（精确查询单条/多条记录）"""
    try:
        data = request.get_json(silent=True) or {}

        table_name = data.get('table')
        if not table_name:
            return jsonify({'success': False, 'message': 'table is required'}), 400

        if not metadata_registry.validate_table(table_name):
            return jsonify({'success': False, 'message': f'Invalid table: {table_name}'}), 400

        table_schema = metadata_registry.get_table_schema(table_name)
        columns_meta = {col['name']: col for col in table_schema['columns']}

        columns = data.get('columns', ['*'])
        if columns != ['*']:
            columns = metadata_registry.validate_columns(table_name, columns)
            if not columns:
                columns = list(columns_meta.keys())

        filters = data.get('filters', [])
        for f in filters:
            if 'field' not in f or 'operator' not in f:
                return jsonify({'success': False, 'message': 'Each filter must have field and operator'}), 400

        order_by = data.get('order_by', [])
        page = int(data.get('page', 1))
        page_size = int(data.get('page_size', DEFAULT_PAGE_SIZE))

        data_sql, count_sql, params = build_detail_query(
            table_name=table_name,
            columns=columns,
            filters=filters,
            order_by=order_by,
            page=page,
            page_size=page_size,
            columns_metadata=columns_meta
        )

        conn = duck_db.get_connection()

        count_result = conn.execute(count_sql, params).fetchone()
        total = count_result[0] if count_result else 0

        if page_size > MAX_QUERY_ROWS:
            return jsonify({'success': False, 'message': 'page_size too large, max is ' + str(MAX_QUERY_ROWS)}), 400

        result_obj = conn.execute(data_sql, params)
        column_names = [d[0] for d in result_obj.description]
        results = result_obj.fetchall()

        result_list = []
        for row in results:
            row_dict = dict(zip(column_names, row))
            for col_name, col_info in columns_meta.items():
                if col_info.get('sensitive') and col_name in row_dict and row_dict[col_name]:
                    row_dict[col_name] = mask_sensitive(row_dict[col_name], col_info['type'])
            result_list.append(row_dict)

        return jsonify({
            'success': True,
            'data': {
                'list': result_list,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total,
                    'total_pages': (total + page_size - 1) // page_size if page_size > 0 else 1
                }
            }
        })
    except QueryBuilderError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/aggregate', methods=['POST'])
def query_aggregate():
    """聚合查询（多维度统计）"""
    try:
        data = request.get_json(silent=True) or {}

        table_name = data.get('table')
        if not table_name:
            return jsonify({'success': False, 'message': 'table is required'}), 400

        if not metadata_registry.validate_table(table_name):
            return jsonify({'success': False, 'message': f'Invalid table: {table_name}'}), 400

        table_schema = metadata_registry.get_table_schema(table_name)
        columns_meta = {col['name']: col for col in table_schema['columns']}

        group_by = data.get('group_by', [])
        if group_by:
            group_by = metadata_registry.validate_columns(table_name, group_by)

        aggregations = data.get('aggregations', [])
        if not aggregations and not group_by:
            aggregations = [{'field': '*', 'func': 'COUNT', 'alias': 'count'}]

        for agg in aggregations:
            if 'field' not in agg or 'func' not in agg:
                return jsonify({'success': False, 'message': 'Each aggregation must have field and func'}), 400

        filters = data.get('filters', [])
        for f in filters:
            if 'field' not in f or 'operator' not in f:
                return jsonify({'success': False, 'message': 'Each filter must have field and operator'}), 400

        order_by = data.get('order_by', [])
        page = int(data.get('page', 1))
        page_size = int(data.get('page_size', DEFAULT_PAGE_SIZE))

        data_sql, count_sql, params = build_aggregate_query(
            table_name=table_name,
            group_by=group_by,
            aggregations=aggregations,
            filters=filters,
            order_by=order_by,
            page=page,
            page_size=page_size,
            columns_metadata=columns_meta
        )

        conn = duck_db.get_connection()

        count_result = conn.execute(count_sql, params).fetchone()
        total = count_result[0] if count_result else 0

        result_obj = conn.execute(data_sql, params)
        column_names = [d[0] for d in result_obj.description]
        results = result_obj.fetchall()

        result_list = []
        for row in results:
            row_dict = dict(zip(column_names, row))
            result_list.append(row_dict)

        return jsonify({
            'success': True,
            'data': {
                'list': result_list,
                'pagination': {
                    'page': page,
                    'page_size': page_size,
                    'total': total,
                    'total_pages': (total + page_size - 1) // page_size if page_size > 0 else 1
                },
                'meta': {
                    'group_by': group_by,
                    'aggregations': aggregations
                }
            }
        })
    except QueryBuilderError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/export', methods=['POST'])
def export_query_result():
    """导出查询结果为 Excel"""
    try:
        import openpyxl
        from io import BytesIO

        data = request.get_json(silent=True) or {}
        query_type = data.get('query_type', 'detail')

        table_name = data.get('table')
        if not table_name:
            return jsonify({'success': False, 'message': 'table is required'}), 400

        if not metadata_registry.validate_table(table_name):
            return jsonify({'success': False, 'message': f'Invalid table: {table_name}'}), 400

        table_schema = metadata_registry.get_table_schema(table_name)
        columns_meta = {col['name']: col for col in table_schema['columns']}

        filters = data.get('filters', [])
        order_by = data.get('order_by', [])

        if query_type == 'detail':
            columns = data.get('columns', ['*'])
            if columns != ['*']:
                columns = metadata_registry.validate_columns(table_name, columns)
                if not columns:
                    columns = list(columns_meta.keys())

            data_sql, _, params = build_detail_query(
                table_name=table_name,
                columns=columns,
                filters=filters,
                order_by=order_by,
                page=1,
                page_size=MAX_QUERY_ROWS,
                columns_metadata=columns_meta
            )
        else:
            group_by = data.get('group_by', [])
            if group_by:
                group_by = metadata_registry.validate_columns(table_name, group_by)

            aggregations = data.get('aggregations', [])
            if not aggregations and not group_by:
                aggregations = [{'field': '*', 'func': 'COUNT', 'alias': 'count'}]

            data_sql, _, params = build_aggregate_query(
                table_name=table_name,
                group_by=group_by,
                aggregations=aggregations,
                filters=filters,
                order_by=order_by,
                page=1,
                page_size=MAX_QUERY_ROWS,
                columns_metadata=columns_meta
            )

        conn = duck_db.get_connection()
        result_obj = conn.execute(data_sql, params)
        column_names = [d[0] for d in result_obj.description]
        results = result_obj.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_schema['display_name']

        header_row = [columns_meta.get(col, {}).get('display_name', col) for col in column_names]
        ws.append(header_row)

        for row in results:
            row_data = []
            for i, val in enumerate(row):
                col_name = column_names[i]
                col_info = columns_meta.get(col_name, {})
                if col_info.get('sensitive') and val:
                    val = mask_sensitive(val, col_info.get('type', ''))
                if isinstance(val, datetime):
                    row_data.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    row_data.append(val if val is not None else '')
            ws.append(row_data)

        for col_idx, col_name in enumerate(column_names, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{table_schema['display_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except QueryBuilderError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/filterable/<table_name>', methods=['GET'])
def get_filterable_columns(table_name):
    """获取可过滤的字段列表"""
    try:
        if not metadata_registry.validate_table(table_name):
            return jsonify({'success': False, 'message': f'Invalid table: {table_name}'}), 404

        columns = metadata_registry.get_filterable_columns(table_name)
        return jsonify({
            'success': True,
            'data': columns
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/groupable/<table_name>', methods=['GET'])
def get_groupable_columns(table_name):
    """获取可分组的字段列表"""
    try:
        if not metadata_registry.validate_table(table_name):
            return jsonify({'success': False, 'message': f'Invalid table: {table_name}'}), 404

        columns = metadata_registry.get_groupable_columns(table_name)
        return jsonify({
            'success': True,
            'data': columns
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/query/aggregatable/<table_name>', methods=['GET'])
def get_aggregatable_columns(table_name):
    """获取可聚合的字段列表"""
    try:
        if not metadata_registry.validate_table(table_name):
            return jsonify({'success': False, 'message': f'Invalid table: {table_name}'}), 404

        columns = metadata_registry.get_aggregatable_columns(table_name)
        return jsonify({
            'success': True,
            'data': columns
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


def mask_sensitive(value, field_type):
    """脱敏处理"""
    if not value:
        return value
    value_str = str(value)
    if len(value_str) <= 4:
        return '***'
    if 'PHONE' in field_type.upper() or 'VARCHAR' in field_type.upper():
        if len(value_str) == 11:
            return value_str[:3] + '****' + value_str[7:]
        return value_str[:2] + '***' + value_str[-2:]
    return '***'


def _rows_to_dicts(cursor):
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _current_year_month():
    return datetime.now().strftime('%Y-%m')


def _parse_float(value, default=0):
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _current_user_has_all_data_scope():
    user = getattr(g, "current_user", None) or {}
    roles = user.get("roles") or []
    if not roles:
        return False
    return any((role.get("data_scope_type") or "all") == "all" for role in roles)


def _current_user_forced_lead_owner():
    user = getattr(g, "current_user", None) or {}
    if not user or _current_user_has_all_data_scope():
        return ""
    # 当前账号体系尚未建立账号到门店范围的完整映射。对于非全量数据范围，
    # 漏斗驾驶舱先强制限定到 display_name 对应的线索运营区域负责人。
    return (user.get("display_name") or "").strip()


def _parse_terminal_target_excel(file_storage, year_month, operator_name=''):
    wb = load_workbook(file_storage, read_only=False, data_only=True)
    ws = wb.active

    def cell_value(row, col):
        return ws.cell(row=row, column=col).value

    model_groups = []
    current_model = None
    for col in range(8, ws.max_column + 1):
        level2 = cell_value(2, col)
        level3 = cell_value(3, col)
        if level2:
            current_model = str(level2).strip()
        if not current_model or not level3:
            continue
        label = str(level3).strip()
        if current_model == '合计' and label == '终端目标':
            model_groups.append({'model': '__TOTAL__', 'col': col, 'label': label, 'priority': 1})
        elif label == '终端小计':
            model_groups.append({'model': current_model, 'col': col, 'label': label, 'priority': 1})
        elif label == '终端':
            model_groups.append({'model': current_model, 'col': col, 'label': label, 'priority': 2})

    selected_cols = {}
    for item in model_groups:
        model = item['model']
        if model not in selected_cols or item['priority'] < selected_cols[model]['priority']:
            selected_cols[model] = item

    dealers = {}
    errors = []
    for row in range(4, ws.max_row + 1):
        dealer_id = cell_value(row, 5)
        dealer_name = cell_value(row, 7)
        if not dealer_id:
            continue
        dealer_id = str(dealer_id).strip()
        dealer_name = str(dealer_name or '').strip()
        if not dealer_id:
            continue
        total_target = _parse_float(cell_value(row, selected_cols.get('__TOTAL__', {}).get('col')) if '__TOTAL__' in selected_cols else 0)
        model_targets = {}
        for model, item in selected_cols.items():
            if model == '__TOTAL__':
                continue
            value = _parse_float(cell_value(row, item['col']))
            if value:
                model_targets[model] = value
        if dealer_id in dealers:
            errors.append({'row': row, 'dealer_id': dealer_id, 'message': '重复店编号，后出现记录覆盖前值'})
        dealers[dealer_id] = {
            'dealer_id': dealer_id,
            'dealer_name': dealer_name,
            'dealer_total_sales_target': total_target,
            'model_targets': model_targets,
        }

    return dealers, errors


def _import_funnel_sales_targets(file_storage, year_month, operator_name=''):
    duck_db.ensure_funnel_schema()
    dealers, errors = _parse_terminal_target_excel(file_storage, year_month, operator_name)
    conn = duck_db.get_connection()
    current_dealers = {
        row[0]: row[1]
        for row in conn.execute("SELECT dealer_id, dealer_name FROM mart_dealers").fetchall()
    }
    now = datetime.now()
    source_file = file_storage.filename or ''
    rows = []
    skipped = []
    for dealer_id, item in dealers.items():
        if dealer_id not in current_dealers:
            skipped.append({'dealer_id': dealer_id, 'dealer_name': item['dealer_name']})
            continue
        dealer_name = current_dealers.get(dealer_id) or item['dealer_name']
        for model_name, target in item['model_targets'].items():
            rows.append((
                year_month,
                dealer_id,
                dealer_name,
                model_name,
                target,
                item['dealer_total_sales_target'],
                source_file,
                operator_name,
                now,
                now,
            ))

    conn.execute("DELETE FROM funnel_sales_targets WHERE year_month = ?", [year_month])
    if rows:
        conn.executemany("""
            INSERT INTO funnel_sales_targets VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)

    log_id = int(datetime.now().timestamp() * 1000)
    summary = {
        'skipped_dealers': skipped[:100],
        'errors': errors[:100],
    }
    conn.execute("""
        INSERT INTO funnel_import_logs VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        log_id,
        year_month,
        source_file,
        len(dealers),
        len(dealers) - len(skipped),
        len(skipped),
        len(rows),
        len(errors),
        json.dumps(summary, ensure_ascii=False),
        operator_name,
        now,
    ])
    conn.commit()
    duck_db.compute_funnel_targets(year_month)
    return {
        'import_id': log_id,
        'year_month': year_month,
        'file_dealer_count': len(dealers),
        'matched_dealer_count': len(dealers) - len(skipped),
        'skipped_dealer_count': len(skipped),
        'imported_target_count': len(rows),
        'sales_target_sum': sum(float(row[4] or 0) for row in rows),
        'dealer_total_sales_target_sum': sum(float(item['dealer_total_sales_target'] or 0) for item in dealers.values() if item['dealer_id'] in current_dealers),
        'latest_updated_at': now.isoformat(),
        'error_count': len(errors),
        'skipped_dealers': skipped[:20],
        'errors': errors[:20],
    }


def _funnel_filters(include_channels=False, include_model=True):
    filters = []
    params = []
    year_month = request.args.get('year_month') or _current_year_month()
    filters.append("year_month = ?")
    params.append(year_month)
    forced_owner = _current_user_forced_lead_owner()
    dealer_search = request.args.get('dealer_search', '').strip()
    if dealer_search:
        filters.append("(dealer_id ILIKE ? OR dealer_name ILIKE ?)")
        like_value = f"%{dealer_search}%"
        params.extend([like_value, like_value])
    for arg, column in [
        ('region', 'region'),
        ('zone', 'zone'),
        ('dealer_id', 'dealer_id'),
        ('lead_ops_support', 'lead_ops_support'),
    ]:
        value = request.args.get(arg, '')
        if value:
            filters.append(f"{column} = ?")
            params.append(value)
    lead_ops_owner = forced_owner or request.args.get('lead_ops_owner', '')
    if lead_ops_owner:
        filters.append("lead_ops_owner = ?")
        params.append(lead_ops_owner)
    if include_model:
        value = request.args.get('model_name', '')
        if value:
            filters.append("model_name = ?")
            params.append(value)
    if include_channels:
        for arg, column in [
            ('channel_2', 'channel_2'),
            ('channel_3', 'channel_3'),
        ]:
            value = request.args.get(arg, '')
            if value:
                filters.append(f"{column} = ?")
                params.append(value)
    return year_month, " AND ".join(filters), params


def _progress_status(progress_gap_rate, config_error=False):
    if config_error:
        return '配置异常'
    gap = float(progress_gap_rate or 0)
    if gap >= 0.05:
        return '领先'
    if gap >= -0.05:
        return '正常'
    if gap >= -0.15:
        return '轻度落后'
    return '严重落后'


def _enrich_funnel_dealer_rows(rows, year_month):
    progress = duck_db._progress_ratios(year_month)
    data_progress_ratio = float(progress.get('data_progress_ratio') or 0)
    valid_rates = []
    visit_rates = []
    for row in rows:
        online = float(row.get('online_lead_count') or 0)
        valid = float(row.get('valid_lead_count') or 0)
        visit = float(row.get('visit_count') or 0)
        if online > 0:
            row['lead_valid_rate'] = valid * 100.0 / online
            row['lead_visit_rate'] = visit * 100.0 / online
            valid_rates.append(row['lead_valid_rate'])
            visit_rates.append(row['lead_visit_rate'])
        else:
            row['lead_valid_rate'] = 0
            row['lead_visit_rate'] = 0
    avg_valid_rate = sum(valid_rates) / len(valid_rates) if valid_rates else 0
    avg_visit_rate = sum(visit_rates) / len(visit_rates) if visit_rates else 0
    for row in rows:
        dealer_visit_target = float(row.get('dealer_visit_target') or 0)
        dealer_visit_target_to_date = float(row.get('dealer_visit_target_to_date') or 0)
        derived_visit_target_to_date = float(row.get('derived_visit_target_to_date') or 0)
        visit = float(row.get('visit_count') or 0)
        sales_target = float(row.get('sales_target') or 0)
        missing_rate_count = int(row.pop('missing_rate_count', 0) or 0)
        progress_gap_rate = (visit / dealer_visit_target - data_progress_ratio) if dealer_visit_target > 0 else 0
        derived_achievement_rate = visit * 100.0 / derived_visit_target_to_date if derived_visit_target_to_date > 0 else 0
        config_error = dealer_visit_target <= 0 or (sales_target > 0 and missing_rate_count > 0)
        status = _progress_status(progress_gap_rate, config_error)
        tags = []
        if config_error:
            tags.append('配置异常')
        if float(row.get('online_lead_count') or 0) <= 0:
            tags.append('线索不足')
        if avg_valid_rate > 0 and row.get('lead_valid_rate', 0) < avg_valid_rate * 0.8:
            tags.append('有效率低')
        if avg_visit_rate > 0 and row.get('lead_visit_rate', 0) < avg_visit_rate * 0.8:
            tags.append('到店转化低')
        if dealer_visit_target_to_date > 0 and derived_visit_target_to_date > dealer_visit_target_to_date * 1.2:
            tags.append('倒推压力高')
        if status in ('轻度落后', '严重落后') and not tags:
            tags.append('进度落后')
        if not tags:
            tags.append('正常')
        row['data_progress_ratio'] = data_progress_ratio
        row['progress_gap_rate'] = progress_gap_rate * 100.0
        row['progress_status'] = status
        row['status_label'] = status
        row['diagnosis_tags'] = tags
        row['primary_diagnosis'] = tags[0]
        row['derived_achievement_rate'] = derived_achievement_rate
    return rows


def _get_funnel_org_dealer_rows(year_month, where_sql, params, limit=None):
    conn = duck_db.get_connection()
    rows = _rows_to_dicts(conn.execute(f"""
        SELECT
            ft.year_month,
            ft.dealer_id,
            ft.dealer_name,
            ft.region,
            ft.zone,
            COALESCE(MAX(md.province), '') AS province,
            ft.lead_ops_owner,
            ft.lead_ops_support,
            MAX(ft.national_visit_target) AS national_visit_target,
            MAX(ft.dealer_online_lead_share) AS dealer_online_lead_share,
            MAX(ft.dealer_visit_target) AS dealer_visit_target,
            MAX(ft.elapsed_day_ratio) AS elapsed_day_ratio,
            MAX(ft.dealer_visit_target_to_date) AS dealer_visit_target_to_date,
            MAX(ft.dealer_visit_gap) AS dealer_visit_gap,
            MAX(ft.dealer_visit_achievement_rate) AS dealer_visit_achievement_rate,
            SUM(ft.online_lead_count) AS online_lead_count,
            SUM(ft.valid_lead_count) AS valid_lead_count,
            SUM(ft.visit_count) AS visit_count,
            SUM(ft.sales_count) AS sales_count,
            SUM(ft.sales_target) AS sales_target,
            SUM(ft.derived_visit_target) AS derived_visit_target,
            SUM(ft.derived_visit_target_to_date) AS derived_visit_target_to_date,
            SUM(ft.derived_visit_gap) AS derived_visit_gap,
            MAX(ft.projected_month_end_visit) AS projected_month_end_visit,
            SUM(CASE WHEN ft.conversion_rate_source = '未配置' AND ft.sales_target > 0 THEN 1 ELSE 0 END) AS missing_rate_count
        FROM (
            SELECT * FROM funnel_metric_targets
            WHERE {where_sql}
        ) ft
        LEFT JOIN mart_dealers md ON ft.dealer_id = md.dealer_id
        GROUP BY ft.year_month, ft.dealer_id, ft.dealer_name, ft.region, ft.zone, ft.lead_ops_owner, ft.lead_ops_support
        ORDER BY dealer_visit_gap ASC
    """, params))
    rows = _enrich_funnel_dealer_rows(rows, year_month)
    status_filter = request.args.get('progress_status', '').strip()
    diagnosis_filter = request.args.get('diagnosis_tag', '').strip()
    if status_filter:
        rows = [row for row in rows if row.get('progress_status') == status_filter]
    if diagnosis_filter:
        rows = [row for row in rows if diagnosis_filter in (row.get('diagnosis_tags') or [])]
    rows.sort(key=lambda row: float(row.get('dealer_visit_gap') or 0))
    if limit:
        rows = rows[:limit]
    return rows


def _normalize_province_name(value):
    province = (value or '').strip()
    for suffix in ('特别行政区', '壮族自治区', '回族自治区', '维吾尔自治区', '自治区', '省', '市'):
        if province.endswith(suffix):
            return province[:-len(suffix)]
    return province


def _raw_store_province_map():
    if not RAW_DB_PATH.exists():
        return {}
    try:
        conn = sqlite3.connect(str(RAW_DB_PATH))
        conn.row_factory = sqlite3.Row
        rows = conn.execute("""
            SELECT 店编号 AS dealer_id, 省份 AS province
            FROM 门店表
            WHERE 店编号 IS NOT NULL
              AND TRIM(店编号) != ''
              AND 省份 IS NOT NULL
              AND TRIM(省份) != ''
        """).fetchall()
        conn.close()
        return {str(row['dealer_id']).strip(): _normalize_province_name(row['province']) for row in rows}
    except Exception as exc:
        print(f"Failed to load raw store province map: {exc}")
        return {}


def _get_funnel_summary(year_month):
    duck_db.ensure_funnel_schema()
    conn = duck_db.get_connection()
    progress = duck_db._progress_ratios(year_month)
    latest_lead_date = progress.get('latest_lead_date')
    if isinstance(latest_lead_date, (datetime, date)):
        latest_lead_date = latest_lead_date.strftime('%Y-%m-%d')
    year, month = [int(part) for part in year_month.split("-")]
    month_start = f"{year}-{month:02d}-01"
    today = datetime.now().date()
    if today.year == year and today.month == month:
        month_end = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        month_end = date(year, month, calendar.monthrange(year, month)[1]).strftime("%Y-%m-%d")
    visit_count_row = conn.execute("""
        WITH dealer_visits AS (
            SELECT
                f.dealer_id,
                COUNT(DISTINCT m.lead_id || '_' || CAST(m.visit_time AS DATE)) AS visit_count
            FROM (
                SELECT DISTINCT dealer_id, CAST(visit_time AS DATE) AS visit_date
                FROM mart_customer_visit
                WHERE CAST(visit_time AS DATE) >= ?
                  AND CAST(visit_time AS DATE) <= ?
                  AND channel_1 = '线上'
            ) f
            JOIN mart_customer_visit m ON f.dealer_id = m.dealer_id
                AND CAST(m.visit_time AS DATE) = f.visit_date
                AND m.channel_1 = '线上'
            GROUP BY f.dealer_id
        )
        SELECT COALESCE(SUM(visit_count), 0)
        FROM dealer_visits
    """, [month_start, month_end]).fetchone()
    actual_visit_count = visit_count_row[0] if visit_count_row and visit_count_row[0] is not None else 0
    row = conn.execute("""
        WITH dealer_level AS (
            SELECT
                year_month,
                dealer_id,
                region,
                SUM(visit_count) AS visit_count,
                MAX(national_visit_target) AS national_visit_target,
                MAX(elapsed_day_ratio) AS elapsed_day_ratio,
                MAX(dealer_visit_target_to_date) AS dealer_visit_target_to_date,
                MAX(dealer_visit_gap) AS dealer_visit_gap,
                MAX(dealer_visit_achievement_rate) AS dealer_visit_achievement_rate,
                MAX(projected_month_end_visit) AS projected_month_end_visit
            FROM funnel_metric_targets
            WHERE year_month = ?
            GROUP BY year_month, dealer_id, region
        ),
        model_level AS (
            SELECT
                SUM(visit_count) AS model_visit_count,
                SUM(derived_visit_target_to_date) AS derived_visit_target_to_date
            FROM funnel_metric_targets
            WHERE year_month = ?
        )
        SELECT
            COALESCE(SUM(d.visit_count), 0) AS visit_count,
            COALESCE(MAX(d.national_visit_target), 0) AS national_visit_target,
            COALESCE(MAX(d.elapsed_day_ratio), 0) AS elapsed_day_ratio,
            COALESCE(SUM(d.dealer_visit_target_to_date), 0) AS visit_target_to_date,
            CASE WHEN COALESCE(MAX(d.national_visit_target), 0) > 0 THEN COALESCE(SUM(d.dealer_visit_gap), 0) ELSE 0 END AS visit_gap,
            COALESCE(AVG(d.dealer_visit_achievement_rate), 0) AS avg_visit_achievement_rate,
            COALESCE(MAX(m.derived_visit_target_to_date), 0) AS derived_visit_target_to_date,
            CASE WHEN COALESCE(MAX(m.derived_visit_target_to_date), 0) > 0
                THEN COALESCE(MAX(m.model_visit_count), 0) * 100.0 / MAX(m.derived_visit_target_to_date) ELSE 0 END AS derived_achievement_rate,
            COALESCE(SUM(d.projected_month_end_visit), 0) AS projected_month_end_visit,
            COUNT(DISTINCT CASE WHEN d.dealer_visit_achievement_rate < 100 AND d.national_visit_target > 0 THEN d.region END) AS lagging_region_count
        FROM dealer_level d
        CROSS JOIN model_level m
    """, [year_month, year_month]).fetchone()
    top_regions = _rows_to_dicts(conn.execute("""
        SELECT region, SUM(dealer_visit_gap) AS visit_gap
        FROM (
            SELECT DISTINCT year_month, dealer_id, region, dealer_visit_gap
            FROM funnel_metric_targets WHERE year_month = ?
        )
        GROUP BY region
        ORDER BY visit_gap ASC
        LIMIT 3
    """, [year_month]))
    visit_target = conn.execute("""
        SELECT national_visit_target, updated_at FROM funnel_national_visit_targets WHERE year_month = ?
        """, [year_month]).fetchone()
    default_rate = conn.execute("""
        SELECT conversion_rate, updated_at FROM funnel_conversion_rates
        WHERE year_month = ? AND scope_type = 'national'
        ORDER BY updated_at DESC LIMIT 1
    """, [year_month]).fetchone()
    return {
        'year_month': year_month,
        'visit_count': actual_visit_count,
        'national_visit_target': row[1] or 0,
        'elapsed_day_ratio': row[2] or 0,
        'data_progress_ratio': progress.get('data_progress_ratio') or 0,
        'latest_lead_date': latest_lead_date,
        'visit_target_to_date': row[3] or 0,
        'visit_gap': row[4] or 0,
        'visit_achievement_rate': (actual_visit_count * 100.0 / row[3]) if row and row[3] else 0,
        'derived_visit_target_to_date': row[6] or 0,
        'derived_achievement_rate': row[7] or 0,
        'projected_month_end_visit': row[8] or 0,
        'lagging_region_count': row[9] or 0,
        'top_lagging_regions': top_regions,
        'visit_target_updated_at': visit_target[1] if visit_target else None,
        'default_conversion_rate': default_rate[0] if default_rate else None,
        'default_conversion_rate_updated_at': default_rate[1] if default_rate else None,
    }


def _status_to_map_status(progress_status):
    if progress_status in ('领先', '正常'):
        return 'normal'
    if progress_status == '轻度落后':
        return 'watch'
    return 'lag'


def _visit_trend_monthly_payload(year_month):
    duck_db.ensure_funnel_schema()
    conn = duck_db.get_connection()
    progress = duck_db._progress_ratios(year_month)
    latest_lead_date = progress.get('latest_lead_date')
    if isinstance(latest_lead_date, (datetime, date)):
        latest_lead_date_str = latest_lead_date.strftime('%Y-%m-%d')
    else:
        latest_lead_date_str = latest_lead_date

    year, month = [int(part) for part in year_month.split("-")]
    month_days = calendar.monthrange(year, month)[1]
    target_row = conn.execute("""
        SELECT COALESCE(MAX(national_visit_target), 0)
        FROM funnel_national_visit_targets
        WHERE year_month = ? AND is_active
    """, [year_month]).fetchone()
    monthly_target = float(target_row[0] or 0) if target_row else 0
    target_daily = monthly_target / month_days if monthly_target else 0

    current_rows = _rows_to_dicts(conn.execute("""
        WITH days AS (
            SELECT DISTINCT CAST(visit_time AS DATE) AS report_date
            FROM mart_customer_visit
            WHERE strftime(CAST(visit_time AS DATE), '%Y-%m') = ?
              AND channel_1 = '线上'
        ),
        visits AS (
            SELECT
                f.report_date,
                f.dealer_id,
                COUNT(DISTINCT m.lead_id || '_' || CAST(m.visit_time AS DATE)) AS visits
            FROM (
                SELECT DISTINCT dealer_id, CAST(visit_time AS DATE) AS report_date
                FROM mart_customer_visit
                WHERE strftime(CAST(visit_time AS DATE), '%Y-%m') = ?
                  AND channel_1 = '线上'
            ) f
            JOIN mart_customer_visit m ON f.dealer_id = m.dealer_id
                AND CAST(m.visit_time AS DATE) = f.report_date
                AND m.channel_1 = '线上'
            GROUP BY f.report_date, f.dealer_id
        ),
        dealer_leads AS (
            SELECT assign_date AS report_date, COUNT(*) AS dealer_leads
            FROM mart_leads
            WHERE strftime(assign_date, '%Y-%m') = ?
              AND channel_1 = '线上'
              AND dealer_id IN (SELECT dealer_id FROM mart_dealers)
            GROUP BY assign_date
        )
        SELECT
            d.report_date,
            COALESCE(SUM(v.visits), 0) AS visits,
            COALESCE(MAX(dl.dealer_leads), 0) AS dealer_leads
        FROM days d
        LEFT JOIN visits v ON d.report_date = v.report_date
        LEFT JOIN dealer_leads dl ON d.report_date = dl.report_date
        GROUP BY d.report_date
        ORDER BY d.report_date
    """, [year_month, year_month, year_month]))

    previous_month_date = date(year, month, 1) - timedelta(days=1)
    previous_year_month = previous_month_date.strftime('%Y-%m')
    previous_rows = _rows_to_dicts(conn.execute("""
        WITH visits AS (
            SELECT
                f.report_date,
                f.dealer_id,
                COUNT(DISTINCT m.lead_id || '_' || CAST(m.visit_time AS DATE)) AS visits
            FROM (
                SELECT DISTINCT dealer_id, CAST(visit_time AS DATE) AS report_date
                FROM mart_customer_visit
                WHERE strftime(CAST(visit_time AS DATE), '%Y-%m') = ?
                  AND channel_1 = '线上'
            ) f
            JOIN mart_customer_visit m ON f.dealer_id = m.dealer_id
                AND CAST(m.visit_time AS DATE) = f.report_date
                AND m.channel_1 = '线上'
            GROUP BY f.report_date, f.dealer_id
        )
        SELECT EXTRACT(day FROM report_date)::INTEGER AS day, COALESCE(SUM(visits), 0) AS visits
        FROM visits
        GROUP BY EXTRACT(day FROM report_date)::INTEGER
    """, [previous_year_month]))
    previous_by_day = {int(row['day']): float(row['visits'] or 0) for row in previous_rows}

    cumulative = 0
    trend = []
    for row in current_rows:
        report_date = row['report_date']
        if isinstance(report_date, datetime):
            report_date = report_date.date()
        if isinstance(report_date, str):
            report_date_obj = datetime.strptime(report_date[:10], '%Y-%m-%d').date()
        else:
            report_date_obj = report_date
        day = int(report_date_obj.day)
        visits = float(row.get('visits') or 0)
        dealer_leads = float(row.get('dealer_leads') or 0)
        last_month = previous_by_day.get(day, 0)
        cumulative += visits
        trend.append({
            'label': f'{month}/{day}',
            'date': report_date_obj.strftime('%Y-%m-%d'),
            'day': day,
            'visits': visits,
            'target': target_daily,
            'lastMonth': last_month,
            'cumulative': cumulative,
            'progress': (cumulative * 100.0 / monthly_target) if monthly_target else 0,
            'mom': ((visits - last_month) * 100.0 / last_month) if last_month else 0,
            'dealerLeads': dealer_leads,
            'visitRate': (visits * 100.0 / dealer_leads) if dealer_leads else 0,
        })

    for index, row in enumerate(trend):
        row['ma7'] = _moving_average(trend, 'visits', 7, index)
        row['ma15'] = _moving_average(trend, 'visits', 15, index)

    summary = _get_funnel_summary(year_month)
    total_dealer_leads = sum(float(row.get('dealerLeads') or 0) for row in trend)
    summary.update({
        'dealer_lead_count': total_dealer_leads,
        'visit_rate': (summary.get('visit_count', 0) * 100.0 / total_dealer_leads) if total_dealer_leads else 0,
    })

    return {
        'year_month': year_month,
        'latest_data_date': latest_lead_date_str,
        'monthly_target': monthly_target,
        'target_daily': target_daily,
        'trend': trend,
        'summary': summary,
    }


def _moving_average(rows, key, window_size, index):
    start = max(0, index - window_size + 1)
    values = [float(row.get(key) or 0) for row in rows[start:index + 1]]
    return sum(values) / len(values) if values else 0


def _visit_trend_province_payload(year_month):
    duck_db.ensure_funnel_schema()
    year_month, where_sql, params = _funnel_filters(include_model=False)
    rows = _get_funnel_org_dealer_rows(year_month, where_sql, params)
    province_map = _raw_store_province_map()
    groups = {}
    for row in rows:
        province = _normalize_province_name(row.get('province') or province_map.get(str(row.get('dealer_id') or '').strip()) or '') or '未归属'
        region = (row.get('region') or '').strip() or '未归属'
        item = groups.setdefault(province, {
            'name': province,
            'region': region,
            'regions': set(),
            'dealer_count': 0,
            'visit_count': 0,
            'visit_target_to_date': 0,
            'visit_gap': 0,
            'normal_count': 0,
            'watch_count': 0,
            'lag_count': 0,
        })
        item['regions'].add(region)
        if item['region'] == '未归属' and region != '未归属':
            item['region'] = region
        item['dealer_count'] += 1
        item['visit_count'] += float(row.get('visit_count') or 0)
        item['visit_target_to_date'] += float(row.get('dealer_visit_target_to_date') or 0)
        item['visit_gap'] += float(row.get('dealer_visit_gap') or 0)
        status = _status_to_map_status(row.get('progress_status'))
        if status == 'normal':
            item['normal_count'] += 1
        elif status == 'watch':
            item['watch_count'] += 1
        else:
            item['lag_count'] += 1

    provinces = []
    for item in groups.values():
        if item['lag_count'] > 0:
            status = 'lag'
        elif item['watch_count'] > 0:
            status = 'watch'
        else:
            status = 'normal'
        item['status'] = status
        item['regions'] = sorted(item['regions'])
        item['region'] = item['regions'][0] if item['regions'] else item['region']
        item['visit_achievement_rate'] = item['visit_count'] * 100.0 / item['visit_target_to_date'] if item['visit_target_to_date'] else 0
        item['desc'] = _province_status_desc(item)
        provinces.append(item)

    provinces.sort(key=lambda item: (item['status'] != 'lag', item['status'] != 'watch', item['visit_gap']))
    return {
        'year_month': year_month,
        'provinces': provinces,
    }


def _province_status_desc(item):
    if item['status'] == 'lag':
        return '进度落后，建议进入漏斗页拆解线索供给、到店率和门店责任。'
    if item['status'] == 'watch':
        return '省内门店状态分化，建议关注近 7 日节奏和落后门店。'
    return '整体节奏正常，可继续日常巡检。'


@app.route('/api/visit-trend/overview', methods=['GET'])
@require_permission('home.view')
def visit_trend_overview():
    try:
        year_month = request.args.get('year_month') or _current_year_month()
        monthly = _visit_trend_monthly_payload(year_month)
        provinces = _visit_trend_province_payload(year_month)
        return jsonify({'success': True, 'data': {**monthly, **provinces}})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/config/visit-targets', methods=['GET', 'POST'])
@require_permission('funnel_target.config.view')
def funnel_visit_targets():
    try:
        duck_db.ensure_funnel_schema()
        conn = duck_db.get_connection()
        if request.method == 'POST':
            if not g.current_user or 'funnel_target.config.manage' not in g.current_user.get('permissions', []):
                return jsonify({'success': False, 'message': '无权限'}), 403
            payload = request.get_json(silent=True) or {}
            year_month = payload.get('year_month') or _current_year_month()
            target = _parse_float(payload.get('national_visit_target'))
            duck_db.set_funnel_visit_target(year_month, target, g.current_user.get('display_name', ''))
            record_audit_log(auth_db, '漏斗目标分析', '配置全国到店目标', 'funnel_visit_target', year_month, after_data=payload)
        rows = _rows_to_dicts(conn.execute("""
            SELECT * FROM funnel_national_visit_targets ORDER BY year_month DESC
        """))
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/config/model-source-values/scan', methods=['POST'])
@require_permission('funnel_target.config.manage')
def funnel_model_source_values_scan():
    try:
        payload = request.get_json(silent=True) or {}
        year_month = payload.get('year_month') or _current_year_month()
        duck_db.scan_funnel_model_source_values(year_month)
        return jsonify({'success': True, 'data': {'year_month': year_month}})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/config/model-source-values', methods=['GET'])
@require_permission('funnel_target.config.view')
def funnel_model_source_values():
    try:
        duck_db.ensure_funnel_schema()
        conn = duck_db.get_connection()
        year_month = request.args.get('year_month') or _current_year_month()
        source_type = request.args.get('source_type') or ''
        mapping_status = request.args.get('mapping_status') or ''
        filters = ["year_month = ?"]
        params = [year_month]
        if source_type:
            filters.append("source_type = ?")
            params.append(source_type)
        if mapping_status:
            filters.append("mapping_status = ?")
            params.append(mapping_status)
        where_sql = " AND ".join(filters)
        rows = _rows_to_dicts(conn.execute(f"""
            SELECT *
            FROM funnel_model_source_values
            WHERE {where_sql}
            ORDER BY
                CASE WHEN mapping_status = '未映射' THEN 0 ELSE 1 END,
                metric_count DESC,
                occurrence_count DESC,
                source_type,
                source_model_value
            LIMIT 1000
        """, params))
        summary_rows = _rows_to_dicts(conn.execute(f"""
            SELECT
                mapping_status,
                source_type,
                COUNT(*) AS source_value_count,
                COALESCE(SUM(metric_count), 0) AS metric_count
            FROM funnel_model_source_values
            WHERE year_month = ?
            GROUP BY mapping_status, source_type
            ORDER BY mapping_status, source_type
        """, [year_month]))
        return jsonify({'success': True, 'data': rows, 'summary': summary_rows})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/config/model-mappings', methods=['GET', 'POST'])
@require_permission('funnel_target.config.view')
def funnel_model_mappings():
    try:
        duck_db.ensure_funnel_schema()
        conn = duck_db.get_connection()
        if request.method == 'POST':
            if not g.current_user or 'funnel_target.config.manage' not in g.current_user.get('permissions', []):
                return jsonify({'success': False, 'message': '无权限'}), 403
            payload = request.get_json(silent=True) or {}
            mappings = payload.get('mappings') or [payload]
            now = datetime.now()
            rows = []
            for item in mappings:
                source_type = (item.get('source_type') or '').strip()
                source_field = (item.get('source_field') or '').strip()
                source_model_value = (item.get('source_model_value') or '').strip()
                standard_model_name = (item.get('standard_model_name') or '').strip()
                if not source_type or not source_field or not source_model_value or not standard_model_name:
                    continue
                rows.append((
                    source_type,
                    source_model_value,
                    standard_model_name,
                    item.get('is_active', True),
                    g.current_user.get('display_name', ''),
                    now,
                    now,
                    source_field,
                    item.get('target_enabled', True),
                ))
            if rows:
                conn.executemany("""
                    INSERT INTO funnel_model_mapping
                    (source_table, source_model_code, standard_model_name, is_active, updated_by, created_at, updated_at, source_field, target_enabled)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT (source_table, source_field, source_model_code) DO UPDATE SET
                        standard_model_name = excluded.standard_model_name,
                        is_active = excluded.is_active,
                        updated_by = excluded.updated_by,
                        updated_at = excluded.updated_at,
                        target_enabled = excluded.target_enabled
                """, rows)
                conn.commit()
                year_month = payload.get('year_month') or _current_year_month()
                duck_db.scan_funnel_model_source_values(year_month)
                record_audit_log(auth_db, '漏斗目标分析', '配置车型映射', 'funnel_model_mapping', year_month, after_data=payload)
        rows = _rows_to_dicts(conn.execute("""
            SELECT *
            FROM funnel_model_mapping
            ORDER BY source_table, source_field, source_model_code
        """))
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/config/conversion-rates', methods=['GET', 'POST'])
@require_permission('funnel_target.config.view')
def funnel_conversion_rates():
    try:
        duck_db.ensure_funnel_schema()
        conn = duck_db.get_connection()
        if request.method == 'POST':
            if not g.current_user or 'funnel_target.config.manage' not in g.current_user.get('permissions', []):
                return jsonify({'success': False, 'message': '无权限'}), 403
            payload = request.get_json(silent=True) or {}
            year_month = payload.get('year_month') or _current_year_month()
            scope_type = payload.get('scope_type') or 'national'
            model_name = payload.get('model_name') or ''
            conversion_rate = _parse_float(payload.get('conversion_rate'))
            duck_db.set_funnel_conversion_rate(year_month, scope_type, model_name, conversion_rate, g.current_user.get('display_name', ''))
            record_audit_log(auth_db, '漏斗目标分析', '配置转化率', 'funnel_conversion_rate', year_month, after_data=payload)
        year_month = request.args.get('year_month') or _current_year_month()
        rows = _rows_to_dicts(conn.execute("""
            SELECT * FROM funnel_conversion_rates WHERE year_month = ? ORDER BY scope_type, model_name
        """, [year_month]))
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/config/sales-targets/import', methods=['POST'])
@require_permission('funnel_target.config.manage')
def funnel_sales_targets_import():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'success': False, 'message': '请上传目标文件'}), 400
        year_month = request.form.get('year_month') or _current_year_month()
        result = _import_funnel_sales_targets(file, year_month, g.current_user.get('display_name', ''))
        record_audit_log(auth_db, '漏斗目标分析', '导入成交目标', 'funnel_sales_target', year_month, after_data=result)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/config/sales-targets', methods=['GET'])
@require_permission('funnel_target.config.view')
def funnel_sales_targets_list():
    try:
        duck_db.ensure_funnel_schema()
        conn = duck_db.get_connection()
        year_month = request.args.get('year_month') or _current_year_month()
        summary = conn.execute("""
            SELECT
                COUNT(*) AS row_count,
                COUNT(DISTINCT dealer_id) AS dealer_count,
                COALESCE(SUM(sales_target), 0) AS sales_target_sum,
                COALESCE(SUM(DISTINCT dealer_total_sales_target), 0) AS dealer_total_sales_target_sum,
                MAX(updated_at) AS latest_updated_at
            FROM funnel_sales_targets
            WHERE year_month = ?
        """, [year_month]).fetchone()
        rows = _rows_to_dicts(conn.execute("""
            SELECT * FROM funnel_sales_targets
            WHERE year_month = ?
            ORDER BY dealer_id, model_name
            LIMIT 1000
        """, [year_month]))
        row_count = summary[0] if summary else 0
        dealer_count = summary[1] if summary else 0
        sales_target_sum = summary[2] if summary else 0
        dealer_total_sales_target_sum = summary[3] if summary else 0
        latest_updated_at = summary[4] if summary else None

        if not rows:
            metric_summary = conn.execute("""
                SELECT
                    COUNT(*) AS row_count,
                    COUNT(DISTINCT dealer_id) AS dealer_count,
                    COALESCE(SUM(sales_target), 0) AS sales_target_sum,
                    COALESCE(SUM(DISTINCT dealer_total_sales_target), 0) AS dealer_total_sales_target_sum,
                    MAX(updated_at) AS latest_updated_at
                FROM funnel_metric_targets
                WHERE year_month = ? AND COALESCE(sales_target, 0) > 0
            """, [year_month]).fetchone()
            metric_rows = _rows_to_dicts(conn.execute("""
                SELECT
                    year_month,
                    dealer_id,
                    dealer_name,
                    model_name,
                    sales_target,
                    dealer_total_sales_target,
                    '明细指标回填' AS source_file,
                    '' AS updated_by,
                    updated_at AS created_at,
                    updated_at
                FROM funnel_metric_targets
                WHERE year_month = ? AND COALESCE(sales_target, 0) > 0
                ORDER BY dealer_id, model_name
                LIMIT 1000
            """, [year_month]))
            if metric_rows:
                rows = metric_rows
                row_count = metric_summary[0] if metric_summary else 0
                dealer_count = metric_summary[1] if metric_summary else 0
                sales_target_sum = metric_summary[2] if metric_summary else 0
                dealer_total_sales_target_sum = metric_summary[3] if metric_summary else 0
                latest_updated_at = metric_summary[4] if metric_summary else None

        return jsonify({
            'success': True,
            'data': rows,
            'summary': {
                'row_count': row_count,
                'dealer_count': dealer_count,
                'sales_target_sum': sales_target_sum,
                'dealer_total_sales_target_sum': dealer_total_sales_target_sum,
                'latest_updated_at': latest_updated_at,
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/recompute', methods=['POST'])
@require_permission('funnel_target.config.manage')
def funnel_recompute():
    try:
        payload = request.get_json(silent=True) or {}
        year_month = payload.get('year_month') or _current_year_month()
        duck_db.compute_funnel_metrics(year_month)
        return jsonify({'success': True, 'data': {'year_month': year_month}})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/home-summary', methods=['GET'])
@require_permission('funnel_target.home_card')
def funnel_home_summary():
    try:
        year_month = request.args.get('year_month') or _current_year_month()
        data = _get_funnel_summary(year_month)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/overview', methods=['GET'])
@require_permission('funnel_target.view')
def funnel_overview():
    try:
        year_month = request.args.get('year_month') or _current_year_month()
        return jsonify({'success': True, 'data': _get_funnel_summary(year_month)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/dashboard-summary', methods=['GET'])
@require_permission('funnel_target.view')
def funnel_dashboard_summary():
    try:
        duck_db.ensure_funnel_schema()
        year_month, where_sql, params = _funnel_filters(include_model=False)
        rows = _get_funnel_org_dealer_rows(year_month, where_sql, params)
        progress = duck_db._progress_ratios(year_month)
        latest_lead_date = progress.get('latest_lead_date')
        if isinstance(latest_lead_date, (datetime, date)):
            latest_lead_date = latest_lead_date.strftime('%Y-%m-%d')
        status_order = ['领先', '正常', '轻度落后', '严重落后', '配置异常']
        status_counts = {status: 0 for status in status_order}
        diagnosis_counts = {}
        for row in rows:
            status = row.get('progress_status') or '正常'
            status_counts[status] = status_counts.get(status, 0) + 1
            for tag in row.get('diagnosis_tags') or []:
                diagnosis_counts[tag] = diagnosis_counts.get(tag, 0) + 1
        total_visit = sum(float(row.get('visit_count') or 0) for row in rows)
        target_to_date = sum(float(row.get('dealer_visit_target_to_date') or 0) for row in rows)
        derived_to_date = sum(float(row.get('derived_visit_target_to_date') or 0) for row in rows)
        return jsonify({
            'success': True,
            'data': {
                'year_month': year_month,
                'dealer_count': len(rows),
                'visit_count': total_visit,
                'visit_target_to_date': target_to_date,
                'visit_achievement_rate': total_visit * 100.0 / target_to_date if target_to_date else 0,
                'visit_gap': total_visit - target_to_date if target_to_date else 0,
                'derived_visit_target_to_date': derived_to_date,
                'derived_achievement_rate': total_visit * 100.0 / derived_to_date if derived_to_date else 0,
                'elapsed_day_ratio': progress.get('time_progress_ratio') or 0,
                'data_progress_ratio': progress.get('data_progress_ratio') or 0,
                'latest_lead_date': latest_lead_date,
                'status_counts': [{'status': status, 'count': status_counts.get(status, 0)} for status in status_order],
                'diagnosis_counts': [{'tag': tag, 'count': count} for tag, count in sorted(diagnosis_counts.items(), key=lambda item: item[1], reverse=True)],
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/dashboard-regions', methods=['GET'])
@require_permission('funnel_target.view')
def funnel_dashboard_regions():
    try:
        duck_db.ensure_funnel_schema()
        level = request.args.get('level') or 'region'
        year_month, where_sql, params = _funnel_filters(include_model=False)
        rows = _get_funnel_org_dealer_rows(year_month, where_sql, params)
        groups = {}
        for row in rows:
            key = row.get('zone') if level == 'zone' else row.get('region')
            key = key or '未归属'
            item = groups.setdefault(key, {
                'name': key,
                'level': level,
                'dealer_count': 0,
                'visit_count': 0,
                'visit_target_to_date': 0,
                'visit_gap': 0,
                'derived_visit_target_to_date': 0,
                'light_lagging_count': 0,
                'serious_lagging_count': 0,
                'config_error_count': 0,
            })
            item['dealer_count'] += 1
            item['visit_count'] += float(row.get('visit_count') or 0)
            item['visit_target_to_date'] += float(row.get('dealer_visit_target_to_date') or 0)
            item['visit_gap'] += float(row.get('dealer_visit_gap') or 0)
            item['derived_visit_target_to_date'] += float(row.get('derived_visit_target_to_date') or 0)
            if row.get('progress_status') == '轻度落后':
                item['light_lagging_count'] += 1
            if row.get('progress_status') == '严重落后':
                item['serious_lagging_count'] += 1
            if row.get('progress_status') == '配置异常':
                item['config_error_count'] += 1
        result = []
        for item in groups.values():
            item['visit_achievement_rate'] = item['visit_count'] * 100.0 / item['visit_target_to_date'] if item['visit_target_to_date'] else 0
            item['derived_achievement_rate'] = item['visit_count'] * 100.0 / item['derived_visit_target_to_date'] if item['derived_visit_target_to_date'] else 0
            result.append(item)
        result.sort(key=lambda item: item['visit_gap'])
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/dealer-models', methods=['GET'])
@require_permission('funnel_target.view')
def funnel_dealer_models():
    try:
        duck_db.ensure_funnel_schema()
        year_month, where_sql, params = _funnel_filters()
        page = int(request.args.get('page', 1))
        page_size = min(int(request.args.get('page_size', 50)), 200)
        offset = (page - 1) * page_size
        sort_by = request.args.get('sort_by') or 'derived_visit_gap'
        sort_order = 'ASC' if request.args.get('sort_order', 'asc').lower() == 'asc' else 'DESC'
        allowed_sort = {'derived_visit_gap', 'dealer_visit_gap', 'visit_count', 'online_lead_count', 'sales_target', 'dealer_id'}
        if sort_by not in allowed_sort:
            sort_by = 'sales_target'
            sort_order = 'DESC'
        conn = duck_db.get_connection()
        total = conn.execute(f"SELECT COUNT(*) FROM funnel_metric_targets WHERE {where_sql}", params).fetchone()[0]
        rows = _rows_to_dicts(conn.execute(f"""
            SELECT * FROM funnel_metric_targets
            WHERE {where_sql}
            ORDER BY {sort_by} {sort_order} NULLS LAST
            LIMIT ? OFFSET ?
        """, params + [page_size, offset]))
        _decorate_rows_with_store_status(rows)
        return jsonify({'success': True, 'data': rows, 'pagination': {'page': page, 'page_size': page_size, 'total': total, 'total_pages': (total + page_size - 1) // page_size}})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/org-dealers', methods=['GET'])
@require_permission('funnel_target.view')
def funnel_org_dealers():
    try:
        duck_db.ensure_funnel_schema()
        year_month, where_sql, params = _funnel_filters(include_model=False)
        rows = _get_funnel_org_dealer_rows(year_month, where_sql, params, limit=500)
        _decorate_rows_with_store_status(rows)
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/channels', methods=['GET'])
@require_permission('funnel_target.view')
def funnel_channels():
    try:
        duck_db.ensure_funnel_schema()
        year_month, where_sql, params = _funnel_filters(include_channels=True)
        conn = duck_db.get_connection()
        rows = _rows_to_dicts(conn.execute(f"""
            SELECT * FROM funnel_metric_monthly
            WHERE {where_sql}
            ORDER BY dealer_id, model_name, online_lead_count DESC, visit_count DESC
            LIMIT 1000
        """, params))
        _decorate_rows_with_store_status(rows)
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/funnel-target/filter-options', methods=['GET'])
@require_permission('funnel_target.view')
def funnel_filter_options():
    try:
        duck_db.ensure_funnel_schema()
        conn = duck_db.get_connection()
        year_month = request.args.get('year_month') or _current_year_month()
        region = request.args.get('region', '').strip()
        forced_owner = _current_user_forced_lead_owner()
        zone_filters = ["zone IS NOT NULL", "TRIM(zone) != ''"]
        zone_params = []
        if region:
            zone_filters.append("region = ?")
            zone_params.append(region)
        if forced_owner:
            zone_filters.append("lead_ops_owner = ?")
            zone_params.append(forced_owner)
        dealer_filters = ["dealer_id IS NOT NULL", "TRIM(dealer_id) != ''"]
        dealer_params = []
        if forced_owner:
            dealer_filters.append("lead_ops_owner = ?")
            dealer_params.append(forced_owner)
        owner_filter_sql = "AND lead_ops_owner = ?" if forced_owner else ""
        owner_filter_params = [forced_owner] if forced_owner else []
        options = {
            'regions': [row[0] for row in conn.execute("""
                SELECT DISTINCT region FROM mart_dealers
                WHERE region IS NOT NULL AND TRIM(region) != ''
                """ + (" AND lead_ops_owner = ?" if forced_owner else "") + """
                ORDER BY region
            """, owner_filter_params).fetchall()],
            'zones': [row[0] for row in conn.execute("""
                SELECT DISTINCT zone FROM mart_dealers
                WHERE """ + " AND ".join(zone_filters) + """
                ORDER BY zone
            """, zone_params).fetchall()],
        }
        region_zone_rows = conn.execute("""
            SELECT DISTINCT region, zone FROM mart_dealers
            WHERE region IS NOT NULL AND TRIM(region) != ''
              AND zone IS NOT NULL AND TRIM(zone) != ''
              """ + (" AND lead_ops_owner = ?" if forced_owner else "") + """
            ORDER BY region, zone
        """, owner_filter_params).fetchall()
        region_zones = {}
        for region_value, zone_value in region_zone_rows:
            region_zones.setdefault(region_value, set()).add(zone_value)
        options['region_zones'] = {key: sorted(value) for key, value in region_zones.items()}
        for key, column in [
            ('models', 'model_name'),
            ('channel_2', 'channel_2'),
            ('channel_3', 'channel_3'),
            ('lead_ops_owners', 'lead_ops_owner'),
            ('lead_ops_supports', 'lead_ops_support'),
        ]:
            options[key] = [row[0] for row in conn.execute(f"""
                SELECT DISTINCT {column} FROM funnel_metric_monthly
                WHERE year_month = ? AND {column} IS NOT NULL AND TRIM({column}) != ''
                ORDER BY {column}
            """, [year_month]).fetchall()]
        options['lead_ops_owner_options'] = [
            {'name': row[0], 'dealer_count': row[1]}
            for row in conn.execute(f"""
                SELECT lead_ops_owner, COUNT(DISTINCT dealer_id) AS dealer_count
                FROM mart_dealers
                WHERE lead_ops_owner IS NOT NULL AND TRIM(lead_ops_owner) != ''
                {owner_filter_sql}
                GROUP BY lead_ops_owner
                ORDER BY dealer_count DESC, lead_ops_owner
            """, owner_filter_params).fetchall()
        ]
        options['dealers'] = [
            {'dealer_id': row[0], 'dealer_name': row[1], 'region': row[2], 'zone': row[3]}
            for row in conn.execute("""
                SELECT dealer_id, dealer_name, region, zone
                FROM mart_dealers
                WHERE """ + " AND ".join(dealer_filters) + """
                ORDER BY dealer_id
                LIMIT 2000
            """, dealer_params).fetchall()
        ]
        return jsonify({'success': True, 'data': options})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


def _overdue_csv_values(value):
    if not value:
        return []
    return [item.strip() for item in value.split(',') if item.strip()]


def _overdue_datetime_sql(expression):
    return f"""
        TRY_CAST(NULLIF(TRIM(
            CASE
                WHEN CAST({expression} AS VARCHAR) LIKE '%/%'
                    THEN replace(CAST({expression} AS VARCHAR), '/', '-')
                ELSE CAST({expression} AS VARCHAR)
            END
        ), '') AS TIMESTAMP)
    """


def _overdue_base_sql():
    sqlite_path = str(RAW_DB_PATH).replace("'", "''")
    assign_time = _overdue_datetime_sql('s."最终下发时间"')
    cutoff_time = _overdue_datetime_sql('s."跟进截止时间"')
    first_follow_time = _overdue_datetime_sql('s."首跟时间"')
    follow2_time = _overdue_datetime_sql('s."二跟时间"')
    follow3_time = _overdue_datetime_sql('s."三跟时间"')
    followup_created_time = _overdue_datetime_sql('f."创建时间"')

    return f"""
        WITH lead_base AS (
            SELECT
                CAST(s."id" AS VARCHAR) AS lead_id,
                COALESCE(NULLIF(TRIM(CAST(s."大区" AS VARCHAR)), ''), CAST(d."大区" AS VARCHAR)) AS region,
                CAST(d."战区" AS VARCHAR) AS zone,
                CAST(s."门店" AS VARCHAR) AS dealer_id,
                COALESCE(NULLIF(TRIM(CAST(s."店简称" AS VARCHAR)), ''), CAST(d."店简称" AS VARCHAR)) AS dealer_name,
                {assign_time} AS assign_time,
                {cutoff_time} AS follow_cutoff_time,
                CAST(s."是否及时跟进" AS VARCHAR) AS timely_follow_text,
                {first_follow_time} AS first_follow_time,
                {follow2_time} AS follow2_time,
                {follow3_time} AS follow3_time,
                CAST(s."线索状态" AS VARCHAR) AS lead_status,
                CAST(s."一级渠道" AS VARCHAR) AS channel_1,
                CAST(s."二级渠道" AS VARCHAR) AS channel_2,
                CAST(s."三级渠道" AS VARCHAR) AS channel_3
            FROM sqlite_scan('{sqlite_path}', '线索表') s
            LEFT JOIN sqlite_scan('{sqlite_path}', '门店表') d
                ON CAST(s."门店" AS VARCHAR) = CAST(d."店编号" AS VARCHAR)
        ),
        first_follow_user AS (
            SELECT
                sub.lead_id,
                sub.follower,
                sub.follower_id
            FROM (
                SELECT
                    CAST(f."门店线索id" AS VARCHAR) AS lead_id,
                    COALESCE(NULLIF(TRIM(CAST(p."姓名" AS VARCHAR)), ''), CAST(f."跟进人" AS VARCHAR)) AS follower,
                    CAST(f."跟进人" AS VARCHAR) AS follower_id,
                    ROW_NUMBER() OVER (
                        PARTITION BY CAST(f."门店线索id" AS VARCHAR)
                        ORDER BY {followup_created_time} ASC NULLS LAST, CAST(f."id" AS VARCHAR) ASC
                    ) AS rn
                FROM sqlite_scan('{sqlite_path}', '跟进表') f
                LEFT JOIN sqlite_scan('{sqlite_path}', '人员表') p
                    ON CAST(f."跟进人" AS VARCHAR) = CAST(p."员工编号" AS VARCHAR)
            ) sub
            WHERE sub.rn = 1
        ),
        overdue AS (
            SELECT
                l.*,
                ffu.follower,
                ffu.follower_id
            FROM lead_base l
            LEFT JOIN first_follow_user ffu
                ON l.lead_id = ffu.lead_id
            WHERE l.channel_1 = '线上'
              AND l.follow_cutoff_time IS NOT NULL
              AND l.timely_follow_text = '否'
        )
    """


def _build_overdue_filters(args):
    where = []
    params = []

    start_date = args.get('start_date', '')
    end_date = args.get('end_date', '')
    if not start_date or not end_date:
        raise ValueError('请选择开始日期和结束日期')
    if start_date > end_date:
        raise ValueError('开始日期不能晚于结束日期')

    where.append("assign_date >= CAST(? AS DATE)")
    params.append(start_date)
    where.append("assign_date <= CAST(? AS DATE)")
    params.append(end_date)

    regions = _overdue_csv_values(args.get('regions', ''))
    if not regions and args.get('region'):
        regions = [args.get('region')]
    if regions:
        where.append(f"region IN ({','.join(['?'] * len(regions))})")
        params.extend(regions)

    zones = _overdue_csv_values(args.get('zones', ''))
    if not zones and args.get('zone'):
        zones = [args.get('zone')]
    if zones:
        where.append(f"zone IN ({','.join(['?'] * len(zones))})")
        params.extend(zones)

    dealer_id = args.get('dealer_id', '').strip()
    if dealer_id:
        where.append("dealer_id = ?")
        params.append(dealer_id)

    dealer_name = args.get('dealer_name', '').strip()
    if dealer_name:
        where.append("dealer_name LIKE ?")
        params.append(f"%{dealer_name}%")

    return " AND ".join(where), params, start_date, end_date


def _format_overdue_value(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return value if value is not None else ''


def _overdue_rows_to_dicts(rows):
    columns = [
        'region', 'zone', 'dealer_id', 'dealer_name', 'lead_id', 'phone',
        'assign_time', 'follow_cutoff_time', 'timely_follow_text',
        'first_follow_time', 'follow2_time', 'follow3_time', 'lead_status',
        'channel_1', 'channel_2', 'channel_3', 'follower', 'follower_id'
    ]
    return [
        {column: _format_overdue_value(row[idx]) for idx, column in enumerate(columns)}
        for row in rows
    ]


def _query_overdue_data(conn, args, include_pagination=True):
    where_sql, params, start_date, end_date = _build_overdue_filters(args)

    valid_sort_columns = {
        'assign_time': 'assign_time',
        'follow_cutoff_time': 'follow_cutoff_time',
        'first_follow_time': 'first_follow_time',
        'dealer_id': 'dealer_id',
        'dealer_name': 'dealer_name',
        'region': 'region',
        'zone': 'zone',
    }
    sort_by = args.get('sort_by', 'assign_time')
    sort_column = valid_sort_columns.get(sort_by, 'assign_time')
    sort_order = 'ASC' if args.get('sort_order', 'desc').lower() == 'asc' else 'DESC'

    summary_sql = f"""
        SELECT
            COUNT(*) AS overdue_count,
            COUNT(DISTINCT dealer_id) AS dealer_count,
            SUM(CASE WHEN first_follow_time IS NOT NULL THEN 1 ELSE 0 END) AS first_followed_count,
            SUM(CASE WHEN first_follow_time IS NULL THEN 1 ELSE 0 END) AS not_first_followed_count
        FROM mart_dealer_overdue_leads
        WHERE {where_sql}
    """
    summary_row = conn.execute(summary_sql, params).fetchone()
    summary = {
        'overdue_count': int(summary_row[0] or 0),
        'dealer_count': int(summary_row[1] or 0),
        'first_followed_count': int(summary_row[2] or 0),
        'not_first_followed_count': int(summary_row[3] or 0),
    }

    total = summary['overdue_count']
    page = max(int(args.get('page', 1)), 1)
    page_size = min(max(int(args.get('page_size', 50)), 1), 500)
    offset = (page - 1) * page_size

    data_sql = f"""
        SELECT
            region, zone, dealer_id, dealer_name, lead_id,
            phone,
            assign_time, follow_cutoff_time, timely_follow_text,
            first_follow_time, follow2_time, follow3_time, lead_status,
            channel_1, channel_2, channel_3, follower, follower_id
        FROM mart_dealer_overdue_leads
        WHERE {where_sql}
        ORDER BY {sort_column} {sort_order} NULLS LAST, lead_id ASC
    """
    data_params = list(params)
    if include_pagination:
        data_sql += " LIMIT ? OFFSET ?"
        data_params.extend([page_size, offset])

    rows = conn.execute(data_sql, data_params).fetchall()

    region_zone_filters = _overdue_region_zone_filters(conn, start_date, end_date)

    result = {
        'summary': summary,
        'items': _overdue_rows_to_dicts(rows),
        'filters': region_zone_filters,
    }
    if include_pagination:
        result['pagination'] = {
            'page': page,
            'page_size': page_size,
            'total': total,
            'total_pages': (total + page_size - 1) // page_size if page_size else 0,
        }
    return result


@app.route('/api/dealer-management/overdue-query', methods=['GET'])
def get_dealer_overdue_query():
    try:
        duck_db.ensure_dealer_overdue_data()
        conn = duck_db.get_connection()
        data = _query_overdue_data(conn, request.args, include_pagination=True)
        return jsonify({'success': True, 'data': data})
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/dealer-management/overdue-query/export', methods=['GET'])
def export_dealer_overdue_query():
    try:
        import openpyxl

        duck_db.ensure_dealer_overdue_data()
        conn = duck_db.get_connection()
        data = _query_overdue_data(conn, request.args, include_pagination=False)
        rows = data['items']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '逾期查询'

        headers = [
            '大区', '战区', '店编号', '店简称', '线索ID',
            '手机', '线索最终下发时间', '首跟截止时间', '是否及时跟进',
            '首跟时间', '二跟时间', '三跟时间', '线索状态',
            '一级渠道', '二级渠道', '三级渠道', '跟进人'
        ]
        ws.append(headers)

        for item in rows:
            ws.append([
                item['region'], item['zone'], item['dealer_id'], item['dealer_name'], item['lead_id'],
                item['phone'],
                item['assign_time'], item['follow_cutoff_time'], item['timely_follow_text'] or '否',
                item['first_follow_time'], item['follow2_time'], item['follow3_time'], item['lead_status'],
                item['channel_1'], item['channel_2'], item['channel_3'], item['follower'],
            ])

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        store_status = request.args.get('store_status', '')
        dealer_name_filter = request.args.get('dealer_name', '').strip()
        dealer_id_filter = request.args.get('dealer_id', '').strip()
        row_dealer_names = sorted({item.get('dealer_name') for item in rows if item.get('dealer_name')})
        if dealer_name_filter:
            dealer_label = dealer_name_filter
        elif len(row_dealer_names) == 1:
            dealer_label = row_dealer_names[0]
        elif dealer_id_filter:
            dealer_label = dealer_id_filter
        else:
            dealer_label = '全部门店'
        safe_dealer_label = ''.join(ch if ch not in r'\/:*?"<>|' else '_' for ch in dealer_label).strip() or '全部门店'
        filename = f"逾期查询_{safe_dealer_label}_{start_date}_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/dealer-daily-report', methods=['GET'])
def get_dealer_daily_report():
    try:
        period = request.args.get('period', 'daily')
        region = request.args.get('region', '')
        zone = request.args.get('zone', '')
        dealer_id = request.args.get('dealer_id', '')
        dealer_name = request.args.get('dealer_name', '')
        store_status = request.args.get('store_status', '')
        sort_by = request.args.get('sort_by', 'lead_count')
        sort_order = request.args.get('sort_order', 'desc')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 50))
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        conn = duck_db.get_connection()

        has_custom_range = start_date and end_date

        if has_custom_range:
            return _get_dealer_report_custom_range(
                conn, start_date, end_date, region, zone,
                dealer_id, dealer_name,
                sort_by, sort_order, page, page_size, store_status
            )
        else:
            return _get_dealer_report_precomputed(
                conn, period, region, zone,
                dealer_id, dealer_name,
                sort_by, sort_order, page, page_size, store_status
            )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


STORE_PORTFOLIO_MONITOR_METRICS = {
    'n60_lead_count', 'n60_follow_30min_count', 'n60_follow_rate',
    'lead_count', 'follow_30min_count', 'follow_30min_task_count', 'follow_30min_rate',
    'day3_3follow_task_count', 'day3_3follow_count', 'day3_3follow_rate',
    'valid_lead_count', 'valid_lead_rate', 'valid_local_lead_count', 'local_lead_count',
    'new_media_self_valid_lead_count', 'new_media_self_lead_count',
    'to_shop_count', 'lead_to_shop_rate', 'local_lead_to_shop_rate',
    'valid_lead_to_shop_rate', 'valid_local_lead_to_shop_rate',
    'online_sales_count', 'online_sales_rate', 'to_shop_conversion_rate',
    'expected_to_shop', 'to_shop_diff',
}


def _with_store_portfolio_derived_metrics(rows):
    for row in rows:
        n60_leads = float(row.get('n60_lead_count') or 0)
        n60_follow = float(row.get('n60_follow_30min_count') or 0)
        row['n60_follow_rate'] = n60_follow * 100.0 / n60_leads if n60_leads else 0
    return rows


@app.route('/api/store-portfolio-monitor/stores', methods=['GET'])
def store_portfolio_monitor_stores():
    try:
        period = request.args.get('period', 'daily')
        region = request.args.get('region', '')
        zone = request.args.get('zone', '')
        dealer_id = request.args.get('dealer_id', '')
        dealer_name = request.args.get('dealer_name', '')
        store_status = request.args.get('store_status', '')
        sort_by = request.args.get('sort_by', 'lead_count')
        sort_order = request.args.get('sort_order', 'desc')
        page = int(request.args.get('page', 1))
        page_size = min(int(request.args.get('page_size', 500)), 1000)
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        conn = duck_db.get_connection()
        if start_date and end_date:
            response = _get_dealer_report_custom_range(
                conn, start_date, end_date, region, zone,
                dealer_id, dealer_name, sort_by, sort_order, page, page_size, store_status
            )
        else:
            response = _get_dealer_report_precomputed(
                conn, period, region, zone,
                dealer_id, dealer_name,
                sort_by, sort_order, page, page_size, store_status
            )

        payload = response.get_json()
        if payload and payload.get('success'):
            payload['data'] = _with_store_portfolio_derived_metrics(payload.get('data') or [])
            return jsonify(payload), response.status_code
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/store-portfolio-monitor/daily-series', methods=['GET'])
def store_portfolio_monitor_daily_series():
    try:
        dealer_id = request.args.get('dealer_id', '').strip()
        metric = request.args.get('metric', 'lead_count').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        store_status = request.args.get('store_status', '')

        if not dealer_id:
            return jsonify({'success': False, 'message': '缺少门店编码'}), 400
        if metric not in STORE_PORTFOLIO_MONITOR_METRICS:
            return jsonify({'success': False, 'message': '不支持的指标'}), 400
        if not start_date or not end_date:
            return jsonify({'success': False, 'message': '缺少日期范围'}), 400

        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        if start_dt > end_dt:
            return jsonify({'success': False, 'message': '开始日期不能晚于结束日期'}), 400
        if (end_dt - start_dt).days > 60:
            return jsonify({'success': False, 'message': '日变化图最多支持60天'}), 400

        conn = duck_db.get_connection()
        rows = []
        current = start_dt
        while current <= end_dt:
            day = current.strftime('%Y-%m-%d')
            response = _get_dealer_report_custom_range(
                conn, day, day, '', '', dealer_id, '',
                metric if metric != 'n60_follow_rate' else 'n60_follow_30min_count',
                'desc', 1, 1, store_status
            )
            payload = response.get_json() or {}
            data = _with_store_portfolio_derived_metrics(payload.get('data') or [])
            row = data[0] if data else {}
            rows.append({
                'day': day,
                'value': float(row.get(metric) or 0),
                'row': row,
            })
            current += timedelta(days=1)

        return jsonify({'success': True, 'data': rows})
    except ValueError:
        return jsonify({'success': False, 'message': '日期格式错误'}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


def _dealer_report_precomputed_sort_expr(sort_by, prefix):
    """Return a non-ambiguous ORDER BY expression for precomputed daily reports."""
    sort_columns = {
        'dealer_name': 'md.dealer_name',
        'region': 'md.region',
        'zone': 'md.zone',
        'n60_lead_count': f'COALESCE(r.{prefix}n60_lead_count, 0)',
        'n60_follow_30min_count': f'COALESCE(r.{prefix}n60_follow_30min_count, 0)',
        'lead_count': f'COALESCE(r.{prefix}lead_count, 0)',
        'follow_30min_count': f'COALESCE(r.{prefix}follow_30min_count, 0)',
        'follow_30min_rate': f'COALESCE(r.{prefix}follow_30min_rate, 0)',
        'day3_3follow_task_count': f'COALESCE(r.{prefix}3day_3follow_task_count, 0)',
        'day3_3follow_count': f'COALESCE(r.{prefix}3day_3follow_count, 0)',
        'day3_3follow_rate': f'COALESCE(r.{prefix}3day_3follow_rate, 0)',
        'valid_lead_count': f'COALESCE(r.{prefix}valid_lead_count, 0)',
        'valid_lead_rate': f'COALESCE(r.{prefix}valid_lead_rate, 0)',
        'valid_local_lead_count': f'COALESCE(r.{prefix}valid_local_lead_count, 0)',
        'local_lead_count': f'COALESCE(r.{prefix}local_lead_count, 0)',
        'to_shop_count': f'COALESCE(r.{prefix}to_shop_count, 0)',
        'lead_to_shop_rate': f'COALESCE(r.{prefix}lead_to_shop_rate, 0)',
        'local_lead_to_shop_rate': f'COALESCE(r.{prefix}local_lead_to_shop_rate, 0)',
        'valid_lead_to_shop_rate': f'COALESCE(r.{prefix}valid_lead_to_shop_rate, 0)',
        'valid_local_lead_to_shop_rate': f'COALESCE(r.{prefix}valid_local_lead_to_shop_rate, 0)',
        'new_media_self_valid_lead_count': f'COALESCE(r.{prefix}new_media_self_valid_lead_count, 0)',
        'new_media_self_lead_count': f'COALESCE(r.{prefix}new_media_self_lead_count, 0)',
        'online_sales_count': 'COALESCE(r.m_online_sales_count, 0)',
        'online_sales_rate': 'COALESCE(r.m_online_sales_rate, 0)',
        'to_shop_conversion_rate': 'COALESCE(r.m_to_shop_conversion_rate, NULL)',
        'expected_to_shop': 'COALESCE(r.m_expected_to_shop, 0)',
        'to_shop_diff': 'COALESCE(r.m_to_shop_diff, 0)',
        'to_shop_eval': "COALESCE(r.m_to_shop_eval, '无')",
    }
    return sort_columns.get(sort_by, f'COALESCE(r.{prefix}lead_count, 0)')


def _dealer_report_aggregate_sort_expr(sort_by):
    """Return an ORDER BY expression for aggregate custom-range report queries."""
    sort_columns = {
        'dealer_name': 'dealer_name',
        'region': 'region',
        'zone': 'zone',
        'n60_lead_count': 'n60_lead_count',
        'n60_follow_30min_count': 'n60_follow_30min_count',
        'lead_count': 'lead_count',
        'follow_30min_count': 'follow_30min_count',
        'follow_30min_rate': 'follow_30min_rate',
        'day3_3follow_task_count': 'day3_3follow_task_count',
        'day3_3follow_count': 'day3_3follow_count',
        'day3_3follow_rate': 'day3_3follow_rate',
        'valid_lead_count': 'valid_lead_count',
        'valid_lead_rate': 'valid_lead_rate',
        'valid_local_lead_count': 'valid_local_lead_count',
        'local_lead_count': 'local_lead_count',
        'to_shop_count': 'to_shop_count',
        'lead_to_shop_rate': 'lead_to_shop_rate',
        'local_lead_to_shop_rate': 'local_lead_to_shop_rate',
        'valid_lead_to_shop_rate': 'valid_lead_to_shop_rate',
        'valid_local_lead_to_shop_rate': 'valid_local_lead_to_shop_rate',
        'new_media_self_valid_lead_count': 'new_media_self_valid_lead_count',
        'new_media_self_lead_count': 'new_media_self_lead_count',
        'online_sales_count': 'online_sales_count',
        'online_sales_rate': 'online_sales_rate',
        'to_shop_conversion_rate': 'to_shop_conversion_rate',
        'expected_to_shop': 'expected_to_shop',
        'to_shop_diff': 'to_shop_diff',
        'to_shop_eval': 'to_shop_eval',
    }
    return sort_columns.get(sort_by, 'lead_count')


def _summarize_dealer_report_rows(rows):
    """Aggregate dealer-level report rows without duplicating per-dealer visit/sales counts."""
    def total(key):
        return sum(float(row.get(key) or 0) for row in rows)

    if not rows:
        return {key: 0 for key in [
            'n60_lead_count', 'n60_follow_30min_count',
            'lead_count', 'follow_30min_count', 'follow_30min_task_count',
            'follow_30min_rate',
            'day3_3follow_task_count', 'day3_3follow_count', 'day3_3follow_rate',
            'valid_lead_count', 'valid_lead_rate',
            'valid_local_lead_count', 'local_lead_count',
            'to_shop_count', 'lead_to_shop_rate',
            'local_lead_to_shop_rate', 'valid_lead_to_shop_rate',
            'valid_local_lead_to_shop_rate',
            'new_media_self_valid_lead_count', 'new_media_self_lead_count',
            'online_sales_count', 'online_sales_rate',
            'expected_to_shop', 'to_shop_diff',
        ]} | {'to_shop_conversion_rate': None, 'to_shop_eval': None}

    n60_lead_count = total('n60_lead_count')
    n60_follow_30min_count = total('n60_follow_30min_count')
    lead_count = total('lead_count')
    follow_30min_count = total('follow_30min_count')
    follow_30min_task_count = total('follow_30min_task_count')
    day3_3follow_task_count = total('day3_3follow_task_count')
    day3_3follow_count = total('day3_3follow_count')
    valid_lead_count = total('valid_lead_count')
    valid_local_lead_count = total('valid_local_lead_count')
    local_lead_count = total('local_lead_count')
    to_shop_count = total('to_shop_count')
    new_media_self_valid_lead_count = total('new_media_self_valid_lead_count')
    new_media_self_lead_count = total('new_media_self_lead_count')
    online_sales_count = total('online_sales_count')
    expected_to_shop = online_sales_count * 4.0
    to_shop_diff = to_shop_count - expected_to_shop

    def rate(numerator, denominator):
        return numerator * 100.0 / denominator if denominator else 0

    if online_sales_count == 0:
        to_shop_eval = '无'
    elif to_shop_count > 2 * expected_to_shop:
        to_shop_eval = '到店转化率低'
    elif to_shop_count >= 0.6 * expected_to_shop:
        to_shop_eval = '正常'
    else:
        to_shop_eval = '到店录入存在问题'

    return {
        'n60_lead_count': n60_lead_count,
        'n60_follow_30min_count': n60_follow_30min_count,
        'lead_count': lead_count,
        'follow_30min_count': follow_30min_count,
        'follow_30min_task_count': follow_30min_task_count,
        'follow_30min_rate': rate(follow_30min_count, follow_30min_task_count),
        'day3_3follow_task_count': day3_3follow_task_count,
        'day3_3follow_count': day3_3follow_count,
        'day3_3follow_rate': rate(day3_3follow_count, day3_3follow_task_count),
        'valid_lead_count': valid_lead_count,
        'valid_lead_rate': rate(valid_lead_count, lead_count),
        'valid_local_lead_count': valid_local_lead_count,
        'local_lead_count': local_lead_count,
        'to_shop_count': to_shop_count,
        'lead_to_shop_rate': rate(to_shop_count, lead_count),
        'local_lead_to_shop_rate': rate(to_shop_count, local_lead_count),
        'valid_lead_to_shop_rate': rate(to_shop_count, valid_lead_count),
        'valid_local_lead_to_shop_rate': rate(to_shop_count, valid_local_lead_count),
        'new_media_self_valid_lead_count': new_media_self_valid_lead_count,
        'new_media_self_lead_count': new_media_self_lead_count,
        'online_sales_count': online_sales_count,
        'online_sales_rate': rate(online_sales_count, local_lead_count),
        'to_shop_conversion_rate': (online_sales_count * 100.0 / to_shop_count) if to_shop_count else None,
        'expected_to_shop': expected_to_shop,
        'to_shop_diff': to_shop_diff,
        'to_shop_eval': to_shop_eval,
    }


def _get_dealer_report_precomputed(conn, period, region, zone, dealer_id, dealer_name, sort_by, sort_order, page, page_size, store_status=""):
    """使用预计算表查询（单日或当月累计）"""
    dealer_where = []
    dealer_params = []
    if region:
        dealer_where.append("md.region = ?")
        dealer_params.append(region)
    if zone:
        dealer_where.append("md.zone = ?")
        dealer_params.append(zone)
    if dealer_id:
        dealer_where.append("md.dealer_id LIKE ?")
        dealer_params.append(f"%{dealer_id}%")
    if dealer_name:
        dealer_where.append("md.dealer_name LIKE ?")
        dealer_params.append(f"%{dealer_name}%")
    store_status_ids = _store_status_filter_ids(store_status)
    store_status_filter, store_status_params = _dealer_id_filter_sql("md", store_status_ids)
    if store_status_filter:
        dealer_where.append(store_status_filter.replace(" AND ", "", 1))
        dealer_params.extend(store_status_params)

    dealer_filter = " AND ".join(dealer_where) if dealer_where else "1=1"

    prefix = 'm_' if period == 'monthly' else 'd_'
    sort_expr = _dealer_report_precomputed_sort_expr(sort_by, prefix)
    sort_direction = 'DESC' if sort_order.lower() == 'desc' else 'ASC'

    count_sql = f"SELECT COUNT(*) FROM mart_dealers md WHERE {dealer_filter}"
    total = conn.execute(count_sql, dealer_params).fetchone()[0]

    offset = (page - 1) * page_size

    data_sql = f"""
        SELECT
            r.report_date,
            COALESCE(r.period_type, ?) AS period_type,
            md.dealer_id,
            md.dealer_name,
            md.region,
            md.zone,
            md.province,
            md.region_manager,
            md.zone_manager,
            md.inspector,
            COALESCE(r.{prefix}n60_lead_count, 0) AS n60_lead_count,
            COALESCE(r.{prefix}n60_follow_30min_count, 0) AS n60_follow_30min_count,
            COALESCE(r.{prefix}lead_count, 0) AS lead_count,
            COALESCE(r.{prefix}follow_30min_count, 0) AS follow_30min_count,
            COALESCE(r.{prefix}follow_30min_task_count, 0) AS follow_30min_task_count,
            COALESCE(r.{prefix}follow_30min_rate, 0) AS follow_30min_rate,
            COALESCE(r.{prefix}3day_3follow_task_count, 0) AS day3_3follow_task_count,
            COALESCE(r.{prefix}3day_3follow_count, 0) AS day3_3follow_count,
            COALESCE(r.{prefix}3day_3follow_rate, 0) AS day3_3follow_rate,
            COALESCE(r.{prefix}valid_lead_count, 0) AS valid_lead_count,
            COALESCE(r.{prefix}valid_lead_rate, 0) AS valid_lead_rate,
            COALESCE(r.{prefix}valid_local_lead_count, 0) AS valid_local_lead_count,
            COALESCE(r.{prefix}local_lead_count, 0) AS local_lead_count,
            COALESCE(r.{prefix}to_shop_count, 0) AS to_shop_count,
            COALESCE(r.{prefix}lead_to_shop_rate, 0) AS lead_to_shop_rate,
            COALESCE(r.{prefix}local_lead_to_shop_rate, 0) AS local_lead_to_shop_rate,
            COALESCE(r.{prefix}valid_lead_to_shop_rate, 0) AS valid_lead_to_shop_rate,
            COALESCE(r.{prefix}valid_local_lead_to_shop_rate, 0) AS valid_local_lead_to_shop_rate,
            COALESCE(r.{prefix}new_media_self_valid_lead_count, 0) AS new_media_self_valid_lead_count,
            COALESCE(r.{prefix}new_media_self_lead_count, 0) AS new_media_self_lead_count,
            COALESCE(r.m_online_sales_count, 0) AS online_sales_count,
            COALESCE(r.m_online_sales_rate, 0) AS online_sales_rate,
            COALESCE(r.m_to_shop_conversion_rate, NULL) AS to_shop_conversion_rate,
            COALESCE(r.m_expected_to_shop, 0) AS expected_to_shop,
            COALESCE(r.m_to_shop_diff, 0) AS to_shop_diff,
            COALESCE(r.m_to_shop_eval, '无') AS to_shop_eval
        FROM mart_dealers md
        LEFT JOIN (
            SELECT * FROM (
                SELECT *, ROW_NUMBER() OVER (PARTITION BY dealer_id ORDER BY report_date DESC) AS _rn
                FROM report_dealer_daily
                WHERE period_type = ?
            ) _sub WHERE _sub._rn = 1
        ) r ON md.dealer_id = r.dealer_id
        WHERE {dealer_filter}
        ORDER BY {sort_expr} {sort_direction} NULLS LAST, md.dealer_id ASC
        LIMIT ? OFFSET ?
    """
    results = conn.execute(data_sql, [period, period] + dealer_params + [page_size, offset]).fetchall()

    summary_where = []
    summary_params = [period]
    if region:
        summary_where.append("region = ?")
        summary_params.append(region)
    if zone:
        summary_where.append("zone = ?")
        summary_params.append(zone)
    if dealer_id:
        summary_where.append("dealer_id LIKE ?")
        summary_params.append(f"%{dealer_id}%")
    if dealer_name:
        summary_where.append("dealer_name LIKE ?")
        summary_params.append(f"%{dealer_name}%")
    if store_status_filter:
        summary_where.append("CAST(dealer_id AS VARCHAR) IN (" + ", ".join(["?"] * len(store_status_params)) + ")" if store_status_params else "dealer_id IS NULL")
        summary_params.extend(store_status_params)
    summary_where_sql = " AND ".join(summary_where)
    if summary_where_sql:
        summary_where_sql = " AND " + summary_where_sql

    summary_sql = f"""
        SELECT
            SUM({prefix}n60_lead_count), SUM({prefix}n60_follow_30min_count),
            SUM({prefix}lead_count), SUM({prefix}follow_30min_count),
            SUM({prefix}follow_30min_task_count),
            CASE WHEN SUM({prefix}follow_30min_task_count) > 0
                THEN SUM({prefix}follow_30min_count) * 100.0 / SUM({prefix}follow_30min_task_count) ELSE 0 END,
            SUM({prefix}3day_3follow_task_count), SUM({prefix}3day_3follow_count),
            CASE WHEN SUM({prefix}3day_3follow_task_count) > 0
                THEN SUM({prefix}3day_3follow_count) * 100.0 / SUM({prefix}3day_3follow_task_count) ELSE 0 END,
            SUM({prefix}valid_lead_count),
            CASE WHEN SUM({prefix}lead_count) > 0
                THEN SUM({prefix}valid_lead_count) * 100.0 / SUM({prefix}lead_count) ELSE 0 END,
            SUM({prefix}valid_local_lead_count), SUM({prefix}local_lead_count),
            SUM({prefix}to_shop_count),
            CASE WHEN SUM({prefix}lead_count) > 0
                THEN SUM({prefix}to_shop_count) * 100.0 / SUM({prefix}lead_count) ELSE 0 END,
            CASE WHEN SUM({prefix}local_lead_count) > 0
                THEN SUM({prefix}to_shop_count) * 100.0 / SUM({prefix}local_lead_count) ELSE 0 END,
            CASE WHEN SUM({prefix}valid_lead_count) > 0
                THEN SUM({prefix}to_shop_count) * 100.0 / SUM({prefix}valid_lead_count) ELSE 0 END,
            CASE WHEN SUM({prefix}valid_local_lead_count) > 0
                THEN SUM({prefix}to_shop_count) * 100.0 / SUM({prefix}valid_local_lead_count) ELSE 0 END,
            SUM({prefix}new_media_self_valid_lead_count),
            SUM({prefix}new_media_self_lead_count),
            SUM(m_online_sales_count),
            CASE WHEN SUM({prefix}local_lead_count) > 0
                THEN SUM(m_online_sales_count) * 100.0 / SUM({prefix}local_lead_count) ELSE 0 END,
            CASE WHEN SUM({prefix}to_shop_count) > 0
                THEN SUM(m_online_sales_count) * 100.0 / SUM({prefix}to_shop_count) ELSE NULL END,
            SUM(m_online_sales_count) * 4.0,
            SUM({prefix}to_shop_count) - SUM(m_online_sales_count) * 4.0,
            NULL
        FROM (
            SELECT * FROM (
                SELECT *, ROW_NUMBER() OVER (PARTITION BY dealer_id ORDER BY report_date DESC) AS _rn
                FROM report_dealer_daily
                WHERE period_type = ?
            ) _sub WHERE _sub._rn = 1
        ) _dedup
        WHERE 1=1{summary_where_sql}
    """
    summary_result = conn.execute(summary_sql, summary_params).fetchone()

    columns = [
        'report_date', 'period_type', 'dealer_id', 'dealer_name',
        'region', 'zone', 'province', 'region_manager', 'zone_manager', 'inspector',
        'n60_lead_count', 'n60_follow_30min_count',
        'lead_count', 'follow_30min_count', 'follow_30min_task_count', 'follow_30min_rate',
        'day3_3follow_task_count', 'day3_3follow_count', 'day3_3follow_rate',
        'valid_lead_count', 'valid_lead_rate',
        'valid_local_lead_count', 'local_lead_count',
        'to_shop_count', 'lead_to_shop_rate',
        'local_lead_to_shop_rate', 'valid_lead_to_shop_rate',
        'valid_local_lead_to_shop_rate',
        'new_media_self_valid_lead_count', 'new_media_self_lead_count',
        'online_sales_count', 'online_sales_rate', 'to_shop_conversion_rate',
        'expected_to_shop', 'to_shop_diff', 'to_shop_eval'
    ]

    data = []
    for row in results:
        data.append(dict(zip(columns, row)))
    _decorate_rows_with_store_status(data)

    summary = None
    if summary_result:
        summary_keys = [
            'n60_lead_count', 'n60_follow_30min_count',
            'lead_count', 'follow_30min_count', 'follow_30min_task_count', 'follow_30min_rate',
            'day3_3follow_task_count', 'day3_3follow_count', 'day3_3follow_rate',
            'valid_lead_count', 'valid_lead_rate',
            'valid_local_lead_count', 'local_lead_count',
            'to_shop_count', 'lead_to_shop_rate',
            'local_lead_to_shop_rate', 'valid_lead_to_shop_rate',
            'valid_local_lead_to_shop_rate',
            'new_media_self_valid_lead_count', 'new_media_self_lead_count',
            'online_sales_count', 'online_sales_rate', 'to_shop_conversion_rate',
            'expected_to_shop', 'to_shop_diff', 'to_shop_eval'
        ]
        summary = dict(zip(summary_keys, summary_result))

    region_zone_filters = _mart_dealer_region_zone_filters(conn)

    return jsonify({
        'success': True,
        'data': data,
        'summary': summary,
        'pagination': {
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size if page_size > 0 else 1
        },
        'filters': region_zone_filters
    })


def _get_dealer_report_custom_range(conn, start_date, end_date, region, zone, dealer_id, dealer_name, sort_by, sort_order, page, page_size, store_status=""):
    """自定义日期范围实时计算"""
    date_where = f"l.assign_date >= DATE '{start_date}' AND l.assign_date <= DATE '{end_date}'"
    visit_date_where = f"visit_date >= DATE '{start_date}' AND visit_date <= DATE '{end_date}'"

    dealer_where = []
    dealer_sql_where = []
    dealer_params = []
    if region:
        dealer_where.append("d.region = ?")
        dealer_sql_where.append("d.region = ?")
        dealer_params.append(region)
    if zone:
        dealer_where.append("d.zone = ?")
        dealer_sql_where.append("d.zone = ?")
        dealer_params.append(zone)
    if dealer_id:
        dealer_where.append("d.dealer_id LIKE ?")
        dealer_sql_where.append("d.dealer_id LIKE ?")
        dealer_params.append(f"%{dealer_id}%")
    if dealer_name:
        dealer_where.append("d.dealer_name LIKE ?")
        dealer_sql_where.append("d.dealer_name LIKE ?")
        dealer_params.append(f"%{dealer_name}%")
    store_status_ids = _store_status_filter_ids(store_status)
    store_status_filter, store_status_params = _dealer_id_filter_sql("d", store_status_ids)
    if store_status_filter:
        condition = store_status_filter.replace(" AND ", "", 1)
        dealer_where.append(condition)
        dealer_sql_where.append(condition)
        dealer_params.extend(store_status_params)

    dealer_filter = " AND ".join(dealer_where)
    if dealer_filter:
        dealer_filter = " AND " + dealer_filter

    dealer_sql_filter = " AND ".join(dealer_sql_where) if dealer_sql_where else "1=1"

    sort_expr = _dealer_report_aggregate_sort_expr(sort_by)
    sort_direction = 'DESC' if sort_order.lower() == 'desc' else 'ASC'

    cutoff_time = f"{end_date} 18:00:00"
    from datetime import datetime as dt, timedelta as td
    cutoff_dt = dt.strptime(cutoff_time, "%Y-%m-%d %H:%M:%S")
    cutoff_time_72h = (cutoff_dt - td(hours=72)).strftime("%Y-%m-%d %H:%M:%S")

    count_sql = f"SELECT COUNT(*) FROM mart_dealers d WHERE {dealer_sql_filter}"
    total = conn.execute(count_sql, dealer_params).fetchone()[0]

    offset = (page - 1) * page_size

    data_sql = f"""
        WITH base AS (
            SELECT
                d.dealer_id AS _did,
                d.dealer_name, d.region, d.zone, d.province,
                d.region_manager, d.zone_manager,
                COALESCE(NULLIF(TRIM(CAST(d.region_manager AS VARCHAR)), ''), '') AS inspector,
                l.*
            FROM mart_dealers d
            LEFT JOIN mart_leads l ON d.dealer_id = l.dealer_id
                AND l.channel_1 = '线上'
                AND {date_where}
            WHERE {dealer_sql_filter}
        ),
        shop_visit AS (
            SELECT dealer_id, SUM(unique_lead_count) AS visit_count
            FROM fact_daily_visit
            WHERE period_type = 'daily'
              AND channel_1 = '线上'
              AND {visit_date_where}
            GROUP BY dealer_id
        ),
        online_sales AS (
            SELECT
                CAST(s.dealer_id AS VARCHAR) AS dealer_id,
                COUNT(*) AS sales_count
            FROM mart_online_sales s
            WHERE s.is_converted = '1'
              AND s.is_counted = '是'
              AND CAST(s.sales_date AS DATE) >= DATE '{start_date}'
              AND CAST(s.sales_date AS DATE) <= DATE '{end_date}'
            GROUP BY dealer_id
        )
        SELECT
            DATE '{end_date}' AS report_date,
            'custom' AS period_type,
            b._did AS dealer_id,
            MAX(b.dealer_name) AS dealer_name,
            MAX(b.region) AS region,
            MAX(b.zone) AS zone,
            MAX(b.province) AS province,
            MAX(b.region_manager) AS region_manager,
            MAX(b.zone_manager) AS zone_manager,
            MAX(b.inspector) AS inspector,

            SUM(CASE WHEN b.invite_intent = 'AION N60' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) AS n60_lead_count,
            SUM(CASE WHEN b.invite_intent = 'AION N60' AND b.follow_cutoff_time IS NOT NULL AND b.is_followed_in_30min THEN 1 ELSE 0 END) AS n60_follow_30min_count,

            COUNT(b.dealer_id) AS lead_count,
            SUM(CASE WHEN b.is_followed_in_30min THEN 1 ELSE 0 END) AS follow_30min_count,
            SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) AS follow_30min_task_count,
            CASE WHEN SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) > 0
                THEN SUM(CASE WHEN b.is_followed_in_30min THEN 1 ELSE 0 END) * 100.0 / SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END)
                ELSE 0 END AS follow_30min_rate,

            COUNT(*) FILTER (
                WHERE b.follow_cutoff_time IS NOT NULL
                  AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                  AND NOT (
                      b.follow_count = 1
                      AND b.lead_status = '跟进中'
                      AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                  )
            ) AS day3_3follow_task_count,

            SUM(CASE WHEN
                b.follow_count IS NOT NULL
                AND b.follow_cutoff_time IS NOT NULL
                AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                AND NOT (
                    b.follow_count = 1
                    AND b.lead_status = '跟进中'
                    AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                )
                AND (
                    (b.follow_count = 1 AND b.lead_status != '跟进中' AND b.is_followed_in_30min)
                    OR
                    (b.follow_count >= 2 AND b.is_followed_in_30min
                     AND b.follow2_time IS NOT NULL AND b.first_follow_time IS NOT NULL
                     AND epoch(b.follow2_time) - epoch(b.first_follow_time) < 259200)
                )
                THEN 1 ELSE 0 END
            ) AS day3_3follow_count,

            CASE WHEN COUNT(*) FILTER (
                WHERE b.follow_cutoff_time IS NOT NULL
                 AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                 AND NOT (
                     b.follow_count = 1
                     AND b.lead_status = '跟进中'
                     AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                 )
            ) > 0
                THEN SUM(CASE WHEN
                    b.follow_count IS NOT NULL
                    AND b.follow_cutoff_time IS NOT NULL
                    AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                    AND NOT (
                        b.follow_count = 1
                        AND b.lead_status = '跟进中'
                        AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                    )
                    AND (
                        (b.follow_count = 1 AND b.lead_status != '跟进中' AND b.is_followed_in_30min)
                        OR
                        (b.follow_count >= 2 AND b.is_followed_in_30min
                         AND b.follow2_time IS NOT NULL AND b.first_follow_time IS NOT NULL
                         AND epoch(b.follow2_time) - epoch(b.first_follow_time) < 259200)
                    )
                    THEN 1 ELSE 0 END
                ) * 100.0 / COUNT(*) FILTER (
                    WHERE b.follow_cutoff_time IS NOT NULL
                     AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                     AND NOT (
                         b.follow_count = 1
                         AND b.lead_status = '跟进中'
                         AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                     )
                )
                ELSE 0 END AS day3_3follow_rate,

            SUM(CASE WHEN b.channel_3 != 'APP-试驾'
                      AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) AS valid_lead_count,
            CASE WHEN COUNT(b.dealer_id) > 0
                THEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) * 100.0 /
                     COUNT(b.dealer_id)
                ELSE 0 END AS valid_lead_rate,

            SUM(CASE WHEN b.channel_3 != 'APP-试驾'
                      AND b.lead_status NOT IN ('异地', '无效')
                      AND b.lead_status != '异地' THEN 1 ELSE 0 END) AS valid_local_lead_count,
            SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) AS local_lead_count,

            COALESCE(sv.visit_count, 0) AS to_shop_count,
            CASE WHEN COUNT(b.dealer_id) > 0 THEN COALESCE(sv.visit_count, 0) * 100.0 / COUNT(b.dealer_id) ELSE 0 END AS lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END)
                ELSE 0 END AS local_lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END)
                ELSE 0 END AS valid_lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END)
                ELSE 0 END AS valid_local_lead_to_shop_rate,
            SUM(CASE WHEN b.follow_cutoff_time IS NOT NULL
                      AND b.channel_1 = '线上'
                      AND b.lead_status NOT IN ('异地', '无效', '未跟进')
                      AND (b.channel_2 = '新媒体-经销店' OR (b.channel_2 = '新媒体' AND b.channel_3 LIKE '%经销商%'))
                 THEN 1 ELSE 0 END) AS new_media_self_valid_lead_count,
            SUM(CASE WHEN b.follow_cutoff_time IS NOT NULL
                      AND b.channel_1 = '线上'
                      AND (b.channel_2 = '新媒体-经销店' OR (b.channel_2 = '新媒体' AND b.channel_3 LIKE '%经销商%'))
                 THEN 1 ELSE 0 END) AS new_media_self_lead_count,
            COALESCE(os.sales_count, 0) AS online_sales_count,
            CASE WHEN SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(os.sales_count, 0) * 100.0 / SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) ELSE 0 END AS online_sales_rate,
            CASE WHEN COALESCE(sv.visit_count, 0) > 0
                THEN COALESCE(os.sales_count, 0) * 100.0 / COALESCE(sv.visit_count, 0) ELSE NULL END AS to_shop_conversion_rate,
            COALESCE(os.sales_count, 0) * 4.0 AS expected_to_shop,
            COALESCE(sv.visit_count, 0) - (COALESCE(os.sales_count, 0) * 4.0) AS to_shop_diff,
            CASE
                WHEN COALESCE(os.sales_count, 0) = 0 THEN '无'
                WHEN COALESCE(sv.visit_count, 0) > 2 * (COALESCE(os.sales_count, 0) * 4.0) THEN '到店转化率低'
                WHEN COALESCE(sv.visit_count, 0) >= 0.6 * (COALESCE(os.sales_count, 0) * 4.0) THEN '正常'
                ELSE '到店录入存在问题'
            END AS to_shop_eval

        FROM base b
        LEFT JOIN shop_visit sv ON b._did = sv.dealer_id
        LEFT JOIN online_sales os ON b._did = os.dealer_id
        GROUP BY b._did, sv.visit_count, os.sales_count
        ORDER BY {sort_expr} {sort_direction} NULLS LAST, b._did ASC
    """
    results = conn.execute(data_sql, dealer_params).fetchall()

    columns = [
        'report_date', 'period_type', 'dealer_id', 'dealer_name',
        'region', 'zone', 'province', 'region_manager', 'zone_manager', 'inspector',
        'n60_lead_count', 'n60_follow_30min_count',
        'lead_count', 'follow_30min_count', 'follow_30min_task_count', 'follow_30min_rate',
        'day3_3follow_task_count', 'day3_3follow_count', 'day3_3follow_rate',
        'valid_lead_count', 'valid_lead_rate',
        'valid_local_lead_count', 'local_lead_count',
        'to_shop_count', 'lead_to_shop_rate',
        'local_lead_to_shop_rate', 'valid_lead_to_shop_rate',
        'valid_local_lead_to_shop_rate',
        'new_media_self_valid_lead_count', 'new_media_self_lead_count',
        'online_sales_count', 'online_sales_rate', 'to_shop_conversion_rate',
        'expected_to_shop', 'to_shop_diff', 'to_shop_eval'
    ]

    data = []
    for row in results:
        data.append(dict(zip(columns, row)))
    _decorate_rows_with_store_status(data)
    summary = _summarize_dealer_report_rows(data)
    data = data[offset:offset + page_size]

    region_zone_filters = _mart_dealer_region_zone_filters(conn)

    return jsonify({
        'success': True,
        'data': data,
        'summary': summary,
        'pagination': {
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size if page_size > 0 else 1
        },
        'filters': region_zone_filters
    })


@app.route('/api/dealer-daily-report/export', methods=['GET'])
def export_dealer_daily_report():
    try:
        import openpyxl

        period = request.args.get('period', 'daily')
        region = request.args.get('region', '')
        zone = request.args.get('zone', '')
        dealer_id = request.args.get('dealer_id', '')
        dealer_name = request.args.get('dealer_name', '')
        sort_by = request.args.get('sort_by', 'lead_count')
        sort_order = request.args.get('sort_order', 'desc')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        store_status = request.args.get('store_status', '')

        conn = duck_db.get_connection()
        has_custom_range = start_date and end_date

        if has_custom_range:
            return _export_dealer_report_custom_range(
                conn, start_date, end_date, region, zone, dealer_id, dealer_name, sort_by, sort_order, store_status
            )
        else:
            return _export_dealer_report_precomputed(
                conn, period, region, zone, dealer_id, dealer_name, sort_by, sort_order, store_status
            )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


def _export_dealer_report_precomputed(conn, period, region, zone, dealer_id, dealer_name, sort_by, sort_order, store_status=""):
    """导出预计算表数据"""
    import openpyxl
    dealer_where = []
    dealer_params = []
    if region:
        dealer_where.append("md.region = ?")
        dealer_params.append(region)
    if zone:
        dealer_where.append("md.zone = ?")
        dealer_params.append(zone)
    if dealer_id:
        dealer_where.append("md.dealer_id LIKE ?")
        dealer_params.append(f"%{dealer_id}%")
    if dealer_name:
        dealer_where.append("md.dealer_name LIKE ?")
        dealer_params.append(f"%{dealer_name}%")
    store_status_ids = _store_status_filter_ids(store_status)
    store_status_filter, store_status_params = _dealer_id_filter_sql("md", store_status_ids)
    if store_status_filter:
        dealer_where.append(store_status_filter.replace(" AND ", "", 1))
        dealer_params.extend(store_status_params)

    dealer_filter = " AND ".join(dealer_where) if dealer_where else "1=1"

    prefix = 'm_' if period == 'monthly' else 'd_'
    sort_expr = _dealer_report_precomputed_sort_expr(sort_by, prefix)
    sort_direction = 'DESC' if sort_order.lower() == 'desc' else 'ASC'

    p_label = '本月' if period == 'monthly' else '本日'

    data_sql = f"""
        SELECT
            md.region, md.zone, md.dealer_id, md.dealer_name, md.province,
            COALESCE(r.{prefix}n60_lead_count, 0) AS n60_lead_count,
            COALESCE(r.{prefix}n60_follow_30min_count, 0) AS n60_follow_30min_count,
            COALESCE(r.{prefix}lead_count, 0) AS lead_count,
            COALESCE(r.{prefix}follow_30min_count, 0) AS follow_30min_count,
            COALESCE(r.{prefix}follow_30min_task_count, 0) AS follow_30min_task_count,
            COALESCE(r.{prefix}follow_30min_rate, 0) AS follow_30min_rate,
            COALESCE(r.{prefix}3day_3follow_task_count, 0) AS day3_3follow_task_count,
            COALESCE(r.{prefix}3day_3follow_count, 0) AS day3_3follow_count,
            COALESCE(r.{prefix}3day_3follow_rate, 0) AS day3_3follow_rate,
            COALESCE(r.{prefix}valid_lead_count, 0) AS valid_lead_count,
            COALESCE(r.{prefix}valid_lead_rate, 0) AS valid_lead_rate,
            COALESCE(r.{prefix}valid_local_lead_count, 0) AS valid_local_lead_count,
            COALESCE(r.{prefix}local_lead_count, 0) AS local_lead_count,
            COALESCE(r.{prefix}to_shop_count, 0) AS to_shop_count,
            COALESCE(r.{prefix}lead_to_shop_rate, 0) AS lead_to_shop_rate,
            COALESCE(r.{prefix}local_lead_to_shop_rate, 0) AS local_lead_to_shop_rate,
            COALESCE(r.{prefix}valid_lead_to_shop_rate, 0) AS valid_lead_to_shop_rate,
            COALESCE(r.{prefix}valid_local_lead_to_shop_rate, 0) AS valid_local_lead_to_shop_rate,
        FROM mart_dealers md
        LEFT JOIN (
            SELECT * FROM (
                SELECT *, ROW_NUMBER() OVER (PARTITION BY dealer_id ORDER BY report_date DESC) AS _rn
                FROM report_dealer_daily
                WHERE period_type = ?
            ) _sub WHERE _sub._rn = 1
        ) r ON md.dealer_id = r.dealer_id
        WHERE {dealer_filter}
        ORDER BY {sort_expr} {sort_direction} NULLS LAST, md.dealer_id ASC
    """
    results = conn.execute(data_sql, [period] + dealer_params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "门店运营日报"

    headers = _xlsx_headers_with_store_status(
        ['大区', '战区', '门店编号', '门店名称', '省份',
         f'{p_label}_N60线索数', f'{p_label}_N60及时跟进数',
         f'{p_label}_线索量', f'{p_label}_30分钟跟进数', f'{p_label}_30分钟跟进任务数', f'{p_label}_30分钟跟进率(%)',
         f'{p_label}_三天三次跟进任务数', f'{p_label}_三天三次跟进数', f'{p_label}_三天三次跟进率(%)',
         f'{p_label}_有效线索量', f'{p_label}_线索有效率(%)',
         f'{p_label}_有效线索量_本地', f'{p_label}_线索量_本地',
         f'{p_label}_到店数', f'{p_label}_线索到店率(%)',
         f'{p_label}_线索到店率_本地(%)', f'{p_label}_有效线索到店率(%)',
         f'{p_label}_有效线索到店率_本地(%)'],
        4,
    )

    ws.append(headers)

    for row in _xlsx_rows_with_store_status(results, dealer_index=2, insert_index=4):
        row_data = [val if val is not None else '' for val in row]
        ws.append(row_data)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"门店运营日报_{p_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _dealer_daily_page_export_headers(p_label):
    """Return custom-range export columns aligned with DealerDailyReport page metrics."""
    return [
        '大区', '战区', '门店编号', '门店名称', '省份',
        f'{p_label}_N60线索数', f'{p_label}_N60及时跟进数',
        f'{p_label}_线索量', f'{p_label}_30分钟跟进数', f'{p_label}_30分钟跟进任务数', f'{p_label}_30分钟跟进率(%)',
        f'{p_label}_三天三次跟进任务数', f'{p_label}_三天三次跟进数', f'{p_label}_三天三次跟进率(%)',
        f'{p_label}_有效线索量', f'{p_label}_线索有效率(%)',
        f'{p_label}_有效线索量_本地', f'{p_label}_线索量_本地',
        f'{p_label}_新媒体自店有效线索量', f'{p_label}_新媒体自店线索量',
        f'{p_label}_到店数', f'{p_label}_线索到店率(%)',
        f'{p_label}_线索到店率_本地(%)', f'{p_label}_有效线索到店率(%)',
        f'{p_label}_有效线索到店率_本地(%)',
        f'{p_label}_线上线索成交数', f'{p_label}_线上线索成交率(%)',
        f'{p_label}_到店成交率(%)', f'{p_label}_到店数预期(25%)',
        f'{p_label}_到店数差异', f'{p_label}_到店数评估',
    ]


def _export_dealer_report_custom_range(conn, start_date, end_date, region, zone, dealer_id, dealer_name, sort_by, sort_order, store_status=""):
    """导出自定义日期范围数据"""
    import openpyxl

    date_where = f"l.assign_date >= DATE '{start_date}' AND l.assign_date <= DATE '{end_date}'"
    visit_date_where = f"visit_date >= DATE '{start_date}' AND visit_date <= DATE '{end_date}'"

    dealer_sql_where = []
    dealer_params = []
    if region:
        dealer_sql_where.append("d.region = ?")
        dealer_params.append(region)
    if zone:
        dealer_sql_where.append("d.zone = ?")
        dealer_params.append(zone)
    if dealer_id:
        dealer_sql_where.append("d.dealer_id LIKE ?")
        dealer_params.append(f"%{dealer_id}%")
    if dealer_name:
        dealer_sql_where.append("d.dealer_name LIKE ?")
        dealer_params.append(f"%{dealer_name}%")
    store_status_ids = _store_status_filter_ids(store_status)
    store_status_filter, store_status_params = _dealer_id_filter_sql("d", store_status_ids)
    if store_status_filter:
        dealer_sql_where.append(store_status_filter.replace(" AND ", "", 1))
        dealer_params.extend(store_status_params)

    dealer_sql_filter = " AND ".join(dealer_sql_where) if dealer_sql_where else "1=1"

    sort_expr = _dealer_report_aggregate_sort_expr(sort_by)
    sort_direction = 'DESC' if sort_order.lower() == 'desc' else 'ASC'

    cutoff_time = f"{end_date} 18:00:00"
    from datetime import datetime as dt, timedelta as td
    cutoff_dt = dt.strptime(cutoff_time, "%Y-%m-%d %H:%M:%S")
    cutoff_time_72h = (cutoff_dt - td(hours=72)).strftime("%Y-%m-%d %H:%M:%S")

    data_sql = f"""
        WITH base AS (
            SELECT
                d.dealer_id AS _did,
                d.dealer_name, d.region, d.zone, d.province,
                d.region_manager, d.zone_manager,
                COALESCE(NULLIF(TRIM(CAST(d.region_manager AS VARCHAR)), ''), '') AS inspector,
                l.*
            FROM mart_dealers d
            LEFT JOIN mart_leads l ON d.dealer_id = l.dealer_id
                AND l.channel_1 = '线上'
                AND {date_where}
            WHERE {dealer_sql_filter}
        ),
        shop_visit AS (
            SELECT dealer_id, SUM(unique_lead_count) AS visit_count
            FROM fact_daily_visit
            WHERE period_type = 'daily'
              AND channel_1 = '线上'
              AND {visit_date_where}
            GROUP BY dealer_id
        ),
        online_sales AS (
            SELECT
                CAST(s.dealer_id AS VARCHAR) AS dealer_id,
                COUNT(*) AS sales_count
            FROM mart_online_sales s
            WHERE s.is_converted = '1'
              AND s.is_counted = '是'
              AND CAST(s.sales_date AS DATE) >= DATE '{start_date}'
              AND CAST(s.sales_date AS DATE) <= DATE '{end_date}'
            GROUP BY dealer_id
        )
        SELECT
            MAX(b.region) AS region,
            MAX(b.zone) AS zone,
            b._did AS dealer_id,
            MAX(b.dealer_name) AS dealer_name,
            MAX(b.province) AS province,
            SUM(CASE WHEN b.invite_intent = 'AION N60' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) AS n60_lead_count,
            SUM(CASE WHEN b.invite_intent = 'AION N60' AND b.follow_cutoff_time IS NOT NULL AND b.is_followed_in_30min THEN 1 ELSE 0 END) AS n60_follow_30min_count,
            COUNT(b.dealer_id) AS lead_count,
            SUM(CASE WHEN b.is_followed_in_30min THEN 1 ELSE 0 END) AS follow_30min_count,
            SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) AS follow_30min_task_count,
            CASE WHEN SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) > 0
                THEN SUM(CASE WHEN b.is_followed_in_30min THEN 1 ELSE 0 END) * 100.0 / SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END)
                ELSE 0 END AS follow_30min_rate,
            COUNT(*) FILTER (
                WHERE b.follow_cutoff_time IS NOT NULL
                  AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                  AND NOT (
                      b.follow_count = 1
                      AND b.lead_status = '跟进中'
                      AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                  )
            ) AS day3_3follow_task_count,
            SUM(CASE WHEN
                b.follow_count IS NOT NULL AND b.follow_cutoff_time IS NOT NULL
                AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                AND NOT (
                    b.follow_count = 1
                    AND b.lead_status = '跟进中'
                    AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                )
                AND ((b.follow_count = 1 AND b.lead_status != '跟进中' AND b.is_followed_in_30min)
                     OR (b.follow_count >= 2 AND b.is_followed_in_30min
                         AND b.follow2_time IS NOT NULL AND b.first_follow_time IS NOT NULL
                         AND epoch(b.follow2_time) - epoch(b.first_follow_time) < 259200))
                THEN 1 ELSE 0 END) AS day3_3follow_count,
            CASE WHEN COUNT(*) FILTER (
                WHERE b.follow_cutoff_time IS NOT NULL
                  AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                  AND NOT (
                      b.follow_count = 1
                      AND b.lead_status = '跟进中'
                      AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                  )
            ) > 0
                THEN SUM(CASE WHEN
                    b.follow_count IS NOT NULL AND b.follow_cutoff_time IS NOT NULL
                    AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                    AND NOT (
                        b.follow_count = 1
                        AND b.lead_status = '跟进中'
                        AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                    )
                    AND ((b.follow_count = 1 AND b.lead_status != '跟进中' AND b.is_followed_in_30min)
                         OR (b.follow_count >= 2 AND b.is_followed_in_30min
                             AND b.follow2_time IS NOT NULL AND b.first_follow_time IS NOT NULL
                             AND epoch(b.follow2_time) - epoch(b.first_follow_time) < 259200))
                    THEN 1 ELSE 0 END) * 100.0 / COUNT(*) FILTER (
                        WHERE b.follow_cutoff_time IS NOT NULL
                          AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                          AND NOT (
                              b.follow_count = 1
                              AND b.lead_status = '跟进中'
                              AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                          ))
                ELSE 0 END AS day3_3follow_rate,
            SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) AS valid_lead_count,
            CASE WHEN COUNT(b.dealer_id) > 0 THEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) * 100.0 / COUNT(b.dealer_id) ELSE 0 END AS valid_lead_rate,
            SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END) AS valid_local_lead_count,
            SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) AS local_lead_count,
            SUM(CASE WHEN b.follow_cutoff_time IS NOT NULL
                      AND b.channel_1 = '线上'
                      AND b.lead_status NOT IN ('异地', '无效', '未跟进')
                      AND (b.channel_2 = '新媒体-经销店' OR (b.channel_2 = '新媒体' AND b.channel_3 LIKE '%经销商%'))
                 THEN 1 ELSE 0 END) AS new_media_self_valid_lead_count,
            SUM(CASE WHEN b.follow_cutoff_time IS NOT NULL
                      AND b.channel_1 = '线上'
                      AND (b.channel_2 = '新媒体-经销店' OR (b.channel_2 = '新媒体' AND b.channel_3 LIKE '%经销商%'))
                 THEN 1 ELSE 0 END) AS new_media_self_lead_count,
            COALESCE(sv.visit_count, 0) AS to_shop_count,
            CASE WHEN COUNT(b.dealer_id) > 0 THEN COALESCE(sv.visit_count, 0) * 100.0 / COUNT(b.dealer_id) ELSE 0 END AS lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) ELSE 0 END AS local_lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) ELSE 0 END AS valid_lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END) ELSE 0 END AS valid_local_lead_to_shop_rate,
            COALESCE(os.sales_count, 0) AS online_sales_count,
            CASE WHEN SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(os.sales_count, 0) * 100.0 / SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) ELSE 0 END AS online_sales_rate,
            CASE WHEN COALESCE(sv.visit_count, 0) > 0
                THEN COALESCE(os.sales_count, 0) * 100.0 / COALESCE(sv.visit_count, 0) ELSE NULL END AS to_shop_conversion_rate,
            COALESCE(os.sales_count, 0) * 4.0 AS expected_to_shop,
            COALESCE(sv.visit_count, 0) - (COALESCE(os.sales_count, 0) * 4.0) AS to_shop_diff,
            CASE
                WHEN COALESCE(os.sales_count, 0) = 0 THEN '无'
                WHEN COALESCE(sv.visit_count, 0) > 2 * (COALESCE(os.sales_count, 0) * 4.0) THEN '到店转化率低'
                WHEN COALESCE(sv.visit_count, 0) >= 0.6 * (COALESCE(os.sales_count, 0) * 4.0) THEN '正常'
                ELSE '到店录入存在问题'
            END AS to_shop_eval
        FROM base b
        LEFT JOIN shop_visit sv ON b._did = sv.dealer_id
        LEFT JOIN online_sales os ON b._did = os.dealer_id
        GROUP BY b._did, sv.visit_count, os.sales_count
        ORDER BY {sort_expr} {sort_direction} NULLS LAST, b._did ASC
    """
    results = conn.execute(data_sql, dealer_params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "门店运营日报"

    p_label = f"{start_date}至{end_date}"
    headers = _xlsx_headers_with_store_status(
        _dealer_daily_page_export_headers(p_label),
        4,
    )

    ws.append(headers)

    for row in _xlsx_rows_with_store_status(results, dealer_index=2, insert_index=4):
        row_data = [val if val is not None else '' for val in row]
        ws.append(row_data)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"门店运营日报_自定义_{start_date}_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _query_dealer_data(conn, start_date, end_date, sort_by='dealer_name', sort_order='asc'):
    date_where = f"l.assign_date >= DATE '{start_date}' AND l.assign_date <= DATE '{end_date}'"
    visit_date_where = f"visit_date >= DATE '{start_date}' AND visit_date <= DATE '{end_date}'"

    valid_sort_columns = [
        'n60_lead_count', 'n60_follow_30min_count',
        'lead_count', 'follow_30min_count', 'follow_30min_rate',
        'day3_3follow_task_count', 'day3_3follow_count', 'day3_3follow_rate',
        'valid_lead_count', 'valid_lead_rate',
        'valid_local_lead_count', 'local_lead_count',
        'to_shop_count', 'lead_to_shop_rate',
        'local_lead_to_shop_rate', 'valid_lead_to_shop_rate',
        'valid_local_lead_to_shop_rate',
        'new_media_self_valid_lead_count', 'new_media_self_lead_count',
        'online_sales_count', 'online_sales_rate', 'to_shop_conversion_rate',
        'expected_to_shop', 'to_shop_diff', 'to_shop_eval',
        'dealer_name', 'region', 'zone'
    ]
    if sort_by not in valid_sort_columns:
        sort_by = 'dealer_name'
    sort_direction = 'DESC' if sort_order.lower() == 'desc' else 'ASC'

    cutoff_time = f"{end_date} 18:00:00"
    from datetime import timedelta as td
    cutoff_dt = datetime.strptime(cutoff_time, "%Y-%m-%d %H:%M:%S")
    cutoff_time_72h = (cutoff_dt - td(hours=72)).strftime("%Y-%m-%d %H:%M:%S")

    data_sql = f"""
        WITH base AS (
            SELECT
                d.dealer_id AS _did,
                d.dealer_name, d.region, d.zone, d.province,
                l.*
            FROM mart_dealers d
            LEFT JOIN mart_leads l ON d.dealer_id = l.dealer_id
                AND l.channel_1 = '线上'
                AND {date_where}
        ),
        shop_visit AS (
            SELECT dealer_id, SUM(unique_lead_count) AS visit_count
            FROM fact_daily_visit
            WHERE period_type = 'daily'
              AND channel_1 = '线上'
              AND {visit_date_where}
            GROUP BY dealer_id
        ),
        online_sales AS (
            SELECT
                CAST(s.dealer_id AS VARCHAR) AS dealer_id,
                COUNT(*) AS sales_count
            FROM mart_online_sales s
            WHERE s.is_converted = '1'
              AND s.is_counted = '是'
              AND CAST(s.sales_date AS DATE) >= DATE '{start_date}'
              AND CAST(s.sales_date AS DATE) <= DATE '{end_date}'
            GROUP BY dealer_id
        )
        SELECT
            MAX(b.region) AS region,
            MAX(b.zone) AS zone,
            b._did AS dealer_id,
            MAX(b.dealer_name) AS dealer_name,
            MAX(b.province) AS province,
            SUM(CASE WHEN b.invite_intent = 'AION N60' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) AS n60_lead_count,
            SUM(CASE WHEN b.invite_intent = 'AION N60' AND b.follow_cutoff_time IS NOT NULL AND b.is_followed_in_30min THEN 1 ELSE 0 END) AS n60_follow_30min_count,
            COUNT(b.dealer_id) AS lead_count,
            SUM(CASE WHEN b.is_followed_in_30min THEN 1 ELSE 0 END) AS follow_30min_count,
            SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) AS follow_30min_task_count,
            CASE WHEN SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END) > 0
                THEN SUM(CASE WHEN b.is_followed_in_30min THEN 1 ELSE 0 END) * 100.0 / SUM(CASE WHEN b.channel_1 = '线上' AND b.follow_cutoff_time IS NOT NULL THEN 1 ELSE 0 END)
                ELSE 0 END AS follow_30min_rate,
            COUNT(*) FILTER (
                WHERE b.follow_cutoff_time IS NOT NULL
                  AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                  AND NOT (
                      b.follow_count = 1
                      AND b.lead_status = '跟进中'
                      AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                  )
            ) AS day3_3follow_task_count,
            SUM(CASE WHEN
                b.follow_count IS NOT NULL AND b.follow_cutoff_time IS NOT NULL
                AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                AND NOT (
                    b.follow_count = 1
                    AND b.lead_status = '跟进中'
                    AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                )
                AND ((b.follow_count = 1 AND b.lead_status != '跟进中' AND b.is_followed_in_30min)
                     OR (b.follow_count >= 2 AND b.is_followed_in_30min
                         AND b.follow2_time IS NOT NULL AND b.first_follow_time IS NOT NULL
                         AND epoch(b.follow2_time) - epoch(b.first_follow_time) < 259200))
                THEN 1 ELSE 0 END) AS day3_3follow_count,
            CASE WHEN COUNT(*) FILTER (
                WHERE b.follow_cutoff_time IS NOT NULL
                  AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                  AND NOT (
                      b.follow_count = 1
                      AND b.lead_status = '跟进中'
                      AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                  )
            ) > 0
                THEN SUM(CASE WHEN
                    b.follow_count IS NOT NULL AND b.follow_cutoff_time IS NOT NULL
                    AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                    AND NOT (
                        b.follow_count = 1
                        AND b.lead_status = '跟进中'
                        AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                    )
                    AND ((b.follow_count = 1 AND b.lead_status != '跟进中' AND b.is_followed_in_30min)
                         OR (b.follow_count >= 2 AND b.is_followed_in_30min
                             AND b.follow2_time IS NOT NULL AND b.first_follow_time IS NOT NULL
                             AND epoch(b.follow2_time) - epoch(b.first_follow_time) < 259200))
                    THEN 1 ELSE 0 END) * 100.0 / COUNT(*) FILTER (
                        WHERE b.follow_cutoff_time IS NOT NULL
                          AND b.raw_assign_time < TRY_CAST('{cutoff_time}' AS TIMESTAMP)
                          AND NOT (
                              b.follow_count = 1
                              AND b.lead_status = '跟进中'
                              AND b.raw_assign_time >= TRY_CAST('{cutoff_time_72h}' AS TIMESTAMP)
                          ))
                ELSE 0 END AS day3_3follow_rate,
            SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) AS valid_lead_count,
            CASE WHEN COUNT(b.dealer_id) > 0 THEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) * 100.0 / COUNT(b.dealer_id) ELSE 0 END AS valid_lead_rate,
            SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END) AS valid_local_lead_count,
            SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) AS local_lead_count,
            SUM(CASE WHEN b.follow_cutoff_time IS NOT NULL
                      AND b.channel_1 = '线上'
                      AND b.lead_status NOT IN ('异地', '无效', '未跟进')
                      AND (b.channel_2 = '新媒体-经销店' OR (b.channel_2 = '新媒体' AND b.channel_3 LIKE '%经销商%'))
                 THEN 1 ELSE 0 END) AS new_media_self_valid_lead_count,
            SUM(CASE WHEN b.follow_cutoff_time IS NOT NULL
                      AND b.channel_1 = '线上'
                      AND (b.channel_2 = '新媒体-经销店' OR (b.channel_2 = '新媒体' AND b.channel_3 LIKE '%经销商%'))
                 THEN 1 ELSE 0 END) AS new_media_self_lead_count,
            COALESCE(sv.visit_count, 0) AS to_shop_count,
            CASE WHEN COUNT(b.dealer_id) > 0 THEN COALESCE(sv.visit_count, 0) * 100.0 / COUNT(b.dealer_id) ELSE 0 END AS lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) ELSE 0 END AS local_lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') THEN 1 ELSE 0 END) ELSE 0 END AS valid_lead_to_shop_rate,
            CASE WHEN SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(sv.visit_count, 0) * 100.0 / SUM(CASE WHEN b.channel_3 != 'APP-试驾' AND b.lead_status NOT IN ('异地', '无效') AND b.lead_status != '异地' THEN 1 ELSE 0 END) ELSE 0 END AS valid_local_lead_to_shop_rate,
            COALESCE(os.sales_count, 0) AS online_sales_count,
            CASE WHEN SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) > 0
                THEN COALESCE(os.sales_count, 0) * 100.0 / SUM(CASE WHEN b.lead_status != '异地' THEN 1 ELSE 0 END) ELSE 0 END AS online_sales_rate,
            CASE WHEN COALESCE(sv.visit_count, 0) > 0
                THEN COALESCE(os.sales_count, 0) * 100.0 / COALESCE(sv.visit_count, 0) ELSE NULL END AS to_shop_conversion_rate,
            COALESCE(os.sales_count, 0) * 4.0 AS expected_to_shop,
            COALESCE(sv.visit_count, 0) - (COALESCE(os.sales_count, 0) * 4.0) AS to_shop_diff,
            CASE
                WHEN COALESCE(os.sales_count, 0) = 0 THEN '无'
                WHEN COALESCE(sv.visit_count, 0) > 2 * (COALESCE(os.sales_count, 0) * 4.0) THEN '到店转化率低'
                WHEN COALESCE(sv.visit_count, 0) >= 0.6 * (COALESCE(os.sales_count, 0) * 4.0) THEN '正常'
                ELSE '到店录入存在问题'
            END AS to_shop_eval
        FROM base b
        LEFT JOIN shop_visit sv ON b._did = sv.dealer_id
        LEFT JOIN online_sales os ON b._did = os.dealer_id
        GROUP BY b._did, sv.visit_count, os.sales_count
        ORDER BY {sort_by} {sort_direction} NULLS LAST, b._did ASC
    """
    return conn.execute(data_sql, []).fetchall()


def _safe_set(ws, row, col, value):
    if value is None:
        ws.cell(row=row, column=col).value = 0
    else:
        ws.cell(row=row, column=col).value = value


_FIELD_INDEX_MAP = {
    'region': 0,
    'zone': 1,
    'dealer_id': 2,
    'dealer_name': 3,
    'province': 4,
    'n60_lead_count': 5,
    'n60_follow_30min_count': 6,
    'lead_count': 7,
    'follow_30min_count': 8,
    'follow_30min_task_count': 9,
    'follow_30min_rate': 10,
    'day3_3follow_task_count': 11,
    'day3_3follow_count': 12,
    'day3_3follow_rate': 13,
    'valid_lead_count': 14,
    'valid_lead_rate': 15,
    'valid_local_lead_count': 16,
    'local_lead_count': 17,
    'new_media_self_valid_lead_count': 18,
    'new_media_self_lead_count': 19,
    'to_shop_count': 20,
    'lead_to_shop_rate': 21,
    'local_lead_to_shop_rate': 22,
    'valid_lead_to_shop_rate': 23,
    'valid_local_lead_to_shop_rate': 24,
    'online_sales_count': 25,
    'online_sales_rate': 26,
    'to_shop_conversion_rate': 27,
    'expected_to_shop': 28,
    'to_shop_diff': 29,
    'to_shop_eval': 30,
}


def _load_field_mapping():
    mapping_path = BASE_DIR / "templates" / "线索运营日报模板-字段映射.csv"
    dealer_mappings = []
    region_mappings = []
    with open(mapping_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            sheet = row['Sheet'].strip()
            col_letter = row['列号'].strip()
            time_dim = row['时间维度'].strip()
            calc_type = row.get('计算类型', '').strip()
            field_expr = row.get('系统字段名/公式', '').strip()
            col_num = sum((ord(c) - ord('A') + 1) * (26 ** (len(col_letter) - i - 1)) for i, c in enumerate(col_letter.upper()))
            entry = {
                'col_num': col_num,
                'time_dim': time_dim,
                'calc_type': calc_type,
                'field_expr': field_expr,
            }
            if sheet == '店端日报':
                dealer_mappings.append(entry)
            elif sheet == '大区日报':
                region_mappings.append(entry)

    return {'dealer': dealer_mappings, 'region': region_mappings}


@app.route('/api/dealer-daily-report/export-template', methods=['GET'])
def export_dealer_daily_report_template():
    try:
        import openpyxl

        date = request.args.get('date', '')
        store_status = request.args.get('store_status', '')

        if not date:
            return jsonify({'success': False, 'message': '请选择日期'}), 400

        month_start = date[:8] + '01'

        conn = duck_db.get_connection()

        monthly_data = _query_dealer_data(conn, month_start, date)
        daily_data = _query_dealer_data(conn, date, date)
        monthly_data = _store_status_filtered_tuple_rows(monthly_data, store_status, dealer_index=2)
        daily_data = _store_status_filtered_tuple_rows(daily_data, store_status, dealer_index=2)

        daily_map = {}
        for row in daily_data:
            daily_map[row[2]] = row

        all_mappings = _load_field_mapping()
        dealer_mappings = all_mappings['dealer']
        region_mappings = all_mappings['region']

        template_path = BASE_DIR / "templates" / "线索运营日报模板.xlsx"
        wb = openpyxl.load_workbook(template_path)

        ws_dealer = wb["店端日报"]
        ws_dealer['A1'] = f"线索运营日报-门店(截止时间{date})"

        for i, mrow in enumerate(monthly_data):
            row_num = 4 + i
            dealer_id = mrow[2]
            drow = daily_map.get(dealer_id)

            for m in dealer_mappings:
                field_idx = _FIELD_INDEX_MAP.get(m['field_expr'])
                if field_idx is None:
                    continue
                if m['time_dim'] == '通用':
                    _safe_set(ws_dealer, row_num, m['col_num'], mrow[field_idx])
                elif m['time_dim'] == '本月':
                    _safe_set(ws_dealer, row_num, m['col_num'], mrow[field_idx])
                elif m['time_dim'] == '本日':
                    if drow:
                        _safe_set(ws_dealer, row_num, m['col_num'], drow[field_idx])
                if m['field_expr'].endswith('_rate'):
                    cell = ws_dealer.cell(row=row_num, column=m['col_num'])
                    if cell.value is not None and cell.value != '':
                        cell.value = round(cell.value / 100, 4)
                    cell.number_format = '0.00%'

        ws_region = wb["大区日报"]
        ws_region['A1'] = f"线索运营日报-大区（截止时间{date}）"

        region_monthly = {}
        region_daily = {}
        region_dealer_ids_monthly = {}
        region_dealer_ids_daily = {}

        for row in monthly_data:
            region_name = row[0]
            dealer_id = row[2]
            if region_name not in region_monthly:
                region_monthly[region_name] = []
                region_dealer_ids_monthly[region_name] = set()
            region_monthly[region_name].append(row)
            region_dealer_ids_monthly[region_name].add(dealer_id)

        for row in daily_data:
            region_name = row[0]
            dealer_id = row[2]
            if region_name not in region_daily:
                region_daily[region_name] = []
                region_dealer_ids_daily[region_name] = set()
            region_daily[region_name].append(row)
            region_dealer_ids_daily[region_name].add(dealer_id)

        all_regions = sorted(region_monthly.keys())

        def _resolve_field_for_region(rows_list, m):
            if m['calc_type'] == '分组键':
                return rows_list[0][0] if rows_list else ''
            if m['calc_type'] == '计数':
                return len(set(r[2] for r in rows_list))

            if m['calc_type'] == '绝对值':
                field_idx = _FIELD_INDEX_MAP.get(m['field_expr'])
                if field_idx is None:
                    return 0
                return sum(r[field_idx] if r[field_idx] is not None else 0 for r in rows_list)

            if m['calc_type'] == '比率':
                parts = m['field_expr'].split('/')
                if len(parts) != 2:
                    return 0
                num_field = parts[0].strip()
                den_field = parts[1].strip()
                num_idx = _FIELD_INDEX_MAP.get(num_field)
                den_idx = _FIELD_INDEX_MAP.get(den_field)
                if num_idx is None or den_idx is None:
                    return 0
                num_sum = sum(r[num_idx] if r[num_idx] is not None else 0 for r in rows_list)
                den_sum = sum(r[den_idx] if r[den_idx] is not None else 0 for r in rows_list)
                if den_sum == 0:
                    return 0
                return round(num_sum / den_sum, 4)

            return 0

        region_row_start = 5
        all_region_data = []

        for r_idx, region_name in enumerate(all_regions):
            row_num = region_row_start + r_idx
            month_rows = region_monthly.get(region_name, [])
            day_rows = region_daily.get(region_name, [])

            row_data = {}
            for m in region_mappings:
                if m['time_dim'] == '本月' or m['time_dim'] == '通用':
                    val = _resolve_field_for_region(month_rows, m)
                else:
                    val = _resolve_field_for_region(day_rows, m)
                row_data[m['col_num']] = val
                _safe_set(ws_region, row_num, m['col_num'], val)
                if m['calc_type'] == '比率':
                    ws_region.cell(row=row_num, column=m['col_num']).number_format = '0.00%'

            all_region_data.append(row_data)

        total_row_num = region_row_start + len(all_regions)

        total_dealer_count = len(set(r[2] for r in monthly_data))

        ws_region.merge_cells(start_row=total_row_num, start_column=1, end_row=total_row_num, end_column=2)
        total_cell = ws_region.cell(row=total_row_num, column=1)
        total_cell.value = f"合计({total_dealer_count}店)"
        total_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        for m in region_mappings:
            if m['calc_type'] in ('分组键', '计数'):
                continue
            col = m['col_num']

            if m['calc_type'] == '绝对值':
                total_val = sum(rd.get(col, 0) for rd in all_region_data)
                _safe_set(ws_region, total_row_num, col, total_val)
            elif m['calc_type'] == '比率':
                parts = m['field_expr'].split('/')
                num_field = parts[0].strip()
                den_field = parts[1].strip()

                num_idx = _FIELD_INDEX_MAP.get(num_field)
                den_idx = _FIELD_INDEX_MAP.get(den_field)

                if num_idx is not None and den_idx is not None:
                    total_num = 0
                    total_den = 0
                    time_dim_for_data = m['time_dim']
                    if time_dim_for_data == '本月' or time_dim_for_data == '通用':
                        source_data = monthly_data
                    else:
                        source_data = daily_data
                    for row in source_data:
                        total_num += row[num_idx] if row[num_idx] is not None else 0
                        total_den += row[den_idx] if row[den_idx] is not None else 0
                    if total_den > 0:
                        val = round(total_num / total_den, 4)
                    else:
                        val = 0
                else:
                    val = 0
                _safe_set(ws_region, total_row_num, col, val)
                ws_region.cell(row=total_row_num, column=col).number_format = '0.00%'

        for col in range(1, ws_region.max_column + 1):
            c = ws_region.cell(row=total_row_num, column=col)
            c.font = openpyxl.styles.Font(name='微软雅黑', size=11, bold=(col <= 2))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"线索运营日报_门店_{date}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/dealer-daily-report/export-custom-range', methods=['GET'])
def export_dealer_daily_report_custom_range():
    try:
        import openpyxl

        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        store_status = request.args.get('store_status', '')

        if not start_date or not end_date:
            return jsonify({'success': False, 'message': '请选择起始日期和结束日期'}), 400

        if start_date > end_date:
            return jsonify({'success': False, 'message': '起始日期不能晚于结束日期'}), 400

        conn = duck_db.get_connection()
        results = _query_dealer_data(conn, start_date, end_date)
        results = _store_status_filtered_tuple_rows(results, store_status, dealer_index=2)

        if not results:
            return jsonify({'success': False, 'message': '该日期范围暂无数据，请调整筛选条件'}), 404

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "门店运营日报"

        p_label = f"{start_date}至{end_date}"
        headers = _xlsx_headers_with_store_status(
            _dealer_daily_page_export_headers(p_label),
            4,
        )

        ws.append(headers)

        for row in _xlsx_rows_with_store_status(results, dealer_index=2, insert_index=4):
            row_data = [val if val is not None else '' for val in row]
            ws.append(row_data)

        rate_cols = [12, 15, 17, 23, 24, 25, 26, 28, 29] if ENABLE_STORE_STATUS_EXPORT else [11, 14, 16, 22, 23, 24, 25, 27, 28]
        for data_row in range(2, len(results) + 2):
            for rc in rate_cols:
                cell = ws.cell(row=data_row, column=rc)
                if cell.value is not None and cell.value != '':
                    cell.value = round(cell.value / 100, 4)
                    cell.number_format = '0.00%'

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"门店运营日报_自定义_{start_date}_{end_date}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':
    init_system(force_refresh=False)
    metadata_registry.initialize()
    port = int(os.getenv("PORT", "5001"))
    print(f"Starting Leads Analytics Server on http://0.0.0.0:{port}")
    app.run(host='0.0.0.0', port=port, debug=False)
